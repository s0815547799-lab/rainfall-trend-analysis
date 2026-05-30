"""
Calibration / Validation Split Analysis
========================================
Period  : 1981-2014
CAL     : 1981-2000  (20 years, 7,305 days)
VAL     : 2001-2014  (14 years, 5,113 days)
Model   : ACCESS-ESM1-5  (1 model demo — extend as needed)

Outputs
-------
  calval_metrics.xlsx          — full metric table per period
  Fig_CalVal_Overview.png/pdf  — 6-panel metric comparison (daily)
  Fig_CalVal_Monthly.png/pdf   — monthly climatology cal vs val
  Fig_CalVal_Scatter.png/pdf   — scatter: cal metrics vs val metrics
"""

import os, math, warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.ticker as ticker
from matplotlib.lines import Line2D
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

warnings.filterwarnings("ignore")

# ── Path configuration ─────────────────────────────────────────────────
# Input files are resolved in order:
#   1. CALVAL_DATA_DIR environment variable  (export CALVAL_DATA_DIR=/path/to/dir)
#   2. data/calval/ subdirectory relative to this script
# Output directory:
#   1. CALVAL_OUT_DIR environment variable   (export CALVAL_OUT_DIR=/path/to/dir)
#   2. results/calval/ relative to this script
_SCRIPT_DIR = Path(__file__).parent
BASE    = Path(os.environ.get("CALVAL_DATA_DIR",
                              str(_SCRIPT_DIR / "data" / "calval")))
OUT_DIR = Path(os.environ.get("CALVAL_OUT_DIR",
                              str(_SCRIPT_DIR / "results" / "calval")))

# ── Style ──────────────────────────────────────────────────────────────
plt.rcParams.update({
    "font.family":     "serif",
    "font.serif":      ["Times New Roman","DejaVu Serif"],
    "font.size":       12,
    "axes.titlesize":  13,"axes.titleweight":"bold",
    "axes.labelsize":  12,"axes.labelweight":"bold",
    "xtick.labelsize": 10,"ytick.labelsize":10,
    "legend.fontsize": 10,
    "lines.linewidth": 2.0,
    "axes.linewidth":  1.3,
    "axes.spines.top": False,"axes.spines.right":False,
    "axes.grid":       True,
    "grid.linestyle":  "-","grid.linewidth":0.4,
    "grid.alpha":      0.4,"grid.color":"#B0BEC5",
    "savefig.dpi":     300,"savefig.bbox":"tight",
    "savefig.pad_inches":0.15,
    "mathtext.fontset":"stix",
    "pdf.fonttype":42,"ps.fonttype":42,
})

C_CAL  = "#1565C0"   # blue  = calibration
C_VAL  = "#C62828"   # red   = validation
C_FULL = "#2E7D32"   # green = full period
C_RAW  = "#FF8F00"   # amber = raw CMIP6
C_OBS  = "#212121"

CAL_LABEL  = "Calibration (1981-2000)"
VAL_LABEL  = "Validation  (2001-2014)"
FULL_LABEL = "Full Period (1981-2014)"
WET_THR    = 1.0
MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]

OBS_FILE   = BASE / "Observed_Rain_daily_198101_201412_28sta.xlsx"
RAW_FILE   = BASE / "pr_day_ACCESS-ESM1-5_historical_r1i1p1f1_gn_19810101-20141231_28sta.xlsx"
BC_FILE    = BASE / "bc_pr_day_ACCESS-ESM1-5_historical_r1i1p1f1_gn_19810101-20141231_28sta.xlsx"
MODEL_NAME = "ACCESS-ESM1-5"


# ── Data loading ────────────────────────────────────────────────────────
def load_df(path):
    df = pd.read_excel(path)
    df["date"] = pd.to_datetime({"year":df["YEAR"],"month":df["MONTH"],"day":df["DAY"]})
    df = df.set_index("date").drop(columns=["YEAR","MONTH","DAY"])
    df = df.where(df >= 0)          # negative → NaN
    df = df.replace([-99,-999,-9999,-9.99e+20,9.99e+20,1e+20], np.nan)
    return df.astype(float)


# ── Metrics ─────────────────────────────────────────────────────────────
def compute_metrics(o, s):
    NULL = {k: np.nan for k in
            ["n","RMSE","MAE","MBE","Pbias","RSR","r","r_sq",
             "NSE","KGE","d","sigma_r","beta","Pbias_abs"]}
    mask = (~np.isnan(o)) & (~np.isnan(s)) & (o >= 0) & (s >= 0)
    o, s = o[mask], s[mask]
    if len(o) < 10:
        return NULL
    e = s - o
    rmse   = float(np.sqrt(np.mean(e**2)))
    mae    = float(np.mean(np.abs(e)))
    mbe    = float(np.mean(e))
    mo, ms = float(np.mean(o)), float(np.mean(s))
    so, ss = float(np.std(o,ddof=1)), float(np.std(s,ddof=1))
    pbias  = 100*np.sum(e)/np.sum(o) if np.sum(o) != 0 else np.nan
    rsr    = rmse/so if so > 0 else np.nan
    r      = float(np.corrcoef(o,s)[0,1]) if so > 0 and ss > 0 else np.nan
    r_sq   = r**2 if not np.isnan(r) else np.nan
    sig_r  = ss/so if so > 0 else np.nan
    beta   = ms/mo if mo != 0 else np.nan
    dn     = np.sum((o-mo)**2)
    nse    = float(1-np.sum(e**2)/dn) if dn > 0 else np.nan
    kge    = float(1-math.sqrt((r-1)**2+(sig_r-1)**2+(beta-1)**2)) \
             if not(np.isnan(r) or np.isnan(sig_r) or np.isnan(beta)) else np.nan
    dd     = np.sum((np.abs(s-mo)+np.abs(o-mo))**2)
    d      = float(1-np.sum(e**2)/dd) if dd > 0 else np.nan
    return dict(n=int(len(o)), RMSE=round(rmse,4), MAE=round(mae,4),
                MBE=round(mbe,4), Pbias=round(float(pbias),2),
                RSR=round(float(rsr),4) if not np.isnan(rsr) else np.nan,
                r=round(r,4) if not np.isnan(r) else np.nan,
                r_sq=round(r_sq,4) if not np.isnan(r_sq) else np.nan,
                NSE=round(float(nse),4) if not np.isnan(nse) else np.nan,
                KGE=round(float(kge),4) if not np.isnan(kge) else np.nan,
                d=round(float(d),4) if not np.isnan(d) else np.nan,
                sigma_r=round(float(sig_r),4) if not np.isnan(sig_r) else np.nan,
                beta=round(float(beta),4) if not np.isnan(beta) else np.nan,
                Pbias_abs=round(abs(float(pbias)),2) if not np.isnan(pbias) else np.nan)

# ── Main ────────────────────────────────────────────────────────────────
def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Loading data...")
    obs = load_df(OBS_FILE)
    raw = load_df(RAW_FILE)
    bc  = load_df(BC_FILE)
    stns = [str(c) for c in obs.columns]
    obs.columns = raw.columns = bc.columns = stns

    # Align common dates
    common = obs.index.intersection(raw.index).intersection(bc.index)
    obs = obs.loc[common]; raw = raw.loc[common]; bc = bc.loc[common]

    cal_mask  = obs.index.year <= 2000
    val_mask  = obs.index.year >  2000
    print(f"  Full : {len(obs):,} days  |  CAL: {cal_mask.sum():,}  |  VAL: {val_mask.sum():,}")
    print(f"  Stations: {len(stns)}")

    # ── Build metrics table ──────────────────────────────────────────────
    print("Computing metrics...")
    records = []
    for stn in stns:
        o = obs[stn].values; r_ = raw[stn].values; b = bc[stn].values
        for label, mask in [("Full (1981-2014)", slice(None)),
                            ("CAL  (1981-2000)", cal_mask),
                            ("VAL  (2001-2014)", val_mask)]:
            om = o[mask]; rm = r_[mask]; bm = b[mask]
            mr = compute_metrics(om, rm)
            mb = compute_metrics(om, bm)
            for ds, mtr in [("Raw",mr),("BC",mb)]:
                row = {"Station":stn, "Period":label, "Dataset":ds}
                row.update(mtr)
                records.append(row)

    met_df = pd.DataFrame(records)

    # Regional means per period/dataset
    reg_rows = []
    for label in met_df["Period"].unique():
        for ds in ["Raw","BC"]:
            sub = met_df[(met_df["Period"]==label) & (met_df["Dataset"]==ds)]
            row = {"Station":"REGIONAL MEAN", "Period":label, "Dataset":ds}
            for col in ["RMSE","MAE","MBE","Pbias","RSR","r","r_sq",
                        "NSE","KGE","d","sigma_r","beta","Pbias_abs"]:
                row[col] = round(float(sub[col].mean()),4) if col in sub else np.nan
            row["n"] = int(sub["n"].mean())
            reg_rows.append(row)

    met_df = pd.concat([met_df, pd.DataFrame(reg_rows)], ignore_index=True)

    # ── Pivot for comparison ─────────────────────────────────────────────────
    METRICS = ["KGE","NSE","r","RMSE","Pbias_abs","RSR","d","MAE"]
    M_LABELS= ["KGE","NSE","r (Pearson)","RMSE (mm/day)","|Pbias| (%)","RSR","d","MAE (mm/day)"]
    LOWER_BETTER = {"RMSE","Pbias_abs","RSR","MAE"}

    # Regional mean BC per period for quick comparison
    def reg_mean(period, ds, met):
        row = met_df[(met_df["Station"]=="REGIONAL MEAN") &
                     (met_df["Period"]==period) &
                     (met_df["Dataset"]==ds)]
        return float(row[met].values[0]) if len(row) else np.nan

    print("\n  Regional mean BC metrics:")
    print(f"  {'Metric':<12} {'Full':>10} {'CAL':>10} {'VAL':>10}  {'CAL-VAL':>10}")
    for m in ["KGE","NSE","r","RMSE","Pbias_abs"]:
        f  = reg_mean("Full (1981-2014)","BC",m)
        c  = reg_mean("CAL  (1981-2000)","BC",m)
        v  = reg_mean("VAL  (2001-2014)","BC",m)
        diff = v-c if not(np.isnan(c) or np.isnan(v)) else np.nan
        print(f"  {m:<12} {f:>10.3f} {c:>10.3f} {v:>10.3f}  {diff:>+10.3f}")

    # ── Figure 1: Overview 6-panel ───────────────────────────────────────────
    print("\nFig 1: Overview comparison...")
    stns_reg = [s for s in stns]   # all stations

    fig, axes = plt.subplots(2, 3, figsize=(18, 11))
    fig.subplots_adjust(hspace=0.42, wspace=0.30,
                        left=0.07, right=0.97, top=0.88, bottom=0.09)

    plot_mets  = ["KGE","NSE","r","RMSE","Pbias_abs","RSR"]
    plot_lbls  = ["KGE","NSE","Pearson r","RMSE (mm/day)","|Pbias| (%)","RSR"]

    for pi, (met, lbl) in enumerate(zip(plot_mets, plot_lbls)):
        ax = axes[pi//3, pi%3]
        tag = chr(97+pi)

        cal_bc   = [float(met_df[(met_df["Station"]==s)&(met_df["Period"]=="CAL  (1981-2000)")&(met_df["Dataset"]=="BC")][met].values[0])
                    if len(met_df[(met_df["Station"]==s)&(met_df["Period"]=="CAL  (1981-2000)")&(met_df["Dataset"]=="BC")]) else np.nan
                    for s in stns_reg]
        val_bc   = [float(met_df[(met_df["Station"]==s)&(met_df["Period"]=="VAL  (2001-2014)")&(met_df["Dataset"]=="BC")][met].values[0])
                    if len(met_df[(met_df["Station"]==s)&(met_df["Period"]=="VAL  (2001-2014)")&(met_df["Dataset"]=="BC")]) else np.nan
                    for s in stns_reg]
        full_raw = [float(met_df[(met_df["Station"]==s)&(met_df["Period"]=="Full (1981-2014)")&(met_df["Dataset"]=="Raw")][met].values[0])
                    if len(met_df[(met_df["Station"]==s)&(met_df["Period"]=="Full (1981-2014)")&(met_df["Dataset"]=="Raw")]) else np.nan
                    for s in stns_reg]

        x = np.arange(len(stns_reg))
        bw = 0.28
        ax.bar(x - bw,   full_raw, width=bw, color=C_RAW, alpha=0.75,
               edgecolor="#BF6000", lw=0.7, label="Raw (Full)",  zorder=3)
        ax.bar(x,        cal_bc,  width=bw, color=C_CAL, alpha=0.85,
               edgecolor="#0D47A1", lw=0.7, label=CAL_LABEL, zorder=3)
        ax.bar(x + bw,   val_bc,  width=bw, color=C_VAL, alpha=0.85,
               edgecolor="#B71C1C", lw=0.7, label=VAL_LABEL, zorder=3)
        ax.axhline(0, color="#546E7A", lw=0.8, ls="-")

        ax.set_xticks(x)
        ax.set_xticklabels([f"S{i+1}" for i in range(len(stns_reg))],
                           rotation=90, fontsize=7)
        ax.set_ylabel(lbl, fontsize=11)
        ax.set_title(f"({tag}) {lbl}", loc="left", fontsize=12, fontweight="bold")
        ax.yaxis.set_minor_locator(ticker.AutoMinorLocator())

        if pi == 0:
            ax.legend(fontsize=9, frameon=True, edgecolor="#B0BEC5",
                      facecolor="white", framealpha=0.93, ncol=1, loc="lower right")

    fig.suptitle(
        f"Calibration / Validation Split — Daily Scale  |  {MODEL_NAME}\n"
        f"CAL: 1981–2000 (20 yr, 7,305 days)    VAL: 2001–2014 (14 yr, 5,113 days)    Stations: {len(stns)}",
        fontsize=13, fontweight="bold"
    )
    fig.patch.set_facecolor("white")
    fig.savefig(str(OUT_DIR/"Fig_CalVal_Overview.png"), dpi=300, bbox_inches="tight")
    fig.savefig(str(OUT_DIR/"Fig_CalVal_Overview.pdf"), bbox_inches="tight")
    plt.close(fig)
    print("  → Fig_CalVal_Overview saved")

    # ── Figure 2: Cal vs Val scatter ─────────────────────────────────────────
    print("Fig 2: Cal vs Val scatter...")
    fig, axes = plt.subplots(1, 3, figsize=(15, 5.5))
    fig.subplots_adjust(wspace=0.32, left=0.07, right=0.97, top=0.86, bottom=0.12)

    scatter_mets = ["KGE","NSE","r"]
    scatter_lbls = ["KGE","NSE","Pearson r"]

    for pi, (met, lbl) in enumerate(zip(scatter_mets, scatter_lbls)):
        ax = axes[pi]
        cv = [(float(met_df[(met_df["Station"]==s)&(met_df["Period"]=="CAL  (1981-2000)")&(met_df["Dataset"]=="BC")][met].values[0]),
               float(met_df[(met_df["Station"]==s)&(met_df["Period"]=="VAL  (2001-2014)")&(met_df["Dataset"]=="BC")][met].values[0]))
              for s in stns_reg
              if len(met_df[(met_df["Station"]==s)&(met_df["Period"]=="CAL  (1981-2000)")&(met_df["Dataset"]=="BC")]) > 0]
        cal_v = [v[0] for v in cv]; val_v = [v[1] for v in cv]

        vmin = min(min(cal_v), min(val_v)) - 0.05
        vmax = max(max(cal_v), max(val_v)) + 0.05
        ax.plot([vmin, vmax], [vmin, vmax], color=C_FULL, lw=1.4, ls="--",
                alpha=0.7, label="1:1 line", zorder=2)
        sc = ax.scatter(cal_v, val_v, c=range(len(cal_v)),
                        cmap="tab20", s=80, edgecolors="white",
                        linewidths=0.6, zorder=4, alpha=0.9)
        for i, (cx, vy) in enumerate(zip(cal_v, val_v)):
            ax.text(cx+0.005, vy+0.005, f"S{i+1}", fontsize=7,
                    color="#37474F", alpha=0.8)

        # regression line
        z = np.polyfit([v for v in cal_v if not np.isnan(v)],
                       [v for v in val_v if not np.isnan(v)], 1)
        xr = np.array([vmin, vmax])
        ax.plot(xr, np.polyval(z, xr), color="#7B1FA2", lw=1.5,
                ls="-", alpha=0.7, label=f"Trend (slope={z[0]:.2f})")

        ax.set_xlim(vmin, vmax); ax.set_ylim(vmin, vmax)
        ax.set_xlabel(f"{lbl} — Calibration", fontsize=11)
        ax.set_ylabel(f"{lbl} — Validation", fontsize=11)
        ax.set_title(f"({chr(97+pi)}) {lbl}: CAL vs VAL",
                     loc="left", fontsize=12, fontweight="bold")
        ax.legend(fontsize=9, loc="upper left")
        ax.set_aspect("equal", "box")

    fig.suptitle(
        f"Calibration vs Validation Scatter — BC (QDM) Performance  |  {MODEL_NAME}\n"
        "Points above 1:1 line = VAL better than CAL; below = degradation in validation",
        fontsize=12, fontweight="bold"
    )
    fig.patch.set_facecolor("white")
    fig.savefig(str(OUT_DIR/"Fig_CalVal_Scatter.png"), dpi=300, bbox_inches="tight")
    fig.savefig(str(OUT_DIR/"Fig_CalVal_Scatter.pdf"), bbox_inches="tight")
    plt.close(fig)
    print("  → Fig_CalVal_Scatter saved")

    # ── Figure 3: Monthly climatology cal vs val ─────────────────────────────
    print("Fig 3: Monthly climatology...")

    def monthly_clim(df_daily, mask):
        sub = df_daily.loc[mask]
        mon = sub.resample("MS").apply(lambda g: g.sum(min_count=int(0.8*len(g))))
        reg = mon.mean(axis=1)
        clim = reg.groupby(reg.index.month).agg(["mean","std"]).reindex(range(1,13))
        return clim["mean"].values, clim["std"].values

    months = np.arange(1, 13)

    obs_cal_m,  obs_cal_s  = monthly_clim(obs, cal_mask)
    obs_val_m,  obs_val_s  = monthly_clim(obs, val_mask)
    raw_cal_m,  raw_cal_s  = monthly_clim(raw, cal_mask)
    raw_val_m,  raw_val_s  = monthly_clim(raw, val_mask)
    bc_cal_m,   bc_cal_s   = monthly_clim(bc,  cal_mask)
    bc_val_m,   bc_val_s   = monthly_clim(bc,  val_mask)

    fig, axes = plt.subplots(1, 2, figsize=(16, 6.5))
    fig.subplots_adjust(wspace=0.28, left=0.07, right=0.97, top=0.86, bottom=0.11)

    for pi, (period_lbl, obs_m, obs_s, raw_m, raw_s, bc_m, bc_s) in enumerate([
        ("(a) Calibration 1981–2000", obs_cal_m, obs_cal_s, raw_cal_m, raw_cal_s, bc_cal_m, bc_cal_s),
        ("(b) Validation  2001–2014", obs_val_m, obs_val_s, raw_val_m, raw_val_s, bc_val_m, bc_val_s),
    ]):
        ax = axes[pi]
        col_obs = C_OBS
        col_raw = C_RAW
        col_bc  = C_CAL if pi==0 else C_VAL

        ax.fill_between(months, obs_m-obs_s, obs_m+obs_s, color="#B0BEC5", alpha=0.3, zorder=1)
        ax.plot(months, obs_m, color=col_obs, lw=2.5, marker="o", ms=5,
                label="Observed", zorder=5)
        ax.fill_between(months, raw_m-raw_s, raw_m+raw_s, color="#FFCCBC", alpha=0.3, zorder=1)
        ax.plot(months, raw_m, color=col_raw, lw=2.0, marker="^", ms=5,
                label="Raw CMIP6", zorder=4)
        ax.fill_between(months, bc_m-bc_s, bc_m+bc_s,
                        color=C_CAL+"44" if pi==0 else C_VAL+"44", alpha=0.3, zorder=1)
        ax.plot(months, bc_m, color=col_bc, lw=2.2, marker="s", ms=5,
                label="BC (QDM)", zorder=4)

        ax.set_xticks(months)
        ax.set_xticklabels(MONTH_ABBR, fontsize=10)
        ax.set_ylabel("Mean Monthly Rainfall — Regional Mean (mm)", fontsize=11)
        ax.set_title(period_lbl, loc="left", fontsize=13, fontweight="bold")
        ax.legend(fontsize=10, frameon=True, edgecolor="#B0BEC5",
                  facecolor="white", framealpha=0.93, ncol=1)
        ax.yaxis.set_minor_locator(ticker.AutoMinorLocator())

    fig.suptitle(
        f"Monthly Climatology — Observed vs Raw vs BC (QDM)  |  {MODEL_NAME}  |  Regional Mean (28 stations)\n"
        "Shaded band = ±1 standard deviation",
        fontsize=12, fontweight="bold"
    )
    fig.patch.set_facecolor("white")
    fig.savefig(str(OUT_DIR/"Fig_CalVal_Monthly.png"), dpi=300, bbox_inches="tight")
    fig.savefig(str(OUT_DIR/"Fig_CalVal_Monthly.pdf"), bbox_inches="tight")
    plt.close(fig)
    print("  → Fig_CalVal_Monthly saved")

    # ── Figure 4: Degradation heatmap ───────────────────────────────────────
    print("Fig 4: Cal-Val difference heatmap...")

    higher_better = {"KGE","NSE","r","d"}
    mets_heat = ["KGE","NSE","r","RMSE","Pbias_abs","d"]
    mlbls_heat= ["KGE","NSE","r","RMSE","Pbias_abs","d"]

    diff_mat = np.full((len(mets_heat), len(stns_reg)), np.nan)
    for si, stn in enumerate(stns_reg):
        for mi, met in enumerate(mets_heat):
            cv = met_df[(met_df["Station"]==stn)&(met_df["Period"]=="CAL  (1981-2000)")&(met_df["Dataset"]=="BC")][met]
            vv = met_df[(met_df["Station"]==stn)&(met_df["Period"]=="VAL  (2001-2014)")&(met_df["Dataset"]=="BC")][met]
            if len(cv) and len(vv):
                c = float(cv.values[0]); v = float(vv.values[0])
                if met in higher_better:
                    diff_mat[mi, si] = v - c    # positive = val better
                else:
                    diff_mat[mi, si] = c - v    # positive = val better (lower error)

    fig, ax = plt.subplots(figsize=(max(14, len(stns_reg)*0.55 + 3), 4.5))
    fig.subplots_adjust(left=0.10, right=0.97, top=0.82, bottom=0.18)

    amx = np.nanmax(np.abs(diff_mat)) + 0.01
    im = ax.imshow(diff_mat, cmap="RdYlGn", vmin=-amx, vmax=amx,
                   aspect="auto", interpolation="nearest")
    for mi in range(len(mets_heat)):
        for si in range(len(stns_reg)):
            v = diff_mat[mi, si]
            if not np.isnan(v):
                mid = 0; rng = amx*2 + 1e-9
                tc = "white" if abs(v)/amx > 0.65 else "black"
                ax.text(si, mi, f"{v:+.2f}", ha="center", va="center",
                        fontsize=7, fontweight="bold", color=tc)

    ax.set_xticks(range(len(stns_reg)))
    ax.set_xticklabels([f"S{i+1}" for i in range(len(stns_reg))], fontsize=9)
    ax.set_yticks(range(len(mets_heat)))
    ax.set_yticklabels(mlbls_heat, fontsize=11)
    ax.set_xlabel("Station", fontsize=11, fontweight="bold")
    cb = plt.colorbar(im, ax=ax, orientation="vertical", pad=0.02, fraction=0.025)
    cb.set_label("VAL − CAL  (positive = VAL better)", fontsize=10)
    ax.set_title(
        "Calibration → Validation Performance Change — BC (QDM)  |  Green = improvement maintained / better in VAL  |  Red = degradation in VAL",
        fontsize=11, fontweight="bold", pad=8
    )
    fig.suptitle(f"{MODEL_NAME}  |  Daily Scale", fontsize=12, fontweight="bold")
    fig.patch.set_facecolor("white")
    fig.savefig(str(OUT_DIR/"Fig_CalVal_Heatmap.png"), dpi=300, bbox_inches="tight")
    fig.savefig(str(OUT_DIR/"Fig_CalVal_Heatmap.pdf"), bbox_inches="tight")
    plt.close(fig)
    print("  → Fig_CalVal_Heatmap saved")

    # ── Excel output ─────────────────────────────────────────────────────────
    print("\nBuilding Excel...")

    THIN = Side(style="thin", color="BDBDBD")
    MED  = Side(style="medium", color="1F4E79")
    def _tb(): return Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
    def _xf(h): return PatternFill("solid",fgColor=h)

    def xsc(ws,r,c,val=None,bold=False,fc=None,bg=None,align="center",sz=10,wrap=True):
        cell=ws.cell(row=r,column=c)
        if val is not None: cell.value=val
        cell.font=Font(bold=bold,name="Calibri",size=sz,color=fc if fc else "1A1A1A")
        cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
        if bg: cell.fill=_xf(bg)
        cell.border=_tb()
        return cell

    def mxsc(ws,r,c1,c2,val,**kw):
        ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
        return xsc(ws,r,c1,val,**kw)

    wb = Workbook(); wb.remove(wb.active)

    # Sheet 1 — Summary table
    ws1 = wb.create_sheet("CalVal Summary")
    ws1.freeze_panes = "D3"

    mxsc(ws1,1,1,20,
         f"Calibration / Validation Split — {MODEL_NAME}  |  CAL: 1981-2000  |  VAL: 2001-2014  |  28 Stations  |  Daily Scale",
         bold=True, fc="FFFFFF", bg="13293D", sz=12)

    headers = ["Station","Period","Dataset","n","KGE","NSE","r","RMSE","MAE",
               "MBE","Pbias","Pbias_abs","RSR","d","sigma_r","beta"]
    hdr_bg = {"Raw":"FFEBEE","BC":"E3F2FD"}
    for ci, h in enumerate(headers, 1):
        xsc(ws1,2,ci,h,bold=True,fc="FFFFFF",bg="2E75B6",sz=10)

    row_i = 3
    period_bg = {
        "Full (1981-2014)": "F1F8E9",
        "CAL  (1981-2000)": "E3F2FD",
        "VAL  (2001-2014)": "FFEBEE",
    }
    for _, row in met_df.iterrows():
        pbg = period_bg.get(row["Period"],"FFFFFF")
        ds_bg = "FFEBEE" if row["Dataset"]=="Raw" else "E3F2FD"
        bg = pbg
        bold_row = row["Station"] == "REGIONAL MEAN"
        for ci, h in enumerate(headers, 1):
            v = row.get(h, "")
            if isinstance(v, float) and np.isnan(v): v = "-"
            elif isinstance(v, float): v = round(v, 3)
            xsc(ws1, row_i, ci, v, bold=bold_row, bg=bg, sz=9)
        row_i += 1

    for ci, w in enumerate([14,22,8,8,8,8,8,10,10,10,10,10,8,8,8,8], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # Sheet 2 — Regional mean comparison
    ws2 = wb.create_sheet("Regional Mean Comparison")
    mxsc(ws2,1,1,10,
         "Regional Mean — CAL vs VAL Comparison (BC only)",
         bold=True,fc="FFFFFF",bg="13293D",sz=12)
    hdrs2 = ["Metric","Full (1981-2014)","CAL (1981-2000)","VAL (2001-2014)",
             "CAL-VAL Diff","Better in VAL?"]
    for ci,h in enumerate(hdrs2,1):
        xsc(ws2,2,ci,h,bold=True,fc="FFFFFF",bg="2E75B6",sz=10)

    ri = 3
    for met in ["KGE","NSE","r","RMSE","MAE","Pbias_abs","RSR","d"]:
        f = reg_mean("Full (1981-2014)","BC",met)
        c = reg_mean("CAL  (1981-2000)","BC",met)
        v = reg_mean("VAL  (2001-2014)","BC",met)
        diff = round(v-c, 3) if not(np.isnan(c) or np.isnan(v)) else "-"
        if met in LOWER_BETTER:
            better = "Yes ✓" if isinstance(diff,float) and diff < 0 else ("No ✗" if isinstance(diff,float) else "-")
            diff_disp = f"{diff:+.3f}" if isinstance(diff,float) else diff
        else:
            better = "Yes ✓" if isinstance(diff,float) and diff > 0 else ("No ✗" if isinstance(diff,float) else "-")
            diff_disp = f"{diff:+.3f}" if isinstance(diff,float) else diff
        bg_better = "C8E6C9" if better=="Yes ✓" else ("FFCDD2" if better=="No ✗" else "FFFFFF")
        for ci, val in enumerate([met,
                                   f"{f:.3f}" if not np.isnan(f) else "-",
                                   f"{c:.3f}" if not np.isnan(c) else "-",
                                   f"{v:.3f}" if not np.isnan(v) else "-",
                                   diff_disp, better], 1):
            bg = "FFF9C4" if ci==5 else (bg_better if ci==6 else "FFFFFF")
            xsc(ws2,ri,ci,val,bold=(ci==6),bg=bg,sz=10)
        ri += 1

    for ci,w in enumerate([14,18,18,18,16,14],1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # Sheet 3 — Interpretation
    ws3 = wb.create_sheet("Interpretation")
    mxsc(ws3,1,1,2,"Cal/Val Split Interpretation Guide",
         bold=True,fc="FFFFFF",bg="13293D",sz=12)
    notes = [
        ("What is Cal/Val split?",
         "The bias correction (QDM) was calibrated on 1981-2000. "
         "We then check whether the corrected model still performs well "
         "on the independent 2001-2014 period (validation). "
         "If performance is similar, the method generalises well (no overfitting)."),
        ("How to interpret?",
         "Compare KGE/NSE/r for CAL vs VAL. "
         "A drop of < 0.05 in KGE from CAL to VAL is generally acceptable. "
         "Larger drops may indicate that the QDM transfer function was tuned too tightly."),
        ("Green cells (Better in VAL)",
         "The validation period shows equal or better performance — the correction generalises well."),
        ("Red cells (No ✗)",
         "Performance drops in the validation period. "
         "If the drop is small (< 0.05 KGE), it is within normal statistical variation. "
         "Larger drops warrant investigation (e.g. non-stationarity in extremes)."),
        ("Reference for Cal/Val practice",
         "Maraun D et al. (2017) Bias correcting climate change simulations - "
         "a critical review. Curr. Clim. Change Rep. 3:211-226."),
        ("Recommended reporting in paper",
         "State calibration period (1981-2000) in Methods. "
         "Report CAL and VAL metrics separately in Supplementary or Table. "
         "Discuss any degradation > 0.05 KGE in Discussion section."),
    ]
    for i,(title,body) in enumerate(notes,3):
        xsc(ws3,i,1,title,bold=True,bg="E3F2FD",sz=10,align="left")
        xsc(ws3,i,2,body,sz=10,align="left")
        ws3.row_dimensions[i].height = 48
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 90

    xl_path = OUT_DIR / "CalVal_Analysis_ACCESS-ESM1-5.xlsx"
    wb.save(str(xl_path))
    print(f"  → Excel saved: {xl_path.name}")

    print("\n" + "="*60)
    print("  CALIBRATION / VALIDATION ANALYSIS COMPLETE")
    print(f"  Files saved to {OUT_DIR}/")
    print("="*60)
    print(f"  Fig_CalVal_Overview.png/pdf")
    print(f"  Fig_CalVal_Scatter.png/pdf")
    print(f"  Fig_CalVal_Monthly.png/pdf")
    print(f"  Fig_CalVal_Heatmap.png/pdf")
    print(f"  CalVal_Analysis_ACCESS-ESM1-5.xlsx")


if __name__ == "__main__":
    main()
