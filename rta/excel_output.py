"""
rta.excel_output — Excel workbook output for rainfall trend analysis.

Extracted and extended from rainfall_trend_analysis_v3.py §16 (lines 1500–1770).
Sheets S1–S6 are preserved exactly from v3.
New sheets S7 (4-Method Comparison), S8 (Field Significance), and
S9 (Dry Season Validation) are appended when the corresponding data is supplied.
"""

import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config import (XC, THIN, MED, SCALE_META, WET_THR, Z_005, Z_001,
                     tb, xfill, xsc, mxsc, cw, rh)


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  write_excel — 9 sheets (S1–S6 exact v3; S7–S9 new)                    ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def write_excel(out_xlsx, stns, smap, trend_df, comp_df,
                desc_df, qc_dict, period,
                comp4_df=None, field_sig_df=None, dry_validation=None):
    """
    Write Excel workbook with up to 9 sheets:

      S1 — Standard MK Results
      S2 — Modified MK Results (Hamed & Rao 1998)
      S3 — MK vs MMK Comparison
      S4 — Sen's Slope Summary (with 95 % CI)
      S5 — Descriptive Statistics
      S6 — Methods & References
      S7 — 4-Method Comparison       (optional; requires comp4_df)
      S8 — Field Significance         (optional; requires field_sig_df)
      S9 — Dry Season Validation      (optional; requires dry_validation)

    Parameters
    ----------
    out_xlsx : str or Path
    stns     : list of station IDs
    smap     : dict  {station_id → short_code}
    trend_df : DataFrame  (Standard MK + Modified MK rows)
    comp_df  : DataFrame  (MK vs MMK comparison)
    desc_df  : DataFrame  (descriptive stats, indexed by station)
    qc_dict  : dict   (unused in sheet writing but kept for API compat.)
    period   : str    e.g. "1981–2020"
    comp4_df : DataFrame or None  — 4-method comparison (S7)
    field_sig_df : DataFrame or None — field significance (S8)
    dry_validation : dict or None  — dry season validation (S9)
    """
    stns = [str(s) for s in stns]
    wb   = Workbook()
    wb.remove(wb.active)

    # ── shared inner helpers ──────────────────────────────────────────────────

    def _title(ws, nc, t1, t2=""):
        mxsc(ws, 1, 1, nc, t1, bold=True, fc="FFFFFF", bg=XC["title"], sz=12,
             align="left")
        rh(ws, 1, 24)
        if t2:
            mxsc(ws, 2, 1, nc, t2, italic=True, fc="FFFFFF", bg=XC["sub"],
                 sz=9)
            rh(ws, 2, 14)

    def _hdr(ws, r, hdrs, bg=XC["hdr"]):
        for ci, h in enumerate(hdrs, 1):
            xsc(ws, r, ci, h, bold=True, fc="FFFFFF", bg=bg, border=tb(),
                sz=9, wrap=True)
        rh(ws, r, 40)

    sc_bg = {
        "annual": XC["ann_h"],
        "wet":    XC["wet_h"],
        "dry":    XC["dry_h"],
    }

    def _write_trend_sheet(ws, method_filter, t1, t2):
        sub_df = (trend_df[trend_df["Method"] == method_filter]
                  .reset_index(drop=True))
        is_mmk = "Modified" in method_filter
        nc = 15 if is_mmk else 13
        _title(ws, nc, t1, t2)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "E4"
        if is_mmk:
            hdr = ["Station", "Code", "Scale", "N", "S", "Var(S)", "Var*(S)",
                   "n_eff", "ρ₁", "Z", "τ (Kendall)", "p-value", "Trend",
                   "* p<0.05", "** p<0.01"]
        else:
            hdr = ["Station", "Code", "Scale", "N", "S", "Var(S)",
                   "Z", "τ (Kendall)", "p-value", "Trend",
                   "* p<0.05", "** p<0.01", "rho_1"]
        _hdr(ws, 3, hdr)
        ri = 4
        for _, row in sub_df.iterrows():
            sk  = row["Scale"]
            s05 = bool(row["sig_05"])
            s01 = bool(row["sig_01"])
            bg  = sc_bg.get(sk, XC["ann_h"])
            if s01:
                bg = XC["sig01"]
            elif s05:
                bg = XC["sig05"]
            if is_mmk:
                vals = [str(row["Station"]), str(row["Code"]),
                        row["Scale_Label"],
                        row["N"], row["S"], row["Var_S"], row["Var_S_adj"],
                        row["n_eff"], row["rho_1"], row["Z"], row["tau"],
                        row["p_value"],
                        str(row["Trend"]),
                        "**" if s01 else ("*" if s05 else "ns"),
                        "**" if s01 else "ns"]
            else:
                vals = [str(row["Station"]), str(row["Code"]),
                        row["Scale_Label"],
                        row["N"], row["S"], row["Var_S"],
                        row["Z"], row["tau"], row["p_value"],
                        str(row["Trend"]),
                        "**" if s01 else ("*" if s05 else "ns"),
                        "**" if s01 else "ns",
                        row["rho_1"]]
            for ci, v in enumerate(vals, 1):
                if isinstance(v, float) and np.isnan(v):
                    v = "—"
                elif isinstance(v, float):
                    v = round(v, 4)
                cell = xsc(ws, ri, ci, v, bg=bg, border=tb(), sz=9,
                           align="left" if ci <= 3 else "right")
                if ci == len(vals) - 2 and "Increasing" in str(v):
                    cell.font = Font(bold=True, color="1B5E20",
                                     name="Calibri", size=9)
                if ci == len(vals) - 2 and "Decreasing" in str(v):
                    cell.font = Font(bold=True, color="B71C1C",
                                     name="Calibri", size=9)
            rh(ws, ri, 15)
            ri += 1
        for ci, w in enumerate(
                [10, 8, 18] + ([7] * 2) + ([11] * (nc - 5)) + [14, 9, 9], 1):
            cw(ws, ci, w)

    # ── S1: Standard MK ───────────────────────────────────────────────────────
    ws1 = wb.create_sheet("S1 Standard MK")
    _write_trend_sheet(
        ws1, "Standard MK",
        f"Standard Mann–Kendall Test Results  |  {period}",
        "Mann (1945) / Kendall (1975)  |  NO autocorrelation correction  |"
        "  * p<0.05  ** p<0.01  |  Two-tailed test")

    # ── S2: Modified MK ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("S2 Modified MK (H&R98)")
    _write_trend_sheet(
        ws2, "Modified MK",
        f"Modified Mann–Kendall Test Results (Hamed & Rao 1998)  |  {period}",
        "Autocorrelation-corrected  |  n* = effective sample size  |"
        "  Var*(S) = Var(S)×(n/n*)  |  * p<0.05  ** p<0.01")

    # ── S3: MK vs MMK Comparison ──────────────────────────────────────────────
    ws3 = wb.create_sheet("S3 MK vs MMK Comparison")
    ws3.sheet_view.showGridLines = False
    nc3 = 17
    _title(ws3, nc3,
           f"Comparison: Standard MK vs Modified MK (Hamed & Rao 1998)  |  {period}",
           "ΔZ = Z_MMK − Z_MK  (negative → autocorr. reduces |Z|)  |"
           "  Agree = both methods reach same trend conclusion  |"
           "  Red Agree=False → autocorrelation changes trend decision")
    hdr3 = ["Station", "Code", "Scale", "ρ₁", "Sig.AC",
             "MK Z", "MK p", "MK Trend", "MK *",
             "MMK Z", "MMK p", "MMK Trend", "MMK *",
             "ΔZ", "Δp", "Agree", "Note"]
    _hdr(ws3, 3, hdr3)
    ri3 = 4
    for _, row in comp_df.iterrows():
        sk     = row["Scale"]
        agree  = bool(row["Agree"])
        sig_ac = bool(row["Sig_AC"])
        bg     = sc_bg.get(sk, XC["ann_h"])
        if not agree:
            bg = "FFE0E0"
        elif sig_ac:
            bg = "FFFDE7"
        vals = [str(row["Station"]), str(row.get("Code", "")),
                SCALE_META.get(sk, {}).get("label", sk),
                row["rho_1"],
                "Yes*" if sig_ac else "No",
                row["MK_Z"], row["MK_p"], str(row["MK_Trend"]),
                "**" if row.get("MK_sig05", False) else "ns",
                row["MMK_Z"], row["MMK_p"], str(row["MMK_Trend"]),
                "**" if row.get("MMK_sig05", False) else "ns",
                row["delta_Z"], row["delta_p"],
                "Yes" if agree else "No",
                ("AC changed conclusion" if not agree
                 else ("AC corrected" if sig_ac else ""))]
        for ci, v in enumerate(vals, 1):
            if isinstance(v, float) and np.isnan(v):
                v = "—"
            elif isinstance(v, float):
                v = round(v, 4)
            bg_cell = bg
            cell = xsc(ws3, ri3, ci, v, bg=bg_cell, border=tb(), sz=9,
                       align="left" if ci in (1, 2, 3, 8, 12, 17) else "right")
            if ci == 16 and v == "No":
                cell.fill = xfill("FFCCBC")
                cell.font = Font(bold=True, color="B71C1C",
                                  name="Calibri", size=9)
            if ci == 5 and "Yes" in str(v):
                cell.font = Font(bold=True, color="E65100",
                                  name="Calibri", size=9)
        rh(ws3, ri3, 15)
        ri3 += 1
    for ci, w in enumerate(
            [10, 8, 18, 8, 7] + [10, 10, 18, 5] * 2 + [9, 9, 5, 30], 1):
        cw(ws3, ci, w)

    # ── S4: Sen's Slope ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("S4 Sens Slope")
    ws4.sheet_view.showGridLines = False
    _title(ws4, 11,
           f"Sen's Slope Estimator + 95% CI  |  {period}",
           "Sen (1968) JASA 63:1379  |  95% CI: Gilbert (1987) rank-based  |"
           "  β = Sen's slope (mm yr⁻¹)  |  Positive β = increasing rainfall")
    hdr4 = ["Station", "Code", "Scale", "Method", "N", "β (mm/yr)",
            "CI_Lower (mm/yr)", "CI_Upper (mm/yr)", "Z", "p-value", "Trend"]
    _hdr(ws4, 3, hdr4)
    ri4 = 4
    for _, row in trend_df.iterrows():
        sk  = row["Scale"]
        s05 = bool(row["sig_05"])
        s01 = bool(row["sig_01"])
        bg  = sc_bg.get(sk, XC["ann_h"])
        if s01:
            bg = XC["sig01"]
        elif s05:
            bg = XC["sig05"]
        vals = [str(row["Station"]), str(row["Code"]), row["Scale_Label"],
                row["Method"], row["N"], row["Slope_Q"],
                row["Slope_lo"], row["Slope_hi"],
                row["Z"], row["p_value"], str(row["Trend"])]
        for ci, v in enumerate(vals, 1):
            if isinstance(v, float) and np.isnan(v):
                v = "—"
            elif isinstance(v, float):
                v = round(v, 3)
            cell = xsc(ws4, ri4, ci, v, bg=bg, border=tb(), sz=9,
                       align="left" if ci <= 4 else "right")
            if ci == 6 and isinstance(v, float):
                fc_v = "1B5E20" if v > 0 else "B71C1C"
                cell.font = Font(bold=s05, color=fc_v,
                                  name="Calibri", size=9)
        rh(ws4, ri4, 15)
        ri4 += 1
    for ci, w in enumerate([10, 8, 18, 16, 7, 12, 14, 14, 10, 10, 18], 1):
        cw(ws4, ci, w)

    # ── S5: Descriptive Statistics ────────────────────────────────────────────
    ws5 = wb.create_sheet("S5 Descriptive Statistics")
    ws5.sheet_view.showGridLines = False
    _title(ws5, 11,
           f"Descriptive Statistics of Annual Rainfall  |  {period}",
           f"Wet-day threshold: ≥{WET_THR} mm/day (WMO)  |  "
           "CV = Coefficient of Variation  |  "
           "Skewness/Kurtosis: Fisher-Pearson")
    hdr5 = ["Station", "Code", "N (yr)", "Mean (mm)", "Median (mm)",
            "Max (mm)", "Min (mm)", "Std (mm)", "CV (%)", "Wet-days/yr",
            "Skewness"]
    _hdr(ws5, 3, hdr5)
    ri5 = 4
    alt = [xfill("E8F4FD"), xfill("FFFFFF")]
    for ni, stn in enumerate(stns, 1):
        bg_fill = alt[ni % 2]
        if stn not in desc_df.index:
            cell = xsc(ws5, ri5, 1, stn, border=tb(), sz=9)
            cell.fill = bg_fill
            rh(ws5, ri5, 15)
            ri5 += 1
            continue
        d    = desc_df.loc[stn]
        vals = [stn, smap.get(stn, stn),
                d["N (yr)"], d["Mean (mm)"], d["Median (mm)"],
                d["Max (mm)"], d["Min (mm)"], d["Std (mm)"],
                d["CV (%)"], d["Wet-days/yr"], d["Skewness"]]
        for ci, v in enumerate(vals, 1):
            if isinstance(v, float) and np.isnan(v):
                v = "—"
            elif isinstance(v, float):
                v = round(v, 1)
            cell = xsc(ws5, ri5, ci, v, border=tb(), sz=9,
                       align="left" if ci <= 2 else "right")
            cell.fill = bg_fill
        rh(ws5, ri5, 16)
        ri5 += 1
    for ci, w in enumerate([10, 8, 8, 12, 12, 12, 12, 10, 8, 10, 10], 1):
        cw(ws5, ci, w)

    # ── S6: Methods & References ──────────────────────────────────────────────
    ws6 = wb.create_sheet("S6 Methods & References")
    ws6.sheet_view.showGridLines = False
    mxsc(ws6, 1, 1, 3,
         "Statistical Methods & References — Rainfall Trend Analysis v2.0",
         bold=True, fc="FFFFFF", bg=XC["title"], sz=13)
    rh(ws6, 1, 26)
    refs = [
        ("Standard Mann–Kendall",
         "Mann (1945); Kendall (1975)",
         "Non-parametric trend test. S = Σ Σ sgn(xⱼ−xᵢ). "
         "Var(S) with tie correction. Z=(S±1)/√Var(S). p two-tailed. "
         "Does NOT account for serial autocorrelation — may overestimate "
         "significance."),
        ("Modified Mann–Kendall",
         "Hamed & Rao (1998) J. Hydrol. 204:182–196",
         "Corrects Var(S) using ranked-series autocorrelations. "
         "n* = n / [1 + 2Σ(1−k/n)ρ_k(ranks)]. Var*(S)=Var(S)×(n/n*). "
         "Only significant ρ_k retained. Recommended when serial "
         "autocorrelation present."),
        ("Sen's Slope + 95% CI",
         "Sen (1968) JASA 63:1379; Gilbert (1987)",
         "Q = median[(xⱼ−xᵢ)/(j−i)] ∀j>i. Magnitude of trend. "
         "95% CI: Cα=z₀.₀₂₅×√Var(S); lo=(N−Cα)/2, hi=(N+Cα)/2+1. "
         "Non-parametric; robust to non-normality and outliers."),
        ("Lag-1 Autocorrelation",
         "Pearson; Box & Jenkins (1976)",
         "r₁ = Σ(xᵢ−x̄)(xᵢ₊₁−x̄)/Σ(xᵢ−x̄)². "
         "Significance: |r₁|>z₀.₀₂₅/√n. "
         "Significant r₁ → use Modified MK (Önöz & Bayazit 2003)."),
        ("Hydrological Seasons",
         "Thai hydro-climatological standard",
         "Wet: May–Oct (monsoon onset to withdrawal). "
         "Dry: Nov–Apr (Nov-Dec of year Y + Jan-Apr of year Y+1). "
         "Wet-day threshold: ≥1.0 mm/day (WMO 2008)."),
        ("Data Quality Control",
         "Tukey (1977); WMO (2008)",
         "Missing data: linear interpolation ≤5 consecutive days. "
         "Outlier detection: upper fence = Q3 + 3×IQR (extreme outlier "
         "threshold). Flagged values retained for transparency."),
        ("Season Definition",
         "Thai Meteorological Department; RID",
         "Wet: May 1 – October 31  (6 months, 184 days). "
         "Dry: November 1 – April 30  (6 months, 181/182 days). "
         "Annual: Calendar year January–December."),
        ("All References",
         "",
         "Mann HB (1945) Econometrica 13:245–259.\n"
         "Kendall MG (1975) Rank Correlation Methods. Griffin, London.\n"
         "Sen PK (1968) JASA 63:1379–1389.\n"
         "Hamed KH, Rao AR (1998) J. Hydrol. 204:182–196.\n"
         "Gilbert RO (1987) Statistical Methods for Environmental Pollution."
         " Van Nostrand.\n"
         "Önöz B, Bayazit M (2003) Hydrol. Sci. J. 48:25–34.\n"
         "Yue S, Wang C (2004) Water Resour. Res. 40:W08307.\n"
         "Box GEP, Jenkins GM (1976) Time Series Analysis. Holden-Day.\n"
         "WMO (2008) Guide to Hydrological Practices. WMO-No. 168."),
    ]
    alt2 = [PatternFill("solid", fgColor="DEEAF1"),
            PatternFill("solid", fgColor="FFFFFF")]
    for ri, (met, ref, desc) in enumerate(refs, 3):
        fl = alt2[ri % 2]
        for ci, v in enumerate([met, ref, desc], 1):
            cell = xsc(ws6, ri, ci, v, bold=(ci <= 2), sz=9.5,
                       align="left", border=tb())
            cell.fill = fl
            if ci == 3:
                cell.alignment = Alignment(horizontal="left",
                                            vertical="top",
                                            wrap_text=True)
        rh(ws6, ri, 68)
    for ci, w in enumerate([26, 42, 68], 1):
        cw(ws6, ci, w)

    # ── S7: 4-Method Comparison ───────────────────────────────────────────────
    if comp4_df is not None:
        ws7 = wb.create_sheet("S7 4-Method Comparison")
        ws7.sheet_view.showGridLines = False
        nc7 = 30
        _title(ws7, nc7,
               f"4-Method Comparison: MK / MMK / PW-MK / TFPW-MK  |  {period}",
               "MK=Standard; MMK=Modified (H&R98); PW=Pre-Whitening MK "
               "(Yue & Wang 2004); TFPW=Trend-Free PW-MK  |  "
               "Green=all agree  Yellow=partial  Red=disagree")
        hdr7 = [
            "Station", "Code", "Scale",
            "rho_1", "Sig_AC",
            "MK_Z", "MK_p", "MK_slope", "MK_sig", "MK_trend",
            "MMK_Z", "MMK_p", "MMK_slope", "MMK_sig", "MMK_trend",
            "PW_Z", "PW_p", "PW_slope", "PW_sig", "PW_trend",
            "TFPW_Z", "TFPW_p", "TFPW_slope", "TFPW_sig", "TFPW_trend",
            "dZ_MMK", "dZ_PW", "dZ_TFPW",
            "dSlope_MMK", "dSlope_PW", "dSlope_TFPW",
            "all_agree", "n_sig_methods",
        ]
        # header is split across two logical groups — use a single row here
        _hdr(ws7, 3, hdr7)
        ri7 = 4
        for _, row in comp4_df.iterrows():
            all_agree   = bool(row.get("all_agree", False))
            n_sig_meth  = int(row.get("n_sig_methods", 0))
            # colour: green=all agree, yellow=partial agreement, red=no agreement
            if all_agree:
                bg = "C8E6C9"       # green
            elif n_sig_meth > 0:
                bg = "FFF9C4"       # yellow
            else:
                bg = "FFCDD2"       # red
            vals = [
                str(row.get("Station", "")),
                str(row.get("Code", "")),
                SCALE_META.get(str(row.get("Scale", "")), {}).get(
                    "label", str(row.get("Scale", ""))),
                row.get("rho_1", np.nan),
                "Yes*" if bool(row.get("Sig_AC", False)) else "No",
                # MK
                row.get("MK_Z", np.nan),
                row.get("MK_p", np.nan),
                row.get("MK_slope", np.nan),
                str(row.get("MK_sig", "ns")),
                str(row.get("MK_trend", "—")),
                # MMK
                row.get("MMK_Z", np.nan),
                row.get("MMK_p", np.nan),
                row.get("MMK_slope", np.nan),
                str(row.get("MMK_sig", "ns")),
                str(row.get("MMK_trend", "—")),
                # PW
                row.get("PW_Z", np.nan),
                row.get("PW_p", np.nan),
                row.get("PW_slope", np.nan),
                str(row.get("PW_sig", "ns")),
                str(row.get("PW_trend", "—")),
                # TFPW
                row.get("TFPW_Z", np.nan),
                row.get("TFPW_p", np.nan),
                row.get("TFPW_slope", np.nan),
                str(row.get("TFPW_sig", "ns")),
                str(row.get("TFPW_trend", "—")),
                # deltas
                row.get("dZ_MMK", np.nan),
                row.get("dZ_PW", np.nan),
                row.get("dZ_TFPW", np.nan),
                row.get("dSlope_MMK", np.nan),
                row.get("dSlope_PW", np.nan),
                row.get("dSlope_TFPW", np.nan),
                # summary
                "Yes" if all_agree else "No",
                n_sig_meth,
            ]
            text_cols = {1, 2, 3, 5, 9, 10, 14, 15, 19, 20, 24, 25, 32}
            for ci, v in enumerate(vals, 1):
                if isinstance(v, float) and np.isnan(v):
                    v = "—"
                elif isinstance(v, float):
                    v = round(v, 4)
                cell = xsc(ws7, ri7, ci, v, bg=bg, border=tb(), sz=9,
                           align="left" if ci in text_cols else "right")
                # highlight all_agree column
                if ci == 32:
                    if v == "No":
                        cell.fill = xfill("FFCCBC")
                        cell.font = Font(bold=True, color="B71C1C",
                                         name="Calibri", size=9)
                    else:
                        cell.font = Font(bold=True, color="1B5E20",
                                         name="Calibri", size=9)
                # highlight significant autocorrelation
                if ci == 5 and "Yes" in str(v):
                    cell.font = Font(bold=True, color="E65100",
                                     name="Calibri", size=9)
            rh(ws7, ri7, 15)
            ri7 += 1
        # column widths
        col_widths = ([10, 8, 20, 8, 7]           # stn/code/scale/rho/sigAC
                      + [9, 9, 10, 7, 16] * 4     # 4 × (Z,p,slope,sig,trend)
                      + [9, 9, 9]                  # dZ
                      + [11, 11, 11]               # dSlope
                      + [9, 12])                   # all_agree, n_sig
        for ci, w in enumerate(col_widths, 1):
            cw(ws7, ci, w)

    # ── S8: Field Significance ────────────────────────────────────────────────
    if field_sig_df is not None:
        ws8 = wb.create_sheet("S8 Field Significance")
        ws8.sheet_view.showGridLines = False
        nc8 = 12
        _title(ws8, nc8,
               f"Field Significance Analysis  |  {period}",
               "Walker (1914) global test; Livezey–Chen (1983) Monte Carlo  |  "
               "Fraction of sig. stations vs. binomial expectation under H₀")
        hdr8 = [
            "Scale",
            "N_stations",
            "N_sig_MK",
            "N_sig_MMK",
            "Frac_sig_MK",
            "Frac_sig_MMK",
            "Walker_p_MK",
            "Walker_sig_MK",
            "LC_p_MK",
            "LC_sig_MK",
            "LC_null_mean",
            "LC_null_95th",
        ]
        _hdr(ws8, 3, hdr8)
        ri8 = 4
        alt8 = [xfill("E3F2FD"), xfill("FFFFFF")]
        for ni, (_, row) in enumerate(field_sig_df.iterrows()):
            bg_fill = alt8[ni % 2]
            walker_sig = bool(row.get("Walker_sig_MK", False))
            lc_sig     = bool(row.get("LC_sig_MK", False))
            if walker_sig or lc_sig:
                bg_fill = xfill("FFF9C4")   # highlight if field-significant
            vals = [
                str(row.get("Scale", "—")),
                row.get("N_stations", "—"),
                row.get("N_sig_MK", "—"),
                row.get("N_sig_MMK", "—"),
                row.get("Frac_sig_MK", np.nan),
                row.get("Frac_sig_MMK", np.nan),
                row.get("Walker_p_MK", np.nan),
                "Yes*" if walker_sig else "No",
                row.get("LC_p_MK", np.nan),
                "Yes*" if lc_sig else "No",
                row.get("LC_null_mean", np.nan),
                row.get("LC_null_95th", np.nan),
            ]
            for ci, v in enumerate(vals, 1):
                if isinstance(v, float) and np.isnan(v):
                    v = "—"
                elif isinstance(v, float):
                    v = round(v, 4)
                cell = xsc(ws8, ri8, ci, v, border=tb(), sz=9,
                           align="left" if ci == 1 else "right")
                cell.fill = bg_fill
                if ci in (8, 10) and "Yes" in str(v):
                    cell.font = Font(bold=True, color="1B5E20",
                                     name="Calibri", size=9)
            rh(ws8, ri8, 15)
            ri8 += 1
        for ci, w in enumerate(
                [20, 12, 10, 11, 12, 13, 12, 14, 10, 10, 13, 13], 1):
            cw(ws8, ci, w)

    # ── S9: Dry Season Validation ─────────────────────────────────────────────
    if dry_validation is not None:
        ws9 = wb.create_sheet("S9 Dry Season Validation")
        ws9.sheet_view.showGridLines = False
        nc9 = 4
        valid_status = dry_validation.get("valid", False)
        status_str   = "PASSED" if valid_status else "FAILED"
        status_bg    = "C8E6C9" if valid_status else "FFCDD2"
        _title(ws9, nc9,
               f"Dry Season Validation Report  |  {period}",
               "Validates that hydrological dry-season year assignment "
               "is consistent across all stations and analysis years.")
        # Validation status banner
        mxsc(ws9, 3, 1, nc9,
             f"Validation Status: {status_str}",
             bold=True,
             fc="1B5E20" if valid_status else "B71C1C",
             bg=status_bg,
             sz=11, align="left")
        rh(ws9, 3, 20)
        # Summary info rows
        summary_rows = [
            ("Years covered",   str(dry_validation.get("years", "—"))),
            ("Number of blocks", str(dry_validation.get("n_blocks", "—"))),
        ]
        ri9 = 4
        for label, val in summary_rows:
            xsc(ws9, ri9, 1, label, bold=True, border=tb(), sz=9,
                align="left", bg=XC["ann_h"])
            xsc(ws9, ri9, 2, val, border=tb(), sz=9, align="left")
            rh(ws9, ri9, 15)
            ri9 += 1
        # Error table (if any)
        errors = dry_validation.get("errors", [])
        if errors:
            ri9 += 1  # blank row spacer
            mxsc(ws9, ri9, 1, nc9, "Validation Errors",
                 bold=True, fc="FFFFFF", bg="C62828", sz=10, align="left")
            rh(ws9, ri9, 18)
            ri9 += 1
            _hdr(ws9, ri9, ["#", "Year", "Station", "Error Description"],
                 bg="E53935")
            ri9 += 1
            for ei, err in enumerate(errors, 1):
                if isinstance(err, dict):
                    yr  = str(err.get("year", "—"))
                    stn = str(err.get("station", "—"))
                    msg = str(err.get("message", str(err)))
                else:
                    yr  = "—"
                    stn = "—"
                    msg = str(err)
                bg_err = xfill("FFEBEE") if ei % 2 == 0 else xfill("FFFFFF")
                for ci, v in enumerate([str(ei), yr, stn, msg], 1):
                    cell = xsc(ws9, ri9, ci, v, border=tb(), sz=9,
                               align="left")
                    cell.fill = bg_err
                rh(ws9, ri9, 15)
                ri9 += 1
        else:
            ri9 += 1
            mxsc(ws9, ri9, 1, nc9, "No errors found — dry season assignment is valid.",
                 italic=True, fc="1B5E20", bg="E8F5E9", sz=9, align="left")
            rh(ws9, ri9, 15)
        for ci, w in enumerate([8, 14, 18, 60], 1):
            cw(ws9, ci, w)

    # ── save ──────────────────────────────────────────────────────────────────
    wb.save(str(out_xlsx))
    n_sheets = 6 + (1 if comp4_df is not None else 0) \
                 + (1 if field_sig_df is not None else 0) \
                 + (1 if dry_validation is not None else 0)
    print(f"  ✓  Excel: {Path(out_xlsx).name}  ({n_sheets} sheets)")
