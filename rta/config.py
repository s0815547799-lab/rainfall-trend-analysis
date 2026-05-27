"""
rta.config — Constants, colour palette, matplotlib style, Excel helpers.

Extracted from rainfall_trend_analysis_v3.py §0 (lines 83–187).
"""

# ── Standard library ──────────────────────────────────────────────────────────
from pathlib import Path

# ── Visualisation ─────────────────────────────────────────────────────────────
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# ── Excel ─────────────────────────────────────────────────────────────────────
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  Version & scalar constants                                              ║
# ╚══════════════════════════════════════════════════════════════════════════╝

VERSION    = "4.0"
WET_THR    = 1.0                    # WMO wet-day threshold (mm/day)
WET_MONTHS = [5, 6, 7, 8, 9, 10]   # Wet season: May–October
DRY_MONTHS = [11, 12, 1, 2, 3, 4]  # Dry season: November–April
MIN_N      = 10                     # minimum years for MK test
ALPHA_005  = 0.05
ALPHA_001  = 0.01
Z_005      = 1.9600
Z_001      = 2.5758
SAVE_PDF   = True
DPI        = 600
MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

MISS_FLAGS = [-99, -999, -9999, -9.99e+20, 9.99e+20, 1e+20]


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  Colour palette (colour-blind safe)                                      ║
# ╚══════════════════════════════════════════════════════════════════════════╝

C = dict(
    annual  = "#37474F",  annual_lt = "#B0BEC5",
    wet     = "#1565C0",  wet_lt    = "#90CAF9",
    dry     = "#E65100",  dry_lt    = "#FFCC80",
    inc     = "#1B5E20",  inc_lt    = "#A5D6A7",
    dec     = "#B71C1C",  dec_lt    = "#EF9A9A",
    ns_col  = "#78909C",  ns_lt     = "#CFD8DC",
    mk_std  = "#6A1B9A",  mk_mod    = "#0277BD",
    gold    = "#F9A825",  grey      = "#546E7A",
)


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  Matplotlib publication style                                            ║
# ╚══════════════════════════════════════════════════════════════════════════╝

plt.rcParams.update({
    "font.family":        "serif",
    "font.serif":         ["Times New Roman", "DejaVu Serif"],
    "font.size":          12,
    "axes.titlesize":     13,
    "axes.labelsize":     12,
    "xtick.labelsize":    11,
    "ytick.labelsize":    11,
    "legend.fontsize":    10.5,
    "figure.titlesize":   13,
    "lines.linewidth":    2.0,
    "axes.linewidth":     1.0,
    "axes.spines.top":    False,
    "axes.spines.right":  False,
    "axes.grid":          True,
    "grid.linestyle":     "--",
    "grid.linewidth":     0.4,
    "grid.alpha":         0.40,
    "grid.color":         "#B0BEC5",
    "savefig.dpi":        DPI,
    "savefig.bbox":       "tight",
    "savefig.pad_inches": 0.15,
    "figure.dpi":         100,
    "mathtext.fontset":   "stix",
    "pdf.fonttype":       42,
    "ps.fonttype":        42,
})


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  Excel style constants                                                   ║
# ╚══════════════════════════════════════════════════════════════════════════╝

THIN = Side(style="thin",   color="BDBDBD")
MED  = Side(style="medium", color="1F4E79")

XC = dict(
    title  = "13293D", sub   = "1F4E79", hdr   = "2E75B6",
    wet_h  = "DDEEFF", dry_h = "FFF3E0",
    ann_h  = "ECEFF1", mon_h = "E8F5E9",
    sig05  = "FFF9C4", sig01 = "FFECB3",
    inc_c  = "E8F5E9", dec_c = "FFEBEE",
    ns_c   = "F5F5F5", white = "FFFFFF",
    mk_h   = "EDE7F6", mmk_h = "E3F2FD",
    diff_h = "FFF8E1",
)


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  Excel cell helper functions                                             ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def tb():
    """Thin border on all four sides."""
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def xfill(h):
    """Solid PatternFill from a hex colour string."""
    return PatternFill("solid", fgColor=h)


def xsc(ws, r, c, val=None, bold=False, italic=False,
        fc=None, bg=None, align="center", sz=10, wrap=True, border=None):
    """Write a styled value into cell (r, c) of worksheet ws."""
    cell = ws.cell(row=r, column=c)
    if val is not None:
        cell.value = val
    cell.font      = Font(bold=bold, italic=italic, name="Calibri",
                          size=sz, color=fc if fc else "1A1A1A")
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    if bg:
        cell.fill = xfill(bg)
    if border:
        cell.border = border
    return cell


def mxsc(ws, r, c1, c2, val, **kw):
    """Merge cells from column c1 to c2 on row r, then style with xsc."""
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    return xsc(ws, r, c1, val, **kw)


def cw(ws, col, w):
    """Set column width for column index col (1-based)."""
    ws.column_dimensions[get_column_letter(col)].width = w


def rh(ws, r, h):
    """Set row height for row r."""
    ws.row_dimensions[r].height = h


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  Scale metadata                                                          ║
# ╚══════════════════════════════════════════════════════════════════════════╝

SCALE_META = {
    "annual": {"label": "Annual (Jan–Dec)",        "unit": "mm yr⁻¹",    "color": C["annual"]},
    "wet":    {"label": "Wet Season (May–Oct)",     "unit": "mm season⁻¹", "color": C["wet"]},
    "dry":    {"label": "Dry Season (Nov–Apr)",     "unit": "mm season⁻¹", "color": C["dry"]},
}


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  Figure save helper                                                      ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def savefig(fig, path_noext: str) -> None:
    """
    Save figure as PNG at 600 DPI.  If SAVE_PDF is True also save PDF.
    Closes the figure afterwards.
    """
    fig.savefig(f"{path_noext}.png", dpi=DPI, bbox_inches="tight", pad_inches=0.15)
    if SAVE_PDF:
        fig.savefig(f"{path_noext}.pdf", bbox_inches="tight", pad_inches=0.15)
    plt.close(fig)
    print(f"    ✓  {Path(path_noext).name}.png" + (" + .pdf" if SAVE_PDF else ""))
