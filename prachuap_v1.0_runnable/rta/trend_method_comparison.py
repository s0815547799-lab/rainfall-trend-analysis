"""
rta.trend_method_comparison — Hierarchical 4-method publication database.

Reads ONLY from existing validated Excel workbooks.  No statistical recomputation.

Source map
==========
WB1  *_Results.xlsx (canonical V4)
     S1  Standard MK           Z, p, tau, rho_1, N, S, Var(S)
     S2  Modified MK (H&R98)   + Var*(S), n_eff, ρ₁(ranked)
     S4  Sens Slope             β + 95 % CI, all 4 methods (144 rows)
     S7  4-Method Comparison    PW-MK + TFPW-MK ONLY here; dZ / dSlope pre-computed
     S8  Field Significance     Walker (1914) + Livezey-Chen (1983)
WB4  Supplementary (optional)
     3_Modified_MK             Significant_Lags — lag index + ρ per lag
     4_Pettitt_CP              Change-point year + Pettitt p-value

Read conventions
================
WB1 : skiprows=[0, 1], header=0   (rows 0–1 = title / subtitle)
WB4 : header=0                     (no title rows)
WB4 Scale : remap short→long before any join

Scale availability
==================
Annual (Jan–Dec)      ✓  34 years (1981–2014),  12 stations
Wet Season (May–Oct)  ✓  34 years,  12 stations
Dry Season (Nov–Apr)  ✓  34 years,  12 stations
Monthly               ✗  not in any workbook output
Seasonal (4-split)    ✗  2-season split only

Output sheet plan — Trend_Method_Comparison_Q1.xlsx
====================================================
A  Method sheets      01_MK_Results, 02_MMK_Results, 03_PW_Results,
                      04_TFPW_Results, 05_SenSlope_Results
B  MK comparisons     11_MK_vs_MMK, 12_MK_vs_PW, 13_MK_vs_TFPW
C  Full pairwise      21_MMK_vs_PW, 22_MMK_vs_TFPW, 23_PW_vs_TFPW
D  Publication tables 31_Significant_Stations, 32_Method_Agreement,
                      33_Autocorrelation_Impact, 34_Field_Significance,
                      35_Method_Summary
E  Master database    90_Master_Station_Database
   Supplementary      91_Pettitt_CP, 99_Metadata
"""

from __future__ import annotations

import re
import warnings
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

# ── Significance thresholds ───────────────────────────────────────────────────
_Z_05 = 1.9600
_Z_01 = 2.5758

# ── Scale constants ───────────────────────────────────────────────────────────
_SCALE_LONG = {
    "annual": "Annual (Jan–Dec)",
    "wet":    "Wet Season (May–Oct)",
    "dry":    "Dry Season (Nov–Apr)",
}
_SCALE_ORDER = [
    "Annual (Jan–Dec)",
    "Wet Season (May–Oct)",
    "Dry Season (Nov–Apr)",
]

# ── WB1 exact sheet names ─────────────────────────────────────────────────────
_SH_S1 = "S1 Standard MK"
_SH_S2 = "S2 Modified MK (H&R98)"
_SH_S4 = "S4 Sens Slope"
_SH_S7 = "S7 4-Method Comparison"
_SH_S8 = "S8 Field Significance"

# ── WB4 exact sheet names ─────────────────────────────────────────────────────
_WB4_LAGS    = "3_Modified_MK"
_WB4_PETTITT = "4_Pettitt_CP"

# ── Tab / header colours per sheet-group ─────────────────────────────────────
_SHEET_GROUPS = {
    "A": ("4472C4", "D9E1F2"),   # Blue   — method sheets
    "B": ("70AD47", "E2EFDA"),   # Green  — MK baseline comparisons
    "C": ("17A589", "D1F2EB"),   # Teal   — full pairwise
    "D": ("ED7D31", "FCE4D6"),   # Orange — publication tables
    "E": ("C9992B", "FFF2CC"),   # Gold   — master database
    "S": ("595959", "EDEDED"),   # Grey   — supplementary
}


# ══════════════════════════════════════════════════════════════════════════════
#  Module-level helpers
# ══════════════════════════════════════════════════════════════════════════════

def _read_wb1(path: Path, sheet: str) -> pd.DataFrame:
    """WB1: rows 0–1 = title/subtitle; row 2 = column headers."""
    df = pd.read_excel(path, sheet_name=sheet, header=0, skiprows=[0, 1])
    if "Station" in df.columns:
        df["Station"] = df["Station"].astype(str).str.strip()
    return df


def _read_wb4(path: Path, sheet: str) -> pd.DataFrame:
    """WB4: row 0 = column headers directly."""
    df = pd.read_excel(path, sheet_name=sheet, header=0)
    if "Station" in df.columns:
        df["Station"] = df["Station"].astype(str).str.strip()
    return df


def _sort(df: pd.DataFrame) -> pd.DataFrame:
    """Sort by canonical Scale order then Station."""
    if "Scale" not in df.columns:
        return df
    df = df.copy()
    df["_ord"] = df["Scale"].map({s: i for i, s in enumerate(_SCALE_ORDER)}).fillna(99)
    cols = ["_ord", "Station"] if "Station" in df.columns else ["_ord"]
    df = df.sort_values(cols).drop(columns="_ord").reset_index(drop=True)
    return df


def _parse_lags(raw) -> str:
    """Parse WB4 Significant_Lags repr → 'k=1(ρ=+0.471); k=9(ρ=−0.494)' or '—'."""
    s = str(raw).strip()
    if s in ("", "nan", "None", "[]"):
        return "—"
    pairs = re.findall(r"\((\d+),\s*np\.float64\(([^)]+)\)\)", s)
    if not pairs:
        return "—"
    parts = []
    for lag, rho in pairs:
        rho_f = float(rho)
        sign = "+" if rho_f >= 0 else "−"
        parts.append(f"k={lag}(ρ={sign}{abs(rho_f):.3f})")
    return "; ".join(parts)


def _add_sig(df: pd.DataFrame, z_col: str, p_col: str, px: str = "") -> pd.DataFrame:
    """
    Append Sig_p05, Sig_p01, Sig_Z05, Sig_Z01, Sig_Consistency.

    px : column prefix, e.g. 'MK' → 'MK_Sig_p05'.  Empty string → 'Sig_p05'.
    """
    pre = f"{px}_" if px else ""
    df[f"{pre}Sig_p05"]        = df[p_col] < 0.05
    df[f"{pre}Sig_p01"]        = df[p_col] < 0.01
    df[f"{pre}Sig_Z05"]        = df[z_col].abs() >= _Z_05
    df[f"{pre}Sig_Z01"]        = df[z_col].abs() >= _Z_01
    df[f"{pre}Sig_Consistency"] = df[f"{pre}Sig_p05"] == df[f"{pre}Sig_Z05"]
    return df


def _add_delta(
    df: pd.DataFrame,
    ref_z: str, ref_p: str, ref_trend: str, ref_sig_z05: str,
    alt_z: str, alt_p: str, alt_trend: str, alt_sig_z05: str,
) -> pd.DataFrame:
    """Append Delta_Z, Delta_p, Direction_Changed, Significance_Changed."""
    df["Delta_Z"]              = (df[alt_z]     - df[ref_z]).round(6)
    df["Delta_p"]              = (df[alt_p]     - df[ref_p]).round(6)
    df["Direction_Changed"]    = df[ref_trend]   != df[alt_trend]
    df["Significance_Changed"] = df[ref_sig_z05] != df[alt_sig_z05]
    return df


# ══════════════════════════════════════════════════════════════════════════════
#  TrendMethodComparison
# ══════════════════════════════════════════════════════════════════════════════

class TrendMethodComparison:
    """
    Build a hierarchical 4-method publication database from validated
    V4 pipeline Excel outputs.  No statistical recomputation.

    Parameters
    ----------
    wb1_path : canonical V4 *_Results.xlsx workbook
    wb4_path : supplementary workbook (Significant_Lags + Pettitt CP); optional
    """

    def __init__(
        self,
        wb1_path: str | Path,
        wb4_path: str | Path | None = None,
    ) -> None:
        self.wb1 = Path(wb1_path)
        self.wb4 = Path(wb4_path) if wb4_path else None
        if not self.wb1.exists():
            raise FileNotFoundError(f"WB1 not found: {self.wb1}")
        if self.wb4 and not self.wb4.exists():
            warnings.warn(f"WB4 not found: {self.wb4} — Pettitt CP and Lag details omitted.")
            self.wb4 = None

    # ── Raw sheet loaders ─────────────────────────────────────────────────────

    def _s1(self) -> pd.DataFrame:
        return _read_wb1(self.wb1, _SH_S1)

    def _s2(self) -> pd.DataFrame:
        return _read_wb1(self.wb1, _SH_S2)

    def _s4(self) -> pd.DataFrame:
        return _read_wb1(self.wb1, _SH_S4)

    def _s7(self) -> pd.DataFrame:
        return _read_wb1(self.wb1, _SH_S7)

    def _s8(self) -> pd.DataFrame:
        df = _read_wb1(self.wb1, _SH_S8)
        df["Scale"] = df["Scale"].map(_SCALE_LONG).fillna(df["Scale"])
        return df

    # ── Shared AC block (n_eff, Correction_Factor, Lag) ──────────────────────

    def _ac_block(self) -> pd.DataFrame:
        """
        Return DataFrame (Station, Scale, n_eff, Correction_Factor, Lag).
        Correction_Factor = Var*(S) / Var(S) computed from S2.
        Lag = parsed significant lags from WB4 (or placeholder).
        """
        s2 = self._s2()
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            cf = (s2["Var*(S)"] / s2["Var(S)"].replace(0, np.nan)).fillna(1.0).round(6)
        ac = s2[["Station", "Scale", "n_eff"]].copy()
        ac["Correction_Factor"] = cf

        if self.wb4 is not None:
            raw = _read_wb4(self.wb4, _WB4_LAGS)
            raw["Scale"] = raw["Scale"].map(_SCALE_LONG).fillna(raw["Scale"])
            raw["Lag"] = raw["Significant_Lags"].apply(_parse_lags)
            ac = ac.merge(raw[["Station", "Scale", "Lag"]], on=["Station", "Scale"], how="left")
        else:
            ac["Lag"] = "—"

        return ac  # cols: Station, Scale, n_eff, Correction_Factor, Lag

    # ═════════════════════════════════════════════════════════════════════════
    #  A — Method-specific sheets
    # ═════════════════════════════════════════════════════════════════════════

    def build_01_mk(self) -> pd.DataFrame:
        """
        01_MK_Results — 36 rows
        Source: S1 (full, richest MK source).
        Adds: Sig_p05, Sig_p01, Sig_Z05, Sig_Z01, Sig_Consistency.
        """
        s1 = self._s1().rename(columns={
            "τ (Kendall)": "Tau",
            "p-value":     "p",
            "* p<0.05":    "Sig_label_05",
            "** p<0.01":   "Sig_label_01",
        })
        # S7 carries Sig_AC flag not present in S1
        s7_ac = self._s7()[["Station", "Scale", "Sig_AC"]].copy()
        s1 = s1.merge(s7_ac, on=["Station", "Scale"], how="left")
        s1 = _add_sig(s1, "Z", "p")
        return _sort(s1[[
            "Station", "Code", "Scale",
            "N", "S", "Var(S)", "rho_1", "Sig_AC",
            "Z", "Tau", "p", "Trend",
            "Sig_label_05", "Sig_label_01",
            "Sig_p05", "Sig_p01", "Sig_Z05", "Sig_Z01", "Sig_Consistency",
        ]])

    def build_02_mmk(self) -> pd.DataFrame:
        """
        02_MMK_Results — 36 rows
        Source: S2 + AC block (n_eff, Correction_Factor, Lag from WB4).
        Adds: Sig_p05, Sig_p01, Sig_Z05, Sig_Z01, Sig_Consistency.
        """
        s2 = self._s2().rename(columns={
            "ρ₁":          "rho_1_ranked",
            "τ (Kendall)": "Tau",
            "p-value":     "p",
            "* p<0.05":    "Sig_label_05",
            "** p<0.01":   "Sig_label_01",
        })
        ac = self._ac_block().drop(columns=["n_eff"])  # s2 already has n_eff
        s7_ac = self._s7()[["Station", "Scale", "Sig_AC"]].copy()
        df = s2.merge(ac,   on=["Station", "Scale"], how="left")
        df = df.merge(s7_ac, on=["Station", "Scale"], how="left")
        df = _add_sig(df, "Z", "p")
        return _sort(df[[
            "Station", "Code", "Scale",
            "N", "S", "Var(S)", "Var*(S)",
            "n_eff", "Correction_Factor", "rho_1_ranked", "Sig_AC", "Lag",
            "Z", "Tau", "p", "Trend",
            "Sig_label_05", "Sig_label_01",
            "Sig_p05", "Sig_p01", "Sig_Z05", "Sig_Z01", "Sig_Consistency",
        ]])

    def build_03_pw(self) -> pd.DataFrame:
        """
        03_PW_Results — 36 rows
        Source: S7 (PW columns) + S1 (rho_1) + AC block for context.
        Adds: Sig_p05, Sig_p01, Sig_Z05, Sig_Z01, Sig_Consistency.
        Also adds Delta_Z_vs_MK (= PW_Z − MK_Z, pre-computed in S7 as dZ_PW).
        """
        s7 = self._s7()
        ac = self._ac_block()

        df = s7.merge(ac, on=["Station", "Scale"], how="left")
        df = df.rename(columns={
            "PW_Z":     "Z",
            "PW_p":     "p",
            "PW_slope": "Slope_mm_yr",
            "PW_sig":   "Sig_label",
            "PW_trend": "Trend",
            "dZ_PW":    "Delta_Z_vs_MK",
        })
        df = _add_sig(df, "Z", "p")
        return _sort(df[[
            "Station", "Code", "Scale",
            "rho_1", "Sig_AC", "n_eff", "Correction_Factor",
            "Z", "p", "Slope_mm_yr", "Trend", "Sig_label",
            "Sig_p05", "Sig_p01", "Sig_Z05", "Sig_Z01", "Sig_Consistency",
            "Delta_Z_vs_MK",
        ]])

    def build_04_tfpw(self) -> pd.DataFrame:
        """
        04_TFPW_Results — 36 rows
        Source: S7 (TFPW columns) + S1 (rho_1) + AC block for context.
        """
        s7 = self._s7()
        ac = self._ac_block()

        df = s7.merge(ac, on=["Station", "Scale"], how="left")
        df = df.rename(columns={
            "TFPW_Z":     "Z",
            "TFPW_p":     "p",
            "TFPW_slope": "Slope_mm_yr",
            "TFPW_sig":   "Sig_label",
            "TFPW_trend": "Trend",
            "dZ_TFPW":    "Delta_Z_vs_MK",
        })
        df = _add_sig(df, "Z", "p")
        return _sort(df[[
            "Station", "Code", "Scale",
            "rho_1", "Sig_AC", "n_eff", "Correction_Factor",
            "Z", "p", "Slope_mm_yr", "Trend", "Sig_label",
            "Sig_p05", "Sig_p01", "Sig_Z05", "Sig_Z01", "Sig_Consistency",
            "Delta_Z_vs_MK",
        ]])

    def build_05_senslope(self) -> pd.DataFrame:
        """
        05_SenSlope_Results — 144 rows (4 methods × 12 stations × 3 scales).
        Source: S4 — only source with per-method 95 % CIs.
        """
        s4 = self._s4().rename(columns={
            "β (mm/yr)":        "Slope_mm_yr",
            "CI_Lower (mm/yr)": "CI_Lower",
            "CI_Upper (mm/yr)": "CI_Upper",
            "p-value":          "p",
        })
        s4 = _add_sig(s4, "Z", "p")
        return _sort(s4[[
            "Station", "Code", "Scale", "Method",
            "N", "Slope_mm_yr", "CI_Lower", "CI_Upper",
            "Z", "p", "Trend",
            "Sig_p05", "Sig_p01", "Sig_Z05", "Sig_Z01", "Sig_Consistency",
        ]])

    # ═════════════════════════════════════════════════════════════════════════
    #  B — MK baseline comparisons
    # ═════════════════════════════════════════════════════════════════════════

    def _mk_ref(self) -> pd.DataFrame:
        """MK reference columns for B / C group joins."""
        s1 = self._s1()
        df = s1.rename(columns={
            "Z":           "MK_Z",
            "τ (Kendall)": "MK_Tau",
            "p-value":     "MK_p",
            "Trend":       "MK_Trend",
            "* p<0.05":    "MK_Sig_label_05",
            "** p<0.01":   "MK_Sig_label_01",
        })
        df = _add_sig(df, "MK_Z", "MK_p", px="MK")
        return df[[
            "Station", "Code", "Scale",
            "N", "S", "Var(S)", "rho_1",
            "MK_Z", "MK_Tau", "MK_p", "MK_Trend",
            "MK_Sig_label_05", "MK_Sig_label_01",
            "MK_Sig_p05", "MK_Sig_p01",
            "MK_Sig_Z05",  "MK_Sig_Z01", "MK_Sig_Consistency",
        ]]

    def build_11_mk_vs_mmk(self) -> pd.DataFrame:
        """
        11_MK_vs_MMK — 36 rows
        MK (S1) vs Modified MK (S2).
        Delta_Z = MMK_Z − MK_Z (positive → MMK has stronger positive trend).
        Includes: n_eff, Correction_Factor, Lag.
        """
        base = self._mk_ref()
        s2 = self._s2().rename(columns={
            "ρ₁":          "rho_1_ranked",
            "τ (Kendall)": "MMK_Tau",
            "Z":           "MMK_Z",
            "p-value":     "MMK_p",
            "Trend":       "MMK_Trend",
            "* p<0.05":    "MMK_Sig_label_05",
            "** p<0.01":   "MMK_Sig_label_01",
        })
        ac = self._ac_block()
        s7_ac = self._s7()[["Station", "Scale", "Sig_AC"]].copy()

        df = (
            base
            .merge(s2[["Station", "Scale", "Var*(S)",
                        "rho_1_ranked",
                        "MMK_Z", "MMK_Tau", "MMK_p", "MMK_Trend",
                        "MMK_Sig_label_05", "MMK_Sig_label_01"]],
                   on=["Station", "Scale"])
            .merge(ac,    on=["Station", "Scale"], how="left")
            .merge(s7_ac, on=["Station", "Scale"], how="left")
        )
        df = _add_sig(df, "MMK_Z", "MMK_p", px="MMK")
        df = _add_delta(df,
                        "MK_Z",  "MK_p",  "MK_Trend",  "MK_Sig_Z05",
                        "MMK_Z", "MMK_p", "MMK_Trend", "MMK_Sig_Z05")
        return _sort(df[[
            "Station", "Code", "Scale",
            "N", "S", "Var(S)", "Var*(S)",
            "rho_1", "rho_1_ranked", "Sig_AC", "n_eff", "Correction_Factor", "Lag",
            "MK_Z",  "MK_p",  "MK_Trend",  "MK_Sig_label_05",
            "MK_Sig_p05",  "MK_Sig_p01",  "MK_Sig_Z05",  "MK_Sig_Z01",  "MK_Sig_Consistency",
            "MMK_Z", "MMK_p", "MMK_Trend", "MMK_Sig_label_05",
            "MMK_Sig_p05", "MMK_Sig_p01", "MMK_Sig_Z05", "MMK_Sig_Z01", "MMK_Sig_Consistency",
            "Delta_Z", "Delta_p", "Direction_Changed", "Significance_Changed",
        ]])

    def build_12_mk_vs_pw(self) -> pd.DataFrame:
        """
        12_MK_vs_PW — 36 rows
        MK (S1) vs PW-MK (S7).  Delta_Z = PW_Z − MK_Z.
        """
        base = self._mk_ref()
        s7 = self._s7()
        ac = self._ac_block()

        df = (
            base
            .merge(s7[["Station", "Scale", "Sig_AC",
                        "PW_Z", "PW_p", "PW_slope", "PW_sig", "PW_trend",
                        "dZ_PW"]],
                   on=["Station", "Scale"])
            .merge(ac, on=["Station", "Scale"], how="left")
        )
        df = df.rename(columns={
            "PW_sig":   "PW_Sig_label",
            "PW_trend": "PW_Trend",
        })
        df = _add_sig(df, "PW_Z", "PW_p", px="PW")
        df = _add_delta(df,
                        "MK_Z", "MK_p", "MK_Trend", "MK_Sig_Z05",
                        "PW_Z", "PW_p", "PW_Trend", "PW_Sig_Z05")
        return _sort(df[[
            "Station", "Code", "Scale",
            "N", "rho_1", "Sig_AC", "n_eff", "Correction_Factor", "Lag",
            "MK_Z",  "MK_p",  "MK_Trend",  "MK_Sig_label_05",
            "MK_Sig_p05",  "MK_Sig_p01",  "MK_Sig_Z05",  "MK_Sig_Z01",  "MK_Sig_Consistency",
            "PW_Z",  "PW_p",  "PW_Trend",  "PW_Sig_label",
            "PW_Sig_p05",  "PW_Sig_p01",  "PW_Sig_Z05",  "PW_Sig_Z01",  "PW_Sig_Consistency",
            "Delta_Z", "Delta_p", "Direction_Changed", "Significance_Changed",
        ]])

    def build_13_mk_vs_tfpw(self) -> pd.DataFrame:
        """
        13_MK_vs_TFPW — 36 rows
        MK (S1) vs TFPW-MK (S7).  Delta_Z = TFPW_Z − MK_Z.
        """
        base = self._mk_ref()
        s7 = self._s7()
        ac = self._ac_block()

        df = (
            base
            .merge(s7[["Station", "Scale", "Sig_AC",
                        "TFPW_Z", "TFPW_p", "TFPW_slope", "TFPW_sig", "TFPW_trend",
                        "dZ_TFPW"]],
                   on=["Station", "Scale"])
            .merge(ac, on=["Station", "Scale"], how="left")
        )
        df = df.rename(columns={
            "TFPW_sig":   "TFPW_Sig_label",
            "TFPW_trend": "TFPW_Trend",
        })
        df = _add_sig(df, "TFPW_Z", "TFPW_p", px="TFPW")
        df = _add_delta(df,
                        "MK_Z",    "MK_p",    "MK_Trend",    "MK_Sig_Z05",
                        "TFPW_Z",  "TFPW_p",  "TFPW_Trend",  "TFPW_Sig_Z05")
        return _sort(df[[
            "Station", "Code", "Scale",
            "N", "rho_1", "Sig_AC", "n_eff", "Correction_Factor", "Lag",
            "MK_Z",   "MK_p",   "MK_Trend",   "MK_Sig_label_05",
            "MK_Sig_p05",   "MK_Sig_p01",   "MK_Sig_Z05",   "MK_Sig_Z01",   "MK_Sig_Consistency",
            "TFPW_Z", "TFPW_p", "TFPW_Trend", "TFPW_Sig_label",
            "TFPW_Sig_p05", "TFPW_Sig_p01", "TFPW_Sig_Z05", "TFPW_Sig_Z01", "TFPW_Sig_Consistency",
            "Delta_Z", "Delta_p", "Direction_Changed", "Significance_Changed",
        ]])

    # ═════════════════════════════════════════════════════════════════════════
    #  C — Full pairwise comparisons
    # ═════════════════════════════════════════════════════════════════════════

    def _mmk_ref_cols(self) -> pd.DataFrame:
        """MMK reference columns for C group: from S2."""
        s2 = self._s2().rename(columns={
            "Z":       "MMK_Z",
            "p-value": "MMK_p",
            "Trend":   "MMK_Trend",
            "* p<0.05":"MMK_Sig_label",
        })
        df = _add_sig(s2, "MMK_Z", "MMK_p", px="MMK")
        return df[["Station", "Scale",
                   "MMK_Z", "MMK_p", "MMK_Trend", "MMK_Sig_label",
                   "MMK_Sig_p05", "MMK_Sig_p01",
                   "MMK_Sig_Z05",  "MMK_Sig_Z01", "MMK_Sig_Consistency"]]

    def build_21_mmk_vs_pw(self) -> pd.DataFrame:
        """
        21_MMK_vs_PW — 36 rows
        Modified MK (S2) vs PW-MK (S7).  Delta_Z = PW_Z − MMK_Z.
        Includes: n_eff, Correction_Factor, Lag.
        """
        mmk = self._mmk_ref_cols()
        s7  = self._s7()
        ac  = self._ac_block()
        s1_ctx = self._s1()[["Station", "Scale", "rho_1"]].copy()

        df = (
            mmk
            .merge(s7[["Station", "Scale", "Sig_AC",
                        "PW_Z", "PW_p", "PW_slope", "PW_sig", "PW_trend"]],
                   on=["Station", "Scale"])
            .merge(ac,     on=["Station", "Scale"], how="left")
            .merge(s1_ctx, on=["Station", "Scale"], how="left")
        )
        s7_ids = self._s7()[["Station", "Code", "Scale"]].copy()
        df = s7_ids.merge(df, on=["Station", "Scale"], how="left")

        df = df.rename(columns={"PW_sig": "PW_Sig_label", "PW_trend": "PW_Trend"})
        df = _add_sig(df, "PW_Z", "PW_p", px="PW")
        df = _add_delta(df,
                        "MMK_Z", "MMK_p", "MMK_Trend", "MMK_Sig_Z05",
                        "PW_Z",  "PW_p",  "PW_Trend",  "PW_Sig_Z05")
        return _sort(df[[
            "Station", "Code", "Scale",
            "rho_1", "Sig_AC", "n_eff", "Correction_Factor", "Lag",
            "MMK_Z", "MMK_p", "MMK_Trend", "MMK_Sig_label",
            "MMK_Sig_p05", "MMK_Sig_p01", "MMK_Sig_Z05", "MMK_Sig_Z01", "MMK_Sig_Consistency",
            "PW_Z",  "PW_p",  "PW_Trend",  "PW_Sig_label",
            "PW_Sig_p05",  "PW_Sig_p01",  "PW_Sig_Z05",  "PW_Sig_Z01",  "PW_Sig_Consistency",
            "Delta_Z", "Delta_p", "Direction_Changed", "Significance_Changed",
        ]])

    def build_22_mmk_vs_tfpw(self) -> pd.DataFrame:
        """
        22_MMK_vs_TFPW — 36 rows
        Modified MK (S2) vs TFPW-MK (S7).  Delta_Z = TFPW_Z − MMK_Z.
        """
        mmk = self._mmk_ref_cols()
        s7  = self._s7()
        ac  = self._ac_block()
        s1_ctx = self._s1()[["Station", "Scale", "rho_1"]].copy()
        s7_ids = s7[["Station", "Code", "Scale"]].copy()

        df = (
            mmk
            .merge(s7[["Station", "Scale", "Sig_AC",
                        "TFPW_Z", "TFPW_p", "TFPW_slope", "TFPW_sig", "TFPW_trend"]],
                   on=["Station", "Scale"])
            .merge(ac,     on=["Station", "Scale"], how="left")
            .merge(s1_ctx, on=["Station", "Scale"], how="left")
        )
        df = s7_ids.merge(df, on=["Station", "Scale"], how="left")

        df = df.rename(columns={"TFPW_sig": "TFPW_Sig_label", "TFPW_trend": "TFPW_Trend"})
        df = _add_sig(df, "TFPW_Z", "TFPW_p", px="TFPW")
        df = _add_delta(df,
                        "MMK_Z",  "MMK_p",  "MMK_Trend",  "MMK_Sig_Z05",
                        "TFPW_Z", "TFPW_p", "TFPW_Trend", "TFPW_Sig_Z05")
        return _sort(df[[
            "Station", "Code", "Scale",
            "rho_1", "Sig_AC", "n_eff", "Correction_Factor", "Lag",
            "MMK_Z",  "MMK_p",  "MMK_Trend",  "MMK_Sig_label",
            "MMK_Sig_p05",  "MMK_Sig_p01",  "MMK_Sig_Z05",  "MMK_Sig_Z01",  "MMK_Sig_Consistency",
            "TFPW_Z", "TFPW_p", "TFPW_Trend", "TFPW_Sig_label",
            "TFPW_Sig_p05", "TFPW_Sig_p01", "TFPW_Sig_Z05", "TFPW_Sig_Z01", "TFPW_Sig_Consistency",
            "Delta_Z", "Delta_p", "Direction_Changed", "Significance_Changed",
        ]])

    def build_23_pw_vs_tfpw(self) -> pd.DataFrame:
        """
        23_PW_vs_TFPW — 36 rows
        PW-MK vs TFPW-MK (both from S7).  Delta_Z = TFPW_Z − PW_Z.
        """
        s7  = self._s7()
        s1_ctx = self._s1()[["Station", "Scale", "rho_1"]].copy()
        ac  = self._ac_block()

        df = (
            s7[["Station", "Code", "Scale", "Sig_AC",
                "PW_Z", "PW_p", "PW_slope", "PW_sig", "PW_trend",
                "TFPW_Z", "TFPW_p", "TFPW_slope", "TFPW_sig", "TFPW_trend"]]
            .merge(s1_ctx, on=["Station", "Scale"], how="left")
            .merge(ac,     on=["Station", "Scale"], how="left")
        )
        df = df.rename(columns={
            "PW_sig":     "PW_Sig_label",   "PW_trend":   "PW_Trend",
            "TFPW_sig":   "TFPW_Sig_label", "TFPW_trend": "TFPW_Trend",
        })
        df = _add_sig(df, "PW_Z",   "PW_p",   px="PW")
        df = _add_sig(df, "TFPW_Z", "TFPW_p", px="TFPW")
        df = _add_delta(df,
                        "PW_Z",   "PW_p",   "PW_Trend",   "PW_Sig_Z05",
                        "TFPW_Z", "TFPW_p", "TFPW_Trend", "TFPW_Sig_Z05")
        return _sort(df[[
            "Station", "Code", "Scale",
            "rho_1", "Sig_AC",
            "PW_Z",   "PW_p",   "PW_Trend",   "PW_Sig_label",
            "PW_Sig_p05",   "PW_Sig_p01",   "PW_Sig_Z05",   "PW_Sig_Z01",   "PW_Sig_Consistency",
            "TFPW_Z", "TFPW_p", "TFPW_Trend", "TFPW_Sig_label",
            "TFPW_Sig_p05", "TFPW_Sig_p01", "TFPW_Sig_Z05", "TFPW_Sig_Z01", "TFPW_Sig_Consistency",
            "Delta_Z", "Delta_p", "Direction_Changed", "Significance_Changed",
        ]])

    # ═════════════════════════════════════════════════════════════════════════
    #  D — Publication tables
    # ═════════════════════════════════════════════════════════════════════════

    def _master_core(self) -> pd.DataFrame:
        """Shared base for D+E group: S7 + S2 (n_eff, CF) + enriched."""
        s7 = self._s7()
        ac = self._ac_block()
        df = s7.merge(ac, on=["Station", "Scale"], how="left")

        z_cols    = ["MK_Z", "MMK_Z", "PW_Z", "TFPW_Z"]
        sig_cols  = ["MK_sig", "MMK_sig", "PW_sig", "TFPW_sig"]
        tr_cols   = ["MK_trend", "MMK_trend", "PW_trend", "TFPW_trend"]

        df["Z_spread"]   = (df[z_cols].max(axis=1) - df[z_cols].min(axis=1)).round(4)
        df["Z_mean"]     = df[z_cols].mean(axis=1).round(4)
        df["n_sig_Z05"]  = (df[z_cols].abs() >= _Z_05).sum(axis=1)
        df["n_sig_Z01"]  = (df[z_cols].abs() >= _Z_01).sum(axis=1)
        df["n_inc"]      = df[tr_cols].apply(lambda r: (r == "Increasing ↑").sum(), axis=1)
        df["n_dec"]      = df[tr_cols].apply(lambda r: (r == "Decreasing ↓").sum(), axis=1)
        df["Direction_Consensus"] = (df["n_inc"] == 4) | (df["n_dec"] == 4)
        df["Robustness"] = df["n_sig_Z05"].map({
            4: "Very robust (4/4)",
            3: "Robust (3/4)",
            2: "Moderate (2/4)",
            1: "Weak (1/4)",
            0: "Not significant (0/4)",
        })
        # Pairwise direction agreement flags
        for a, b in [("MK", "MMK"), ("MK", "PW"), ("MK", "TFPW"),
                     ("MMK", "PW"), ("MMK", "TFPW"), ("PW", "TFPW")]:
            df[f"Dir_{a}_{b}"] = df[f"{a}_trend"] == df[f"{b}_trend"]
        # Pairwise significance agreement flags
        s7_sig_z = {}
        for m in ["MK", "MMK", "PW", "TFPW"]:
            col = f"{m}_sig" if m in ("MK", "MMK", "PW", "TFPW") else f"{m}_sig"
            s7_sig_z[m] = df[f"{m}_Z"].abs() >= _Z_05
        for a, b in [("MK", "MMK"), ("MK", "PW"), ("MK", "TFPW"),
                     ("MMK", "PW"), ("MMK", "TFPW"), ("PW", "TFPW")]:
            df[f"Sig_{a}_{b}"] = s7_sig_z[a] == s7_sig_z[b]
        df["N_sig_pairs"] = sum(
            df[f"Sig_{a}_{b}"].astype(int)
            for a, b in [("MK","MMK"),("MK","PW"),("MK","TFPW"),
                         ("MMK","PW"),("MMK","TFPW"),("PW","TFPW")]
        )
        return df

    def build_31_sig_stations(self) -> pd.DataFrame:
        """
        31_Significant_Stations — ≤36 rows
        Stations significant (Sig_Z05) by at least one method.
        """
        df = self._master_core()
        filt = df["n_sig_Z05"] >= 1
        cols = [
            "Station", "Code", "Scale",
            "rho_1", "Sig_AC", "n_eff", "Correction_Factor", "Lag",
            "MK_Z",  "MK_p",  "MK_sig",  "MK_trend",
            "MMK_Z", "MMK_p", "MMK_sig", "MMK_trend",
            "PW_Z",  "PW_p",  "PW_sig",  "PW_trend",
            "TFPW_Z","TFPW_p","TFPW_sig","TFPW_trend",
            "n_sig_Z05", "n_sig_Z01", "Robustness", "Direction_Consensus",
        ]
        return _sort(df.loc[filt, cols].copy())

    def build_32_method_agreement(self) -> pd.DataFrame:
        """
        32_Method_Agreement — 36 rows
        Pairwise direction and significance agreement for all 6 method pairs.
        N_dir_pairs_agree: how many of 6 direction pairs agree.
        N_sig_pairs_agree: how many of 6 significance pairs agree.
        """
        df = self._master_core()
        dir_cols = [f"Dir_{a}_{b}" for a, b in
                    [("MK","MMK"),("MK","PW"),("MK","TFPW"),
                     ("MMK","PW"),("MMK","TFPW"),("PW","TFPW")]]
        sig_cols = [f"Sig_{a}_{b}" for a, b in
                    [("MK","MMK"),("MK","PW"),("MK","TFPW"),
                     ("MMK","PW"),("MMK","TFPW"),("PW","TFPW")]]
        df["N_dir_pairs_agree"] = df[dir_cols].sum(axis=1)
        df["N_sig_pairs_agree"] = df[sig_cols].sum(axis=1)
        df["All_dir_agree"]  = df["N_dir_pairs_agree"] == 6
        df["All_sig_agree"]  = df["N_sig_pairs_agree"] == 6
        return _sort(df[[
            "Station", "Code", "Scale",
            "rho_1", "Sig_AC", "n_eff", "Correction_Factor",
            "MK_Z",   "MK_sig",   "MK_trend",
            "MMK_Z",  "MMK_sig",  "MMK_trend",
            "PW_Z",   "PW_sig",   "PW_trend",
            "TFPW_Z", "TFPW_sig", "TFPW_trend",
            *dir_cols,
            *sig_cols,
            "N_dir_pairs_agree", "N_sig_pairs_agree",
            "All_dir_agree", "All_sig_agree",
        ]])

    def build_33_autocorr_impact(self) -> pd.DataFrame:
        """
        33_Autocorrelation_Impact — 36 rows
        Effect of autocorrelation on MK conclusions.
        MK_Reliable = False when Sig_AC='Yes' AND MK_Sig_Z05 != MMK_Sig_Z05.
        """
        s1 = self._s1().rename(columns={"Z": "MK_Z", "p-value": "MK_p",
                                          "τ (Kendall)": "MK_Tau",
                                          "Trend": "MK_Trend"})
        s2 = self._s2().rename(columns={"Z": "MMK_Z", "p-value": "MMK_p",
                                          "Var*(S)": "VarS_adj"})
        s7 = self._s7()
        ac = self._ac_block()

        df = (
            s1[["Station", "Code", "Scale", "rho_1",
                "MK_Z", "MK_p", "MK_Tau", "MK_Trend"]]
            .merge(s2[["Station", "Scale", "MMK_Z", "MMK_p", "VarS_adj"]], on=["Station", "Scale"])
            .merge(s7[["Station", "Scale", "Sig_AC",
                        "PW_Z", "PW_p", "TFPW_Z", "TFPW_p",
                        "dZ_MMK", "dZ_PW", "dZ_TFPW"]], on=["Station", "Scale"])
            .merge(ac, on=["Station", "Scale"], how="left")
        )
        df = _add_sig(df, "MK_Z",   "MK_p",   px="MK")
        df = _add_sig(df, "MMK_Z",  "MMK_p",  px="MMK")
        df = _add_sig(df, "PW_Z",   "PW_p",   px="PW")
        df = _add_sig(df, "TFPW_Z", "TFPW_p", px="TFPW")

        df["MK_Reliable"] = ~(
            (df["Sig_AC"] == "Yes") & (df["MK_Sig_Z05"] != df["MMK_Sig_Z05"])
        )
        df["MK_Reliability"] = df["MK_Reliable"].map({
            True:  "Reliable — AC does not change conclusion",
            False: "Unreliable — AC changes conclusion; use MMK",
        })
        df["AC_Bias"] = df["dZ_MMK"].apply(
            lambda d: "AC inflates MK significance" if d < -0.05 else
                      "AC deflates MK significance" if d >  0.05 else
                      "AC bias negligible"
        )
        df["Recommended_Method"] = df["Sig_AC"].apply(
            lambda a: "MMK — significant autocorrelation present" if a == "Yes"
                      else "MK — no significant autocorrelation"
        )
        return _sort(df[[
            "Station", "Code", "Scale",
            "rho_1", "Sig_AC", "n_eff", "Correction_Factor", "Lag",
            "MK_Z",   "MK_p",   "MK_Sig_Z05",  "MK_Sig_Consistency",
            "MMK_Z",  "MMK_p",  "MMK_Sig_Z05", "MMK_Sig_Consistency",
            "PW_Z",   "PW_p",   "PW_Sig_Z05",
            "TFPW_Z", "TFPW_p", "TFPW_Sig_Z05",
            "dZ_MMK", "dZ_PW", "dZ_TFPW",
            "AC_Bias", "MK_Reliable", "MK_Reliability", "Recommended_Method",
        ]])

    def build_34_field_significance(self) -> pd.DataFrame:
        """
        34_Field_Significance — 3 rows (one per scale).
        S8 (Walker + LC) extended with PW / TFPW N_sig counts from S7.
        """
        s8 = self._s8().copy()
        s7 = self._s7()

        for scale in _SCALE_ORDER:
            sub = s7[s7["Scale"] == scale]
            n   = max(len(sub), 1)
            idx = s8.index[s8["Scale"] == scale]
            if len(idx) == 0:
                continue
            i = idx[0]
            s8.loc[i, "N_sig_PW"]      = int(sub["PW_sig"].isin(["*", "**"]).sum())
            s8.loc[i, "N_sig_TFPW"]    = int(sub["TFPW_sig"].isin(["*", "**"]).sum())
            s8.loc[i, "Frac_sig_PW"]   = round(int(s8.loc[i, "N_sig_PW"]) / n, 4)
            s8.loc[i, "Frac_sig_TFPW"] = round(int(s8.loc[i, "N_sig_TFPW"]) / n, 4)

        return s8[[
            "Scale", "N_stations",
            "N_sig_MK", "N_sig_MMK", "N_sig_PW", "N_sig_TFPW",
            "Frac_sig_MK", "Frac_sig_MMK", "Frac_sig_PW", "Frac_sig_TFPW",
            "Walker_p_MK", "Walker_sig_MK",
            "LC_p_MK",     "LC_sig_MK",
            "LC_null_mean", "LC_null_95th",
        ]]

    def build_35_method_summary(self) -> pd.DataFrame:
        """
        35_Method_Summary — 3 rows (one per scale).
        Aggregated: N_sig, Frac_sig, mean |Z|, mean slope for all 4 methods.
        """
        s7 = self._s7()
        ac = self._ac_block()
        df = s7.merge(ac, on=["Station", "Scale"], how="left")

        rows = []
        for scale in _SCALE_ORDER:
            sub = df[df["Scale"] == scale]
            n   = len(sub)
            row: dict = {"Scale": scale, "N_stations": n}
            for m, zc, pc, slc, sc in [
                ("MK",   "MK_Z",   "MK_p",   "MK_slope",   "MK_sig"),
                ("MMK",  "MMK_Z",  "MMK_p",  "MMK_slope",  "MMK_sig"),
                ("PW",   "PW_Z",   "PW_p",   "PW_slope",   "PW_sig"),
                ("TFPW", "TFPW_Z", "TFPW_p", "TFPW_slope", "TFPW_sig"),
            ]:
                row[f"N_sig_Z05_{m}"]    = int((sub[zc].abs() >= _Z_05).sum())
                row[f"N_sig_Z01_{m}"]    = int((sub[zc].abs() >= _Z_01).sum())
                row[f"N_sig_p05_{m}"]    = int((sub[pc] < 0.05).sum())
                row[f"Frac_sig_Z05_{m}"] = round(row[f"N_sig_Z05_{m}"] / n, 4)
                row[f"Mean_absZ_{m}"]    = round(sub[zc].abs().mean(), 4)
                row[f"Mean_Slope_{m}"]   = round(sub[slc].mean(), 4)
            row["N_all4_agree"]  = int(sub["all_agree"].eq("Yes").sum())
            row["N_any_sig_Z05"] = int((sub[["MK_Z","MMK_Z","PW_Z","TFPW_Z"]].abs() >= _Z_05).any(axis=1).sum())
            row["N_none_sig"]    = int((sub[["MK_Z","MMK_Z","PW_Z","TFPW_Z"]].abs() < _Z_05).all(axis=1).sum())
            row["Mean_dZ_MMK"]   = round(sub["dZ_MMK"].mean(), 4)
            row["Mean_dZ_PW"]    = round(sub["dZ_PW"].mean(), 4)
            row["Mean_dZ_TFPW"]  = round(sub["dZ_TFPW"].mean(), 4)
            row["Max_abs_dZ_MMK"]= round(sub["dZ_MMK"].abs().max(), 4)
            row["Max_abs_dZ_PW"] = round(sub["dZ_PW"].abs().max(), 4)
            row["Max_abs_dZ_TFPW"]= round(sub["dZ_TFPW"].abs().max(), 4)
            row["Mean_n_eff"]    = round(sub["n_eff"].mean(), 2)
            row["Mean_CF"]       = round(sub["Correction_Factor"].mean(), 4)
            rows.append(row)

        return pd.DataFrame(rows)

    # ═════════════════════════════════════════════════════════════════════════
    #  E — Master station database
    # ═════════════════════════════════════════════════════════════════════════

    def build_90_master(self) -> pd.DataFrame:
        """
        90_Master_Station_Database — 36 rows
        All 4 method Z/p/slope/sig/trend + AC block + Sen slope (MK CI) +
        all enriched derivations.  Primary cross-reference table.
        """
        df = self._master_core()

        # Add sig columns for all 4 methods
        for m, zc, pc in [("MK","MK_Z","MK_p"),("MMK","MMK_Z","MMK_p"),
                           ("PW","PW_Z","PW_p"),("TFPW","TFPW_Z","TFPW_p")]:
            df = _add_sig(df, zc, pc, px=m)

        # Attach Sen slope (Standard MK method) for quick reference
        s4_mk = (
            self._s4()
            .query("Method == 'Standard MK'")
            .rename(columns={
                "β (mm/yr)":        "Sen_Slope",
                "CI_Lower (mm/yr)": "Sen_CI_Lower",
                "CI_Upper (mm/yr)": "Sen_CI_Upper",
            })
        )[["Station", "Scale", "Sen_Slope", "Sen_CI_Lower", "Sen_CI_Upper"]]
        df = df.merge(s4_mk, on=["Station", "Scale"], how="left")

        cols = [
            "Station", "Code", "Scale",
            "rho_1", "Sig_AC", "n_eff", "Correction_Factor", "Lag",
            # MK
            "MK_Z", "MK_p", "MK_slope", "MK_sig", "MK_trend",
            "MK_Sig_p05", "MK_Sig_p01", "MK_Sig_Z05", "MK_Sig_Z01", "MK_Sig_Consistency",
            # MMK
            "MMK_Z", "MMK_p", "MMK_slope", "MMK_sig", "MMK_trend",
            "MMK_Sig_p05","MMK_Sig_p01","MMK_Sig_Z05","MMK_Sig_Z01","MMK_Sig_Consistency",
            # PW
            "PW_Z", "PW_p", "PW_slope", "PW_sig", "PW_trend",
            "PW_Sig_p05", "PW_Sig_p01", "PW_Sig_Z05", "PW_Sig_Z01", "PW_Sig_Consistency",
            # TFPW
            "TFPW_Z", "TFPW_p", "TFPW_slope", "TFPW_sig", "TFPW_trend",
            "TFPW_Sig_p05","TFPW_Sig_p01","TFPW_Sig_Z05","TFPW_Sig_Z01","TFPW_Sig_Consistency",
            # Pre-computed deltas (MK baseline)
            "dZ_MMK", "dZ_PW", "dZ_TFPW",
            "dSlope_MMK", "dSlope_PW", "dSlope_TFPW",
            # Derived statistics
            "Z_spread", "Z_mean",
            "n_sig_Z05", "n_sig_Z01",
            "n_inc", "n_dec", "Direction_Consensus",
            "all_agree", "n_sig_methods", "N_sig_pairs", "Robustness",
            # Sen slope reference (Standard MK method)
            "Sen_Slope", "Sen_CI_Lower", "Sen_CI_Upper",
        ]
        return _sort(df[cols])

    # ═════════════════════════════════════════════════════════════════════════
    #  Supplementary sheets
    # ═════════════════════════════════════════════════════════════════════════

    def build_91_pettitt(self) -> pd.DataFrame | None:
        """91_Pettitt_CP — 36 rows; WB4 only (absent from canonical V4 pipeline)."""
        if self.wb4 is None:
            return None
        df = _read_wb4(self.wb4, _WB4_PETTITT)
        df["Scale"] = df["Scale"].map(_SCALE_LONG).fillna(df["Scale"])
        s7_ids = self._s7()[["Station", "Code", "Scale"]].copy()
        df = s7_ids.merge(df, on=["Station", "Scale"], how="left")
        return _sort(df[["Station", "Code", "Scale",
                          "Change_Point_Year", "Pettitt_p", "Homogeneity"]])

    def build_99_metadata(self) -> pd.DataFrame:
        """99_Metadata — source map and generation info."""
        rows = [
            ("01_MK_Results",              "A",36,  "WB1","S1",                  "Full MK: N,S,Var(S),rho_1,Z,τ,p + 5 sig cols"),
            ("02_MMK_Results",             "A",36,  "WB1+WB4","S2+3_Modified_MK","Full MMK + n_eff, CF, Lag parsed from WB4"),
            ("03_PW_Results",              "A",36,  "WB1","S7+S1",               "PW-MK (only in S7); rho_1 + AC block as context"),
            ("04_TFPW_Results",            "A",36,  "WB1","S7+S1",               "TFPW-MK (only in S7); rho_1 + AC block as context"),
            ("05_SenSlope_Results",        "A",144, "WB1","S4",                  "All 4 methods × 3 scales with 95% CI (12×3×4=144)"),
            ("11_MK_vs_MMK",               "B",36,  "WB1","S1+S2",               "MK vs AC-corrected MMK; Delta_Z/p, n_eff, CF, Lag"),
            ("12_MK_vs_PW",                "B",36,  "WB1","S1+S7",               "MK vs PW prewhitening; Delta_Z/p, conclusion flags"),
            ("13_MK_vs_TFPW",              "B",36,  "WB1","S1+S7",               "MK vs TFPW trend-free prewhitening"),
            ("21_MMK_vs_PW",               "C",36,  "WB1","S2+S7",               "MMK vs PW; Delta_Z = PW_Z − MMK_Z"),
            ("22_MMK_vs_TFPW",             "C",36,  "WB1","S2+S7",               "MMK vs TFPW; Delta_Z = TFPW_Z − MMK_Z"),
            ("23_PW_vs_TFPW",              "C",36,  "WB1","S7",                  "PW vs TFPW; Delta_Z = TFPW_Z − PW_Z"),
            ("31_Significant_Stations",    "D","≤36","WB1","S7+S2 filtered",     "n_sig_Z05 ≥ 1; any method significant at α=0.05"),
            ("32_Method_Agreement",        "D",36,  "WB1","S7+S2",               "6 pairwise direction + significance agreement flags"),
            ("33_Autocorrelation_Impact",  "D",36,  "WB1","S1+S2+S7",            "AC bias direction, MK reliability, recommended method"),
            ("34_Field_Significance",      "D",3,   "WB1","S8+S7",               "Walker(1914)+LC(1983); PW/TFPW counts added"),
            ("35_Method_Summary",          "D",3,   "WB1","S7+S2",               "Per-scale: N_sig, mean|Z|, mean slope; dZ stats"),
            ("90_Master_Station_Database", "E",36,  "WB1","S7+S2+S4",            "All methods + all sig cols + enriched derivations"),
            ("91_Pettitt_CP",              "S",36,  "WB4","4_Pettitt_CP",        "Change-point year + p-value; absent from WB1"),
            ("99_Metadata",                "S","-", "—","—",                     "Source map, generation date, scale note"),
        ]
        df = pd.DataFrame(rows, columns=[
            "Sheet", "Group", "Rows",
            "Source_Workbook", "Source_Sheet", "Notes",
        ])
        df.insert(0, "Generated",   datetime.now().strftime("%Y-%m-%d %H:%M"))
        df["WB1_path"]   = str(self.wb1)
        df["WB4_path"]   = str(self.wb4) if self.wb4 else "Not provided"
        df["Scale_note"] = (
            "Available: Annual (Jan–Dec), Wet Season (May–Oct), Dry Season (Nov–Apr). "
            "Monthly: NOT available. Seasonal (4-split): NOT available."
        )
        return df

    # ═════════════════════════════════════════════════════════════════════════
    #  Workbook writer
    # ═════════════════════════════════════════════════════════════════════════

    def write_workbook(self, out_path: str | Path) -> None:
        """
        Build all sheets and write Trend_Method_Comparison_Q1.xlsx.

        Raises FileExistsError if the target file already exists (never overwrites).
        """
        out = Path(out_path)
        if out.exists():
            raise FileExistsError(
                f"Output already exists: {out}\n"
                "Remove the existing file or choose a different path."
            )
        out.parent.mkdir(parents=True, exist_ok=True)

        print("Building all sheets …")
        pettitt = self.build_91_pettitt()

        # (name, DataFrame|None, group)
        plan: list[tuple[str, pd.DataFrame | None, str]] = [
            ("01_MK_Results",              self.build_01_mk(),              "A"),
            ("02_MMK_Results",             self.build_02_mmk(),             "A"),
            ("03_PW_Results",              self.build_03_pw(),              "A"),
            ("04_TFPW_Results",            self.build_04_tfpw(),            "A"),
            ("05_SenSlope_Results",        self.build_05_senslope(),        "A"),
            ("11_MK_vs_MMK",               self.build_11_mk_vs_mmk(),       "B"),
            ("12_MK_vs_PW",                self.build_12_mk_vs_pw(),        "B"),
            ("13_MK_vs_TFPW",              self.build_13_mk_vs_tfpw(),      "B"),
            ("21_MMK_vs_PW",               self.build_21_mmk_vs_pw(),       "C"),
            ("22_MMK_vs_TFPW",             self.build_22_mmk_vs_tfpw(),     "C"),
            ("23_PW_vs_TFPW",              self.build_23_pw_vs_tfpw(),      "C"),
            ("31_Significant_Stations",    self.build_31_sig_stations(),    "D"),
            ("32_Method_Agreement",        self.build_32_method_agreement(),"D"),
            ("33_Autocorrelation_Impact",  self.build_33_autocorr_impact(), "D"),
            ("34_Field_Significance",      self.build_34_field_significance(),"D"),
            ("35_Method_Summary",          self.build_35_method_summary(),  "D"),
            ("90_Master_Station_Database", self.build_90_master(),          "E"),
            ("91_Pettitt_CP",              pettitt,                         "S"),
            ("99_Metadata",                self.build_99_metadata(),        "S"),
        ]

        print(f"Writing {len(plan)} sheets → {out.name} …")
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            for name, df, _grp in plan:
                if df is None:
                    pd.DataFrame({"Note": [
                        "WB4 not provided — Pettitt CP data unavailable."
                    ]}).to_excel(writer, sheet_name=name, index=False)
                else:
                    df.to_excel(writer, sheet_name=name, index=False)

        _style_workbook(out, plan)

        print(f"\n{'='*60}")
        print(f"  Written: {out}")
        print(f"  Sheets : {len(plan)}")
        for name, df, grp in plan:
            n = len(df) if df is not None else 0
            print(f"  [{grp}]  {name:<34s}  {n:>4} rows")
        print("="*60)


# ══════════════════════════════════════════════════════════════════════════════
#  Excel styling
# ══════════════════════════════════════════════════════════════════════════════

def _style_workbook(path: Path, plan: list) -> None:
    """Apply tab colours, header fill, column widths, and freeze panes."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = load_workbook(path)
    thin_border = Border(bottom=Side(style="thin", color="BFBFBF"))
    alt_fill    = PatternFill("solid", fgColor="F5F5F5")
    data_font   = Font(size=9)

    for name, _df, group in plan:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        tab_hex, hdr_hex = _SHEET_GROUPS.get(group, _SHEET_GROUPS["S"])

        ws.sheet_properties.tabColor = tab_hex

        # Header row
        hdr_fill = PatternFill("solid", fgColor=hdr_hex)
        hdr_font = Font(bold=True, color="111111", size=9)
        for cell in ws[1]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = thin_border

        # Data rows
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = alt_fill if i % 2 == 0 else PatternFill()
            for cell in row:
                cell.font      = data_font
                cell.fill      = fill
                cell.alignment = Alignment(horizontal="left")

        # Column widths (cap at 32)
        for col in ws.columns:
            hdr_len = len(str(col[0].value or ""))
            max_len = max(
                hdr_len,
                *(len(str(c.value or "")) for c in col[1:]),
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 32)

        ws.freeze_panes = "A2"

    wb.save(path)
