"""
Generates Trend_Method_Comparison_All_vs_MK.xlsx

Reads from:
  results/final_N33_v5/Trend_Method_Comparison/Excel/Master/
      Trend_Method_Comparison_Master.xlsx  (Master_DB sheet)

Writes to:
  results/final_N33_v5/Trend_Method_Comparison/Excel/Master/
      Trend_Method_Comparison_All_vs_MK.xlsx  (NEW — does not overwrite anything)

Sheets
------
  MMK_vs_MK_Station      36 rows — station-level MMK vs MK
  PW_vs_MK_Station       36 rows — station-level PW vs MK
  TFPW_vs_MK_Station     36 rows — station-level TFPW vs MK
  All_Scale_Summary       9 rows — 3 methods × 3 scales summary
  All_Manuscript_Table   12 rows — wide format, all methods per station
  Significant_Changes    variable — rows where significance changed
  Direction_Changes      variable — rows where direction changed
  Method_Ranking          4 rows — ranked by conservativeness
  Reviewer_Summary       variable — labelled multi-section summary

Does NOT modify any existing workbook.
"""
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT   = Path(__file__).parent
MASTER = ROOT / "results/final_N33_v5/Trend_Method_Comparison/Excel/Master"
OUT    = MASTER / "Trend_Method_Comparison_All_vs_MK.xlsx"

_SCALE_ORDER = ["Annual (Jan–Dec)", "Wet Season (May–Oct)", "Dry Season (Nov–Apr)"]
_SCALE_ABBREV = {
    "Annual (Jan–Dec)":     "Annual",
    "Wet Season (May–Oct)": "Wet",
    "Dry Season (Nov–Apr)": "Dry",
}
_Z05 = 1.9600
_Z01 = 2.5758

_ALT_META = {
    "MMK":  "Modified MK (H&R98)",
    "PW":   "PW-MK",
    "TFPW": "TFPW-MK",
}

# 5-category change taxonomy
_CAT_NO_CHANGE   = "No Change"
_CAT_SIG_TO_NS   = "Significant → Non-significant"
_CAT_NS_TO_SIG   = "Non-significant → Significant"
_CAT_INC_TO_DEC  = "Increase → Decrease"
_CAT_DEC_TO_INC  = "Decrease → Increase"

# Excel styling constants
_TH = Side(style="thin",   color="CCCCCC")
_MB = Side(style="medium", color="555555")
_B  = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)

_COLORS = {
    "MMK":   ("0277BD", "E3F2FD"),
    "PW":    ("00897B", "E0F2F1"),
    "TFPW":  ("E65100", "FBE9E7"),
    "COMP":  ("37474F", "ECEFF1"),
    "CHNG":  ("AD1457", "FCE4EC"),
    "DIR":   ("6A1B9A", "EDE7F6"),
    "RANK":  ("1B5E20", "E8F5E9"),
    "REV":   ("4E342E", "FBE9E7"),
}


# ── Excel styling ─────────────────────────────────────────────────────────────

def _style_ws(ws, header_hex: str, alt_hex: str) -> None:
    hf   = PatternFill("solid", fgColor=header_hex)
    af   = PatternFill("solid", fgColor=alt_hex)
    wf   = PatternFill("solid", fgColor="FFFFFF")
    dark = int(header_hex[:2], 16) < 200
    hfont = Font(bold=True, size=9, color="FFFFFF" if dark else "000000")
    dfont = Font(size=9)
    for cell in ws[1]:
        cell.fill      = hf
        cell.font      = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = _B
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = af if i % 2 == 0 else wf
        for cell in row:
            cell.fill      = fill
            cell.font      = dfont
            cell.alignment = Alignment(horizontal="center")
            cell.border    = _B
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=4)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, 42)
    ws.freeze_panes = "A2"


def _write_excel(sheets: dict, path: Path, color_map: dict | None = None,
                 default_h: str = "546E7A", default_a: str = "ECEFF1") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
    wb = load_workbook(path)
    for ws in wb.worksheets:
        h, a = (color_map or {}).get(ws.title, (default_h, default_a))
        _style_ws(ws, h, a)
    wb.save(path)


# ── Data loading ──────────────────────────────────────────────────────────────

def load_master() -> pd.DataFrame:
    src = MASTER / "Trend_Method_Comparison_Master.xlsx"
    if not src.exists():
        raise FileNotFoundError(f"Master workbook not found: {src}")
    df = pd.read_excel(src, sheet_name="Master_DB", header=0)
    if "Station" in df.columns:
        df["Station"] = df["Station"].astype(str).str.strip()
    return df


# ── Change category helper ────────────────────────────────────────────────────

def _change_category(mk_sig: bool, alt_sig: bool,
                     mk_trend: str, alt_trend: str) -> str:
    """Assign one of the five explicit change categories."""
    # Direction changes take priority
    mk_inc = "Increasing" in str(mk_trend)
    mk_dec = "Decreasing" in str(mk_trend)
    alt_inc = "Increasing" in str(alt_trend)
    alt_dec = "Decreasing" in str(alt_trend)

    if mk_inc and alt_dec:
        return _CAT_INC_TO_DEC
    if mk_dec and alt_inc:
        return _CAT_DEC_TO_INC

    # Significance changes (direction same)
    if mk_sig and not alt_sig:
        return _CAT_SIG_TO_NS
    if not mk_sig and alt_sig:
        return _CAT_NS_TO_SIG

    return _CAT_NO_CHANGE


def _sort_scale(df: pd.DataFrame) -> pd.DataFrame:
    order = {s: i for i, s in enumerate(_SCALE_ORDER)}
    df = df.copy()
    df["_k"] = df["Scale"].map(order).fillna(9)
    df = df.sort_values(["_k", "Station"]).drop(columns=["_k"])
    return df.reset_index(drop=True)


# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_station_sheet(master: pd.DataFrame, alt: str) -> pd.DataFrame:
    """36-row station-level comparison for one alternative method vs MK."""
    rows = []
    for _, r in master.iterrows():
        alt_z    = float(r[f"{alt}_Z"])
        mk_z     = float(r["MK_Z"])
        alt_sig  = abs(alt_z) >= _Z05
        mk_sig   = abs(mk_z)  >= _Z05
        alt_trend = str(r[f"{alt}_trend"])
        mk_trend  = str(r["MK_trend"])

        rows.append({
            "Station":            r["Station"],
            "Code":               r["Code"],
            "Scale":              r["Scale"],
            "MK_Z":               round(mk_z, 4),
            "MK_p":               round(float(r["MK_p"]), 6),
            "MK_Significant":     mk_sig,
            "MK_Trend":           mk_trend,
            "Alternative_Method": _ALT_META[alt],
            "Alt_Z":              round(alt_z, 4),
            "Alt_p":              round(float(r[f"{alt}_p"]), 6),
            "Alt_Significant":    alt_sig,
            "Alt_Trend":          alt_trend,
            "Delta_Z":            round(alt_z - mk_z, 6),
            "Delta_p":            round(float(r[f"{alt}_p"]) - float(r["MK_p"]), 6),
            "Significance_Changed": mk_sig != alt_sig,
            "Direction_Changed":    mk_trend != alt_trend,
            "Change_Category":      _change_category(mk_sig, alt_sig,
                                                      mk_trend, alt_trend),
        })

    return _sort_scale(pd.DataFrame(rows))


def build_all_scale_summary(station_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """9 rows (3 methods × 3 scales) with per-method summary statistics."""
    rows = []
    for alt, df in station_sheets.items():
        for scale in _SCALE_ORDER:
            sub = df[df["Scale"] == scale]
            n         = len(sub)
            n_mk_sig  = int(sub["MK_Significant"].sum())
            n_alt_sig = int(sub["Alt_Significant"].sum())
            n_sig_chg = int(sub["Significance_Changed"].sum())
            n_dir_chg = int(sub["Direction_Changed"].sum())
            n_full_agr = int(
                (~sub["Significance_Changed"] & ~sub["Direction_Changed"]).sum()
            )
            # 5-category counts
            cats = sub["Change_Category"].value_counts().to_dict()
            rows.append({
                "Alternative_Method":              _ALT_META[alt],
                "Scale":                           scale,
                "N_stations":                      n,
                "N_MK_significant":                n_mk_sig,
                "N_Alt_significant":               n_alt_sig,
                "Number_of_changed_significance":  n_sig_chg,
                "Number_of_changed_direction":     n_dir_chg,
                "Agreement_rate_with_MK":          round(n_full_agr / n, 4) if n else None,
                "Mean_abs_Delta_Z":                round(sub["Delta_Z"].abs().mean(), 4),
                "Max_abs_Delta_Z":                 round(sub["Delta_Z"].abs().max(), 4),
                "Mean_Delta_Z":                    round(sub["Delta_Z"].mean(), 4),
                "Std_Delta_Z":                     round(sub["Delta_Z"].std(), 4),
                "N_No_Change":                     cats.get(_CAT_NO_CHANGE, 0),
                "N_Sig_to_NS":                     cats.get(_CAT_SIG_TO_NS, 0),
                "N_NS_to_Sig":                     cats.get(_CAT_NS_TO_SIG, 0),
                "N_Inc_to_Dec":                    cats.get(_CAT_INC_TO_DEC, 0),
                "N_Dec_to_Inc":                    cats.get(_CAT_DEC_TO_INC, 0),
            })
    return pd.DataFrame(rows)


def build_all_manuscript_table(master: pd.DataFrame) -> pd.DataFrame:
    """12-row wide manuscript table: one row per station, all methods + scales."""
    stations = sorted(master["Station"].unique(),
                      key=lambda x: int(x) if x.isdigit() else x)
    code_map = dict(zip(master["Station"], master["Code"]))
    rows = []
    for stn in stations:
        row: dict = {"Station": stn, "Code": code_map[stn]}
        for scale, abbr in _SCALE_ABBREV.items():
            sub = master[(master["Station"] == stn) & (master["Scale"] == scale)]
            if sub.empty:
                continue
            r = sub.iloc[0]
            row[f"{abbr}_MK_Z"]   = round(float(r["MK_Z"]),  3)
            row[f"{abbr}_MK_sig"] = str(r["MK_sig"])
            for alt in ("MMK", "PW", "TFPW"):
                alt_z  = float(r[f"{alt}_Z"])
                mk_z   = float(r["MK_Z"])
                alt_s  = bool(abs(alt_z) >= _Z05)
                mk_s   = bool(abs(mk_z)  >= _Z05)
                row[f"{abbr}_{alt}_Z"]       = round(alt_z, 3)
                row[f"{abbr}_{alt}_sig"]     = str(r[f"{alt}_sig"])
                row[f"{abbr}_{alt}_Delta_Z"] = round(alt_z - mk_z, 3)
                row[f"{abbr}_{alt}_Change"]  = _change_category(
                    mk_s, alt_s, str(r["MK_trend"]), str(r[f"{alt}_trend"]))
        rows.append(row)
    return pd.DataFrame(rows)


def build_significant_changes(station_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """All rows (any method) where Significance_Changed = True."""
    parts = [df[df["Significance_Changed"]].copy()
             for df in station_sheets.values()]
    if not any(len(p) for p in parts):
        return pd.DataFrame(columns=list(next(iter(station_sheets.values())).columns))
    return _sort_scale(pd.concat(parts, ignore_index=True))


def build_direction_changes(station_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    All rows (any method) where Direction_Changed = True.
    Captures any label change: Decreasing→No trend, No trend→Increasing, Inc→Dec, etc.
    """
    parts = [df[df["Direction_Changed"]].copy()
             for df in station_sheets.values()]
    if not any(len(p) for p in parts):
        return pd.DataFrame(columns=list(next(iter(station_sheets.values())).columns))
    return _sort_scale(pd.concat(parts, ignore_index=True))


def build_method_ranking(master: pd.DataFrame,
                         station_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    4-row table ranking all four methods by conservativeness.
    MK is the reference; non-MK methods also show agreement metrics vs MK.
    """
    rows = []
    for m in ("MK", "MMK", "PW", "TFPW"):
        z_col  = f"{m}_Z"
        t_col  = f"{m}_trend"
        sig_05 = master[z_col].abs() >= _Z05
        sig_01 = master[z_col].abs() >= _Z01
        n_sig  = int(sig_05.sum())
        n_inc_sig  = int((sig_05 & master[t_col].str.contains("Increasing", na=False)).sum())
        n_dec_sig  = int((sig_05 & master[t_col].str.contains("Decreasing", na=False)).sum())
        n_inc_all  = int(master[t_col].str.contains("Increasing", na=False).sum())
        n_dec_all  = int(master[t_col].str.contains("Decreasing", na=False).sum())
        n_no_trend = int((master[t_col] == "No trend").sum())
        mean_abs_z = round(master[z_col].abs().mean(), 4)

        row: dict = {
            "Method":               _ALT_META.get(m, "Standard MK"),
            "N_significant_05":     n_sig,
            "N_significant_01":     int(sig_01.sum()),
            "N_significant_Annual": int((sig_05 & (master["Scale"] == _SCALE_ORDER[0])).sum()),
            "N_significant_Wet":    int((sig_05 & (master["Scale"] == _SCALE_ORDER[1])).sum()),
            "N_significant_Dry":    int((sig_05 & (master["Scale"] == _SCALE_ORDER[2])).sum()),
            "N_Increasing_sig":     n_inc_sig,
            "N_Decreasing_sig":     n_dec_sig,
            "N_Increasing_all":     n_inc_all,
            "N_Decreasing_all":     n_dec_all,
            "N_No_trend":           n_no_trend,
            "Mean_abs_Z":           mean_abs_z,
            "Agreement_rate_vs_MK": None,
            "Mean_abs_Delta_Z_vs_MK": None,
        }

        if m != "MK":
            df  = station_sheets[m]
            agr = int((~df["Significance_Changed"] & ~df["Direction_Changed"]).sum())
            row["Agreement_rate_vs_MK"]      = round(agr / len(df), 4)
            row["Mean_abs_Delta_Z_vs_MK"]    = round(df["Delta_Z"].abs().mean(), 4)

        rows.append(row)

    df_rank = pd.DataFrame(rows)

    # Conservativeness rank: fewer significant → more conservative
    df_rank["Conservativeness_Rank"] = (
        df_rank["N_significant_05"]
        .rank(method="min", ascending=True)
        .astype(int)
    )
    df_rank["Liberalness_Rank"] = (
        df_rank["N_significant_05"]
        .rank(method="min", ascending=False)
        .astype(int)
    )

    return df_rank.sort_values("Conservativeness_Rank").reset_index(drop=True)


def build_reviewer_summary(master: pd.DataFrame,
                           station_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Multi-section reviewer-oriented summary.
    Columns: Section | Metric | MMK | PW | TFPW | Notes
    """
    rows = []
    sep = {"Section": "─────────────", "Metric": "", "MMK": "", "PW": "", "TFPW": "",
           "Notes": ""}

    def row(section, metric, mmk, pw, tfpw, notes=""):
        return {"Section": section, "Metric": metric,
                "MMK": mmk, "PW": pw, "TFPW": tfpw, "Notes": notes}

    # ── Section 1: Overall significance ──────────────────────────────────────
    rows.append(row("SIGNIFICANCE", "N_significant_total",
                    *[int(station_sheets[a]["Alt_Significant"].sum()) for a in ("MMK","PW","TFPW")],
                    "All scales combined"))
    rows.append(row("", "N_significant_Annual",
                    *[int(station_sheets[a][station_sheets[a]["Scale"]==_SCALE_ORDER[0]]["Alt_Significant"].sum()) for a in ("MMK","PW","TFPW")]))
    rows.append(row("", "N_significant_Wet",
                    *[int(station_sheets[a][station_sheets[a]["Scale"]==_SCALE_ORDER[1]]["Alt_Significant"].sum()) for a in ("MMK","PW","TFPW")]))
    rows.append(row("", "N_significant_Dry",
                    *[int(station_sheets[a][station_sheets[a]["Scale"]==_SCALE_ORDER[2]]["Alt_Significant"].sum()) for a in ("MMK","PW","TFPW")]))
    rows.append(row("", "N_sig_MK_baseline", 6, 6, 6, "Standard MK = 6 (reference)"))

    rows.append(sep.copy())

    # ── Section 2: Change category counts ────────────────────────────────────
    for cat in (_CAT_SIG_TO_NS, _CAT_NS_TO_SIG, _CAT_INC_TO_DEC, _CAT_DEC_TO_INC, _CAT_NO_CHANGE):
        vals = [int((station_sheets[a]["Change_Category"] == cat).sum())
                for a in ("MMK","PW","TFPW")]
        rows.append(row("CHANGE CATEGORIES", cat, *vals))

    rows.append(sep.copy())

    # ── Section 3: Agreement with MK ─────────────────────────────────────────
    for a in ("MMK","PW","TFPW"):
        df  = station_sheets[a]
        agr = int((~df["Significance_Changed"] & ~df["Direction_Changed"]).sum())
        other = {"MMK":"PW","PW":"TFPW","TFPW":"MMK"}  # not used, just to keep structure
    rows.append(row("AGREEMENT WITH MK", "N_full_agreement (36 combos)",
                    *[int((~station_sheets[a]["Significance_Changed"] &
                           ~station_sheets[a]["Direction_Changed"]).sum())
                      for a in ("MMK","PW","TFPW")]))
    rows.append(row("", "Agreement_rate",
                    *[round((~station_sheets[a]["Significance_Changed"] &
                              ~station_sheets[a]["Direction_Changed"]).mean(), 4)
                      for a in ("MMK","PW","TFPW")]))
    rows.append(row("", "Mean_abs_Delta_Z",
                    *[round(station_sheets[a]["Delta_Z"].abs().mean(), 4)
                      for a in ("MMK","PW","TFPW")]))
    rows.append(row("", "Max_abs_Delta_Z",
                    *[round(station_sheets[a]["Delta_Z"].abs().max(), 4)
                      for a in ("MMK","PW","TFPW")]))

    rows.append(sep.copy())

    # ── Section 4: Stations changing significance ─────────────────────────────
    sig_cats = {_CAT_SIG_TO_NS, _CAT_NS_TO_SIG}
    for a in ("MMK","PW","TFPW"):
        df  = station_sheets[a]
        changed = df[df["Change_Category"].isin(sig_cats)][
            ["Station","Scale","Change_Category"]].copy()
        for _, r in changed.iterrows():
            rows.append({
                "Section": "SIG CHANGES (detail)" if a == "MMK" else "",
                "Metric":  f"{_ALT_META[a]}",
                a:         f"{r['Station']} | {_SCALE_ABBREV[r['Scale']]} | {r['Change_Category']}",
                **{o: "" for o in ("MMK","PW","TFPW") if o != a},
                "Notes": "",
            })
    if not any(station_sheets[a]["Significance_Changed"].any() for a in ("MMK","PW","TFPW")):
        rows.append(row("SIG CHANGES (detail)", "None", "", "", "", "No significance changes detected"))

    rows.append(sep.copy())

    # ── Section 5: Stations changing direction ────────────────────────────────
    dir_cats = {_CAT_INC_TO_DEC, _CAT_DEC_TO_INC}
    found_any = False
    for a in ("MMK","PW","TFPW"):
        df  = station_sheets[a]
        changed = df[df["Change_Category"].isin(dir_cats)][
            ["Station","Scale","Change_Category","MK_Trend","Alt_Trend"]].copy()
        for _, r in changed.iterrows():
            found_any = True
            rows.append({
                "Section": "DIR CHANGES (detail)" if not found_any else "",
                "Metric":  f"{_ALT_META[a]}",
                a:         f"{r['Station']} | {_SCALE_ABBREV[r['Scale']]} | {r['MK_Trend']}→{r['Alt_Trend']}",
                **{o: "" for o in ("MMK","PW","TFPW") if o != a},
                "Notes": "",
            })
    if not found_any:
        rows.append(row("DIR CHANGES (detail)", "None", "", "", "", "No direction changes detected"))

    rows.append(sep.copy())

    # ── Section 6: Largest departures from MK ────────────────────────────────
    rows.append(row("LARGEST |ΔZ|", "Top station-scale by |ΔZ| (per method)",
                    "", "", "", ""))
    for a in ("MMK","PW","TFPW"):
        df   = station_sheets[a]
        top3 = df.nlargest(3, "Delta_Z", keep="all")[["Station","Scale","Delta_Z"]]
        for rank, (_, r) in enumerate(top3.iterrows(), 1):
            rows.append({
                "Section": "",
                "Metric":  f"Rank {rank} — {_ALT_META[a]}",
                a:         f"{r['Station']} | {_SCALE_ABBREV[r['Scale']]} | ΔZ={r['Delta_Z']:+.3f}",
                **{o: "" for o in ("MMK","PW","TFPW") if o != a},
                "Notes": "",
            })

    rows.append(sep.copy())

    # ── Section 7: Autocorrelation summary ───────────────────────────────────
    n_sig_ac = int((master["Sig_AC"] == "Yes*").sum())
    rows.append(row("AUTOCORRELATION", "N_stations_Sig_AC (any scale)",
                    n_sig_ac, n_sig_ac, n_sig_ac,
                    "Sig_AC from Master_DB (MK baseline shared across methods)"))
    if "rho_1" in master.columns:
        rows.append(row("", "Max_rho_1",
                        round(float(master["rho_1"].max()), 4),
                        round(float(master["rho_1"].max()), 4),
                        round(float(master["rho_1"].max()), 4)))
    if "Correction_Factor" in master.columns:
        rows.append(row("", "Max_Correction_Factor (MMK)",
                        round(float(master["Correction_Factor"].max()), 4),
                        "N/A", "N/A",
                        "CF = Var*(S)/Var(S); MMK only"))

    return pd.DataFrame(rows)


# ── Validation ─────────────────────────────────────────────────────────────────

def validate(master: pd.DataFrame,
             station_sheets: dict[str, pd.DataFrame],
             summary: pd.DataFrame) -> list[str]:
    issues: list[str] = []

    # Row counts
    for alt, df in station_sheets.items():
        if len(df) != 36:
            issues.append(f"{alt}_vs_MK_Station: expected 36 rows, got {len(df)}")

    if len(summary) != 9:
        issues.append(f"All_Scale_Summary: expected 9 rows, got {len(summary)}")

    # No duplicate Station × Scale per sheet
    for alt, df in station_sheets.items():
        n_dup = int(df.duplicated(subset=["Station", "Scale"]).sum())
        if n_dup:
            issues.append(f"{alt}: {n_dup} duplicate Station×Scale rows")

    # All master joins present
    master_keys = set(zip(master["Station"], master["Scale"]))
    for alt, df in station_sheets.items():
        missing = master_keys - set(zip(df["Station"], df["Scale"]))
        if missing:
            issues.append(f"{alt}: {len(missing)} missing joins: {missing}")

    # No NaN in key numeric columns
    for alt, df in station_sheets.items():
        for col in ("MK_Z", "Alt_Z", "Delta_Z", "Delta_p"):
            n_nan = int(df[col].isna().sum())
            if n_nan:
                issues.append(f"{alt}.{col}: {n_nan} NaN values")

    # Change_Category covers all 36 rows
    for alt, df in station_sheets.items():
        expected = {_CAT_NO_CHANGE, _CAT_SIG_TO_NS, _CAT_NS_TO_SIG,
                    _CAT_INC_TO_DEC, _CAT_DEC_TO_INC}
        actual = set(df["Change_Category"].unique())
        unknown = actual - expected
        if unknown:
            issues.append(f"{alt}: unexpected Change_Category values: {unknown}")

    return issues


# ── Validation report ──────────────────────────────────────────────────────────

def print_report(master: pd.DataFrame,
                 station_sheets: dict[str, pd.DataFrame],
                 summary: pd.DataFrame,
                 ranking: pd.DataFrame,
                 issues: list[str]) -> None:
    SEP = "─" * 66
    print(f"\n{SEP}")
    print("  Validation Report — Trend_Method_Comparison_All_vs_MK.xlsx")
    print(SEP)

    print("\n  Row counts")
    for alt, df in station_sheets.items():
        ok = "✓" if len(df) == 36 else "✗"
        print(f"    {ok} {alt}_vs_MK_Station      : {len(df)} rows")
    ok = "✓" if len(summary) == 9 else "✗"
    print(f"    {ok} All_Scale_Summary          : {len(summary)} rows (3 methods × 3 scales)")

    print("\n  Duplicate check (Station × Scale uniqueness)")
    for alt, df in station_sheets.items():
        n_dup = int(df.duplicated(subset=["Station","Scale"]).sum())
        ok = "✓" if n_dup == 0 else "✗"
        print(f"    {ok} {alt}: {n_dup} duplicates")

    print("\n  Join completeness (vs Master_DB — 36 keys required)")
    master_keys = set(zip(master["Station"], master["Scale"]))
    for alt, df in station_sheets.items():
        matched = len(master_keys & set(zip(df["Station"], df["Scale"])))
        ok = "✓" if matched == 36 else "✗"
        print(f"    {ok} {alt}: {matched}/36 Station×Scale keys matched")

    print("\n  Summary statistics per method (all scales combined)")
    header = f"    {'Metric':<40}  {'MMK':>8}  {'PW':>8}  {'TFPW':>8}"
    print(header)
    print("    " + "─" * 62)
    metrics = [
        ("Number_of_changed_significance",
         lambda a: int(station_sheets[a]["Significance_Changed"].sum())),
        ("Number_of_changed_direction",
         lambda a: int(station_sheets[a]["Direction_Changed"].sum())),
        ("Agreement_rate_with_MK",
         lambda a: round((~station_sheets[a]["Significance_Changed"] &
                           ~station_sheets[a]["Direction_Changed"]).mean(), 4)),
        ("Mean_abs_Delta_Z",
         lambda a: round(station_sheets[a]["Delta_Z"].abs().mean(), 4)),
        ("Max_abs_Delta_Z",
         lambda a: round(station_sheets[a]["Delta_Z"].abs().max(), 4)),
    ]
    for lbl, fn in metrics:
        vals = [str(fn(a)) for a in ("MMK","PW","TFPW")]
        print(f"    {lbl:<40}  {vals[0]:>8}  {vals[1]:>8}  {vals[2]:>8}")

    print("\n  Change category breakdown (all 36 rows per method)")
    cats = [_CAT_NO_CHANGE, _CAT_SIG_TO_NS, _CAT_NS_TO_SIG,
            _CAT_INC_TO_DEC, _CAT_DEC_TO_INC]
    print(f"    {'Category':<42}  {'MMK':>5}  {'PW':>5}  {'TFPW':>5}")
    print("    " + "─" * 62)
    for cat in cats:
        vals = [int((station_sheets[a]["Change_Category"]==cat).sum())
                for a in ("MMK","PW","TFPW")]
        print(f"    {cat:<42}  {vals[0]:>5}  {vals[1]:>5}  {vals[2]:>5}")

    print("\n  Method ranking by conservativeness (fewest significant = most conservative)")
    for _, r in ranking.iterrows():
        print(f"    Rank {r['Conservativeness_Rank']}: {r['Method']:<30}  "
              f"N_sig={r['N_significant_05']}")

    if issues:
        print(f"\n  ✗  {len(issues)} VALIDATION ISSUE(S) DETECTED:")
        for iss in issues:
            print(f"     • {iss}")
    else:
        print("\n  ✓  All validation checks passed — no issues detected")

    print(f"\n{SEP}\n")


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    if OUT.exists():
        print(f"ERROR: Output already exists and must not be overwritten:\n  {OUT}")
        sys.exit(1)

    print(f"Reading Master_DB …")
    master = load_master()
    print(f"  Loaded {len(master)} rows × {len(master.columns)} cols")

    print("Building station comparison sheets …")
    station_sheets = {alt: build_station_sheet(master, alt)
                      for alt in ("MMK", "PW", "TFPW")}

    print("Building cross-method summary sheets …")
    summary     = build_all_scale_summary(station_sheets)
    manuscript  = build_all_manuscript_table(master)
    sig_changes = build_significant_changes(station_sheets)
    dir_changes = build_direction_changes(station_sheets)
    ranking     = build_method_ranking(master, station_sheets)
    reviewer    = build_reviewer_summary(master, station_sheets)

    issues = validate(master, station_sheets, summary)

    print_report(master, station_sheets, summary, ranking, issues)

    if issues:
        print("Aborting — resolve issues before writing workbook.")
        sys.exit(1)

    sheets = {
        "MMK_vs_MK_Station":    station_sheets["MMK"],
        "PW_vs_MK_Station":     station_sheets["PW"],
        "TFPW_vs_MK_Station":   station_sheets["TFPW"],
        "All_Scale_Summary":    summary,
        "All_Manuscript_Table": manuscript,
        "Significant_Changes":  sig_changes,
        "Direction_Changes":    dir_changes,
        "Method_Ranking":       ranking,
        "Reviewer_Summary":     reviewer,
    }

    color_map = {
        "MMK_vs_MK_Station":    _COLORS["MMK"],
        "PW_vs_MK_Station":     _COLORS["PW"],
        "TFPW_vs_MK_Station":   _COLORS["TFPW"],
        "All_Scale_Summary":    _COLORS["COMP"],
        "All_Manuscript_Table": _COLORS["COMP"],
        "Significant_Changes":  _COLORS["CHNG"],
        "Direction_Changes":    _COLORS["DIR"],
        "Method_Ranking":       _COLORS["RANK"],
        "Reviewer_Summary":     _COLORS["REV"],
    }

    print(f"Writing {OUT.name} …")
    _write_excel(sheets, OUT, color_map=color_map)

    size_kb = OUT.stat().st_size / 1024
    print(f"  Written : {OUT}")
    print(f"  Size    : {size_kb:.1f} KB")
    print(f"  Sheets  : {len(sheets)}")
    for name, df in sheets.items():
        print(f"    [{name}]  {len(df)} rows × {len(df.columns)} cols")


if __name__ == "__main__":
    main()
