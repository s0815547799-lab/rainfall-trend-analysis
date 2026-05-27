"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  Rainfall Trend Analysis — Publication Edition v2.0                         ║
║  Study: Phetchaburi–Prachuap Khiri Khan River Basin, Western Thailand       ║
║  Period: 1981–2014  |  Daily Rainfall Data                                  ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  Analysis Workflow:                                                          ║
║  Step 1  : Data Loading & Quality Control (IQR outlier, linear interp.)    ║
║  Step 2  : Temporal Aggregation                                             ║
║             – Annual (Jan–Dec)                                              ║
║             – Hydrological Wet Season  (May–Oct)                           ║
║             – Hydrological Dry Season  (Nov–Apr)  ← Hydrological Year      ║
║             – Seasonal Cycle Monthly                                        ║
║  Step 3  : Descriptive Statistics (Mean, Max, Min, Std, CV, Wet-days)      ║
║  Step 4  : Lag-k Autocorrelation Assessment                                 ║
║  Step 5  : Standard Mann–Kendall (MK) Test                                 ║
║  Step 6  : Modified Mann–Kendall (MMK) Test  — Hamed & Rao (1998)          ║
║  Step 7  : Sen's Slope Estimator + 95% CI  — Sen (1968) / Gilbert (1987)  ║
║  Step 8  : MK vs MMK Comparison + Statistical Summary                      ║
║  Step 9  : Publication Figures (Fig 1–8)                                   ║
║  Step 10 : Excel Tables (6 sheets)                                         ║
║  Step 11 : Research Summary Document (Markdown → ready for paper writing)  ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  Temporal Scales:                                                           ║
║    Annual   : January–December (calendar year)                             ║
║    Wet      : May–October      (monsoon / wet season)                      ║
║    Dry      : November–April   (dry season, hydrological year)             ║
║    Monthly  : 12-month cycle   (climatological mean)                       ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  Output Files (prefix = Output_TrendV2_<basename>):                        ║
║    Fig1_AnnualTimeSeries.png/.pdf                                           ║
║    Fig2_WetDryTimeSeries.png/.pdf                                           ║
║    Fig3_SenSlope_AllScales.png/.pdf                                         ║
║    Fig4_MK_vs_MMK_Comparison.png/.pdf                                       ║
║    Fig5_Significance_Heatmap.png/.pdf                                       ║
║    Fig6_Autocorrelation.png/.pdf                                            ║
║    Fig7_MonthlyClimatology.png/.pdf                                         ║
║    Fig8_SpatialTrend_Summary.png/.pdf                                       ║
║    Results_TrendAnalysis.xlsx  (6 sheets)                                   ║
║    Research_Summary.md         (paper-ready text)                          ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  References:                                                                ║
║    Mann (1945) Econometrica 13:245–259                                      ║
║    Kendall (1975) Rank Correlation Methods. Griffin, London.               ║
║    Sen (1968) JASA 63:1379–1389                                            ║
║    Hamed & Rao (1998) J. Hydrol. 204:182–196                               ║
║    Gilbert (1987) Statistical Methods for Environmental Pollution          ║
║    Yue & Wang (2004) Water Resour. Res. 40:W08307                          ║
║    Önöz & Bayazit (2003) Hydrol. Sci. J. 48:25–34                         ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

# ── Standard library ──────────────────────────────────────────────────────────
import os, sys, math, warnings, textwrap
from pathlib import Path
from datetime import datetime

# ── Scientific stack ──────────────────────────────────────────────────────────
import numpy as np
import pandas as pd
from scipy import stats as sps
from scipy.stats import norm as scipy_norm

# ── v4 extension modules (rta package) ───────────────────────────────────────
# Gracefully degrade if rta is not installed (keeps v3 runnable standalone)
try:
    _RTA_AVAILABLE = False
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).parent))
    from rta.pw               import pw_mk
    from rta.tfpw             import tfpw_mk
    from rta.field_significance import (walker_test, livezey_chen_mc,
                                        field_sig_summary)
    from rta.checkpoint       import (save   as _ckpt_save,
                                      load   as _ckpt_load,
                                      list_steps as _ckpt_list,
                                      prompt_resume as _ckpt_prompt)
    from rta.spatial          import load_coords, validate_coords
    _RTA_AVAILABLE = True
except ImportError:
    _RTA_AVAILABLE = False
    def pw_mk(x):   return {}
    def tfpw_mk(x): return {}
    def field_sig_summary(*a, **kw): return None
    def walker_test(*a, **kw): return {}
    def _ckpt_save(*a, **kw): pass
    def _ckpt_load(*a, **kw): return None
    def _ckpt_list(*a, **kw): return []
    def _ckpt_prompt(*a, **kw): return 0
    def load_coords(*a, **kw): return None
    def validate_coords(*a, **kw): return {}

# ── Visualisation ─────────────────────────────────────────────────────────────
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.ticker as ticker
import matplotlib.patches as mpatches
import matplotlib.colors as mcolors
from matplotlib.lines import Line2D
import matplotlib.cm as cm

# ── Excel ─────────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §0  CONSTANTS & STYLE                                                  ║
# ╚══════════════════════════════════════════════════════════════════════════╝

VERSION      = "2.0"
WET_THR      = 1.0                    # WMO wet-day threshold (mm/day)
WET_MONTHS   = [5, 6, 7, 8, 9, 10]   # Wet season: May–October
DRY_MONTHS   = [11, 12, 1, 2, 3, 4]  # Dry season: November–April
MIN_N        = 10                     # minimum years for MK test
ALPHA_005    = 0.05
ALPHA_001    = 0.01
Z_005        = 1.9600
Z_001        = 2.5758
SAVE_PDF     = True
DPI          = 600
MONTH_ABBR   = ["Jan","Feb","Mar","Apr","May","Jun",
                 "Jul","Aug","Sep","Oct","Nov","Dec"]

# ── Colour palette (colour-blind safe) ───────────────────────────────────────
C = dict(
    annual  = "#37474F",  annual_lt = "#B0BEC5",
    wet     = "#1565C0",  wet_lt    = "#90CAF9",
    dry     = "#E65100",  dry_lt    = "#FFCC80",
    inc     = "#1B5E20",  inc_lt    = "#A5D6A7",
    dec     = "#B71C1C",  dec_lt    = "#EF9A9A",
    ns_col  = "#78909C",  ns_lt     = "#CFD8DC",
    mk_std  = "#6A1B9A",  mk_mod    = "#0277BD",
    gold    = "#F9A825",  grey      = "#546E7A",
)

# ── Matplotlib publication style ─────────────────────────────────────────────
plt.rcParams.update({
    "font.family":        "serif",
    "font.serif":         ["Times New Roman", "DejaVu Serif"],
    "font.size":          12,
    "axes.titlesize":     13,
    "axes.labelsize":     12,
    "xtick.labelsize":    11,
    "ytick.labelsize":    11,
    "legend.fontsize":    10.5,
    "figure.titlesize":   13,
    "lines.linewidth":    2.0,
    "axes.linewidth":     1.0,
    "axes.spines.top":    False,
    "axes.spines.right":  False,
    "axes.grid":          True,
    "grid.linestyle":     "--",
    "grid.linewidth":     0.4,
    "grid.alpha":         0.40,
    "grid.color":         "#B0BEC5",
    "savefig.dpi":        DPI,
    "savefig.bbox":       "tight",
    "savefig.pad_inches": 0.15,
    "figure.dpi":         100,
    "mathtext.fontset":   "stix",
    "pdf.fonttype":       42,
    "ps.fonttype":        42,
})

# ── Excel style ──────────────────────────────────────────────────────────────
THIN = Side(style="thin",   color="BDBDBD")
MED  = Side(style="medium", color="1F4E79")
XC   = dict(
    title  = "13293D", sub = "1F4E79", hdr = "2E75B6",
    wet_h  = "DDEEFF", dry_h = "FFF3E0",
    ann_h  = "ECEFF1", mon_h = "E8F5E9",
    sig05  = "FFF9C4", sig01 = "FFECB3",
    inc_c  = "E8F5E9", dec_c = "FFEBEE",
    ns_c   = "F5F5F5", white = "FFFFFF",
    mk_h   = "EDE7F6", mmk_h = "E3F2FD",
    diff_h = "FFF8E1",
)

def tb():     return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def xfill(h): return PatternFill("solid", fgColor=h)

def xsc(ws, r, c, val=None, bold=False, italic=False,
        fc=None, bg=None, align="center", sz=10, wrap=True, border=None):
    cell = ws.cell(row=r, column=c)
    if val is not None: cell.value = val
    cell.font      = Font(bold=bold, italic=italic, name="Calibri",
                          size=sz, color=fc if fc else "1A1A1A")
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    if bg:     cell.fill   = xfill(bg)
    if border: cell.border = border
    return cell

def mxsc(ws, r, c1, c2, val, **kw):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    return xsc(ws, r, c1, val, **kw)

def cw(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def rh(ws, r, h):
    ws.row_dimensions[r].height = h

def savefig(fig, path_noext: str):
    fig.savefig(f"{path_noext}.png", dpi=DPI, bbox_inches="tight", pad_inches=0.15)
    if SAVE_PDF:
        fig.savefig(f"{path_noext}.pdf", bbox_inches="tight", pad_inches=0.15)
    plt.close(fig)
    print(f"    ✓  {Path(path_noext).name}.png" + (" + .pdf" if SAVE_PDF else ""))


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §1  DATA LOADING & QUALITY CONTROL                                     ║
# ╚══════════════════════════════════════════════════════════════════════════╝

MISS_FLAGS = [-99, -999, -9999, -9.99e+20, 9.99e+20, 1e+20]


def find_csv(folder: str) -> str:
    """Auto-discover daily rainfall CSV in folder."""
    csvs = sorted(Path(folder).glob("*.csv"))
    obs  = [f for f in csvs
            if "observed" in f.name.lower() or "rain" in f.name.lower()]
    return str((obs or csvs)[0]) if (obs or csvs) else sys.exit("No CSV found")


def load_daily(path: str) -> pd.DataFrame:
    """Load daily rainfall CSV → DatetimeIndex DataFrame."""
    df = pd.read_csv(path)
    for mv in MISS_FLAGS:
        df.replace(mv, np.nan, inplace=True)
    df.columns = [str(c) for c in df.columns]
    stns = [c for c in df.columns if c not in ("YEAR","MONTH","DAY")]
    for s in stns:
        df.loc[df[s] < 0, s] = np.nan
    df["date"] = pd.to_datetime(
        {"year": df["YEAR"], "month": df["MONTH"], "day": df["DAY"]})
    return df.set_index("date")[stns]


def quality_control(df: pd.DataFrame) -> tuple:
    """
    QC: report missing values, detect outliers (IQR 3×),
    fill short gaps (≤5 days) by linear interpolation.
    Returns: (df_clean, qc_report_dict)
    """
    stns = df.columns.tolist()
    qc   = {}
    df   = df.copy()
    for s in stns:
        series     = df[s].copy()
        n_miss     = int(series.isna().sum())
        wet_vals   = series[(series >= WET_THR) & series.notna()]
        q1, q3     = float(wet_vals.quantile(0.25)), float(wet_vals.quantile(0.75))
        iqr        = q3 - q1
        upper_fence= q3 + 3.0 * iqr
        n_out      = int((series > upper_fence).sum())
        filled     = series.interpolate(method="time", limit=5,
                                        limit_direction="both")
        n_fill     = int(filled.notna().sum()) - int(series.notna().sum())
        df[s]      = filled
        qc[s] = dict(n_total=len(series), n_missing=n_miss,
                     pct_miss=round(n_miss/len(series)*100, 2),
                     n_outlier=n_out,
                     upper_fence=round(upper_fence, 1),
                     n_filled=n_fill)
    return df, qc


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §2  TEMPORAL AGGREGATION                                               ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def aggregate_all(df: pd.DataFrame) -> dict:
    """
    Aggregate to 4 temporal scales.
    'dry' season (Nov–Apr) crosses calendar years → shift Nov-Dec to next year
    so each dry-season block is complete (Nov Y → Apr Y+1, labelled year Y+1).
    min_count ensures ≥80% completeness (≥60% for dry/wet).
    """
    scales = {}

    # Annual (Jan–Dec)
    scales["annual"] = df.resample("YS").apply(
        lambda g: g.sum(min_count=int(0.8 * len(g))))

    # Wet season (May–Oct)
    wet = df[df.index.month.isin(WET_MONTHS)]
    scales["wet"] = wet.resample("YS").apply(
        lambda g: g.sum(min_count=int(0.8 * len(g))))

    # Dry season (Nov–Apr) — hydrological year approach
    # Nov/Dec of year Y shift to year Y+1 so the 6 months are continuous
    dry_raw = df[df.index.month.isin(DRY_MONTHS)].copy()
    late_mask = dry_raw.index.month.isin([11, 12])
    new_idx = [d.replace(year=d.year + 1) if m else d
               for d, m in zip(dry_raw.index.to_list(), late_mask)]
    dry_raw.index = pd.DatetimeIndex(new_idx)
    scales["dry"] = dry_raw.resample("YS").apply(
        lambda g: g.sum(min_count=int(0.8 * len(g))))
    # Null out blocks missing any of the 6 required months (Nov–Apr).
    # Boundary blocks at data start/end are incomplete and must not enter MK.
    _dry_req = frozenset([11, 12, 1, 2, 3, 4])
    _mcov = (dry_raw.groupby(dry_raw.index.year)
             .apply(lambda g: frozenset(g.index.month.tolist())))
    _incomp = [yr for yr in scales["dry"].index.year
               if _mcov.get(yr, frozenset()) != _dry_req]
    if _incomp:
        scales["dry"].loc[scales["dry"].index.year.isin(_incomp)] = np.nan

    # Monthly climatology (mean monthly total)
    monthly_all = df.resample("MS").apply(
        lambda g: g.sum(min_count=int(0.8 * len(g))))
    scales["monthly_all"] = monthly_all   # full monthly series

    return scales


def descriptive_stats(scales: dict, df_daily: pd.DataFrame) -> pd.DataFrame:
    """Descriptive statistics for annual series."""
    ann   = scales["annual"]
    stns  = ann.columns.tolist()
    rows  = []
    for s in stns:
        v = ann[s].dropna().values.astype(float)
        d = df_daily[s].dropna()
        w = d[d >= WET_THR]
        n = len(v)
        rows.append({
            "Station":      s,
            "N (yr)":       n,
            "Mean (mm)":    round(float(np.mean(v)), 1)          if n > 0 else np.nan,
            "Median (mm)":  round(float(np.median(v)), 1)        if n > 0 else np.nan,
            "Max (mm)":     round(float(np.max(v)), 1)           if n > 0 else np.nan,
            "Min (mm)":     round(float(np.min(v)), 1)           if n > 0 else np.nan,
            "Std (mm)":     round(float(np.std(v, ddof=1)), 1)   if n > 1 else np.nan,
            "CV (%)":       round(float(np.std(v,ddof=1)/np.mean(v)*100), 1)
                            if n > 1 and np.mean(v) != 0 else np.nan,
            "Wet-days/yr":  round(float(len(w) / n), 1)          if n > 0 else np.nan,
            "Skewness":     round(float(sps.skew(v)), 3)          if n > 3 else np.nan,
            "Kurtosis":     round(float(sps.kurtosis(v,fisher=True)), 3) if n>3 else np.nan,
        })
    return pd.DataFrame(rows).set_index("Station")


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §3  AUTOCORRELATION                                                    ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def lag_k_autocorr(x: np.ndarray, k: int = 1) -> float:
    """Pearson Lag-k autocorrelation."""
    x = np.asarray(x, dtype=float)
    x = x[~np.isnan(x)]
    n = len(x)
    if n < k + 3: return np.nan
    xb  = np.mean(x)
    num = np.sum((x[:n-k] - xb) * (x[k:n] - xb))
    den = np.sum((x - xb) ** 2)
    return float(num / den) if den > 0 else np.nan


def all_lag_autocorr(x: np.ndarray, max_lag: int = None) -> np.ndarray:
    """Autocorrelation for lags 1..max_lag (default n//3)."""
    x = np.asarray(x, dtype=float)
    x = x[~np.isnan(x)]
    n = len(x)
    if n < 4: return np.array([])
    if max_lag is None: max_lag = min(n // 3, n - 1)
    return np.array([lag_k_autocorr(x, k) for k in range(1, max_lag + 1)])


def is_sig_autocorr(r1: float, n: int, alpha: float = 0.05) -> bool:
    """Two-tailed significance of Lag-1 autocorrelation."""
    if np.isnan(r1) or n < 4: return False
    return abs(r1) > scipy_norm.ppf(1 - alpha/2) / math.sqrt(n)


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §4  STANDARD MANN–KENDALL TEST                                         ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def mk_s_ties(x: np.ndarray) -> tuple:
    """Compute S statistic and tie sizes."""
    x = x[~np.isnan(x)]
    n = len(x)
    if n < 4: return np.nan, []
    S = int(np.sum(np.sign(x[j] - x[i])
                   for i in range(n-1) for j in range(i+1, n)))
    _, counts = np.unique(x, return_counts=True)
    ties = counts[counts > 1].tolist()
    return float(S), ties


def _mk_s_fast(x: np.ndarray) -> tuple:
    """Vectorised S statistic (faster than nested loops for large n)."""
    x = np.asarray(x, dtype=float)
    x = x[~np.isnan(x)]
    n = len(x)
    if n < 4: return np.nan, []
    S = 0
    for i in range(n - 1):
        S += int(np.sum(np.sign(x[i+1:] - x[i])))
    _, counts = np.unique(x, return_counts=True)
    ties = counts[counts > 1].tolist()
    return float(S), ties


def mk_variance_ties(n: int, ties: list) -> float:
    """Var(S) with tie correction."""
    tie_sum = sum(t * (t - 1) * (2 * t + 5) for t in ties)
    return (n * (n - 1) * (2 * n + 5) - tie_sum) / 18.0


def standard_mk(x: np.ndarray) -> dict:
    """
    Standard Mann–Kendall Test  (Mann 1945; Kendall 1975).

    Does NOT correct for serial autocorrelation.
    Use this as baseline; compare with Modified MK.
    """
    null = {k: np.nan for k in ["S","n","Var_S","Z","p_value","tau",
                                  "slope_Q","slope_lo","slope_hi",
                                  "trend","sig_05","sig_01"]}
    null.update({"trend":"—","sig_05":False,"sig_01":False})

    x = np.asarray(x, dtype=float)
    x = x[~np.isnan(x)]
    n = int(len(x))
    if n < MIN_N: return null

    S, ties = _mk_s_fast(x)
    if np.isnan(S): return null

    Var_S = mk_variance_ties(n, ties)
    if Var_S <= 0: return null

    Z = (S - 1) / math.sqrt(Var_S) if S > 0 else \
        (S + 1) / math.sqrt(Var_S) if S < 0 else 0.0
    p_val = float(min(2.0 * (1.0 - scipy_norm.cdf(abs(Z))), 1.0))
    tau   = float(S / (0.5 * n * (n - 1)))
    sig05 = p_val < ALPHA_005
    sig01 = p_val < ALPHA_001
    trend = ("Increasing ↑" if (sig05 and Z > 0) else
             "Decreasing ↓" if (sig05 and Z < 0) else "No trend")

    slope_Q, slope_lo, slope_hi = sens_slope(x)

    return {"S":       float(S),          "n":       n,
            "Var_S":   round(Var_S, 2),   "Z":       round(Z, 4),
            "p_value": round(p_val, 6),   "tau":     round(tau, 4),
            "slope_Q": round(slope_Q, 3)  if not np.isnan(slope_Q)  else np.nan,
            "slope_lo":round(slope_lo, 3) if not np.isnan(slope_lo) else np.nan,
            "slope_hi":round(slope_hi, 3) if not np.isnan(slope_hi) else np.nan,
            "trend":   trend, "sig_05": sig05, "sig_01": sig01,
            "method":  "Standard MK"}


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §5  MODIFIED MANN–KENDALL TEST  (Hamed & Rao 1998)                    ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def modified_mk(x: np.ndarray) -> dict:
    """
    Modified Mann–Kendall Test  (Hamed & Rao 1998, J. Hydrol. 204:182–196).

    Corrects Var(S) for serial autocorrelation using ranked-series
    autocorrelations and effective sample size n*:

       n / n* = 1 + (2/n) Σ_{k=1}^{n-1} (n-k) ρ_k(ranks)

    Only statistically significant ρ_k are used (Hamed & Rao 1998).
    Var*(S) = Var(S) × (n / n*)
    """
    null = {k: np.nan for k in ["S","n","Var_S","Var_S_adj","n_eff",
                                  "rho_1","Z","p_value","tau",
                                  "slope_Q","slope_lo","slope_hi",
                                  "trend","sig_05","sig_01"]}
    null.update({"trend":"—","sig_05":False,"sig_01":False})

    x = np.asarray(x, dtype=float)
    x = x[~np.isnan(x)]
    n = int(len(x))
    if n < MIN_N: return null

    S, ties = _mk_s_fast(x)
    if np.isnan(S): return null

    Var_S = mk_variance_ties(n, ties)
    if Var_S <= 0: return null

    # Autocorrelations of ranked series (Hamed & Rao 1998)
    ranks = sps.rankdata(x).astype(float)
    rho   = all_lag_autocorr(ranks, max_lag=min(n // 3, n - 1))

    # Effective sample size correction
    if len(rho) == 0:
        n_over_neff = 1.0
    else:
        se_rho = 1.0 / math.sqrt(n)
        z_crit = scipy_norm.ppf(1 - ALPHA_005 / 2)
        rho_sig = np.where(np.abs(rho) > z_crit * se_rho, rho, 0.0)
        ks   = np.arange(1, len(rho_sig) + 1)
        n_over_neff = max(1.0 + (2.0 / n) * np.sum((n - ks) * rho_sig), 1.0)

    rho_1     = float(rho[0]) if len(rho) > 0 else np.nan
    n_eff     = n / n_over_neff
    Var_S_adj = Var_S * n_over_neff

    Z     = (S - 1) / math.sqrt(Var_S_adj) if S > 0 else \
            (S + 1) / math.sqrt(Var_S_adj) if S < 0 else 0.0
    p_val = float(min(2.0 * (1.0 - scipy_norm.cdf(abs(Z))), 1.0))
    tau   = float(S / (0.5 * n * (n - 1)))
    sig05 = p_val < ALPHA_005
    sig01 = p_val < ALPHA_001
    trend = ("Increasing ↑" if (sig05 and Z > 0) else
             "Decreasing ↓" if (sig05 and Z < 0) else "No trend")

    slope_Q, slope_lo, slope_hi = sens_slope(x)

    return {"S":        float(S),            "n":         n,
            "Var_S":    round(Var_S, 2),
            "Var_S_adj":round(Var_S_adj, 2), "n_eff":     round(n_eff, 2),
            "rho_1":    round(rho_1, 4)      if not np.isnan(rho_1) else np.nan,
            "Z":        round(Z, 4),          "p_value":   round(p_val, 6),
            "tau":      round(tau, 4),
            "slope_Q":  round(slope_Q, 3)    if not np.isnan(slope_Q)  else np.nan,
            "slope_lo": round(slope_lo, 3)   if not np.isnan(slope_lo) else np.nan,
            "slope_hi": round(slope_hi, 3)   if not np.isnan(slope_hi) else np.nan,
            "trend":    trend, "sig_05": sig05, "sig_01": sig01,
            "method":   "Modified MK (H&R98)"}


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §6  SEN'S SLOPE ESTIMATOR  (Sen 1968 / Gilbert 1987)                  ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def sens_slope(x: np.ndarray, alpha: float = 0.05) -> tuple:
    """
    Sen's Slope Estimator + 95% CI (rank-based, Gilbert 1987).

    Q  = median[(xⱼ − xᵢ)/(j − i)]  for all j > i
    CI : Cα = z_{α/2} × √Var(S)
         lo_rank = (N − Cα)/2,  hi_rank = (N + Cα)/2 + 1
    """
    x = np.asarray(x, dtype=float)
    x = x[~np.isnan(x)]
    n = len(x)
    if n < 4: return np.nan, np.nan, np.nan

    slopes = []
    for i in range(n - 1):
        for j in range(i + 1, n):
            slopes.append((x[j] - x[i]) / (j - i))
    slopes = np.sort(slopes)
    N      = len(slopes)
    Q      = float(np.median(slopes))

    _, ties = _mk_s_fast(x)
    Var_S   = mk_variance_ties(n, ties)
    if Var_S <= 0: return Q, np.nan, np.nan

    z_crit  = scipy_norm.ppf(1 - alpha / 2)
    C_alpha = z_crit * math.sqrt(Var_S)
    lo_r    = max(0,     int(round((N - C_alpha) / 2.0)))
    hi_r    = min(N - 1, int(round((N + C_alpha) / 2.0)))

    return Q, float(slopes[lo_r]), float(slopes[hi_r])


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §7  RUN ALL STATIONS × SCALES × METHODS                               ║
# ╚══════════════════════════════════════════════════════════════════════════╝

SCALE_META = {
    "annual": {"label":"Annual (Jan–Dec)",       "unit":"mm yr⁻¹",   "color":C["annual"]},
    "wet":    {"label":"Wet Season (May–Oct)",    "unit":"mm season⁻¹","color":C["wet"]},
    "dry":    {"label":"Dry Season (Nov–Apr)",    "unit":"mm season⁻¹","color":C["dry"]},
}


def run_all(scales: dict, stns: list, smap: dict) -> pd.DataFrame:
    """
    Run Standard MK + Modified MK for all stations × all temporal scales.
    Returns tidy DataFrame with one row per (Station × Scale × Method).
    """
    rows = []
    scale_keys = ["annual", "wet", "dry"]

    for sk in scale_keys:
        df_s = scales[sk]
        meta = SCALE_META[sk]
        for stn in [str(s) for s in stns]:
            if stn not in df_s.columns:
                continue
            arr = df_s[stn].dropna().values.astype(float)
            if len(arr) < MIN_N:
                continue
            # Lag-1 autocorrelation
            r1  = lag_k_autocorr(arr)
            sig_ac = is_sig_autocorr(r1, len(arr))

            _method_list = [
                (standard_mk, "Standard MK"),
                (modified_mk, "Modified MK"),
            ]
            if _RTA_AVAILABLE:
                _method_list += [
                    (pw_mk,    "PW-MK"),
                    (tfpw_mk,  "TFPW-MK"),
                ]
            for method_fn, method_name in _method_list:
                res = method_fn(arr)
                rows.append({
                    "Station":     stn,
                    "Code":        smap.get(stn, stn),
                    "Scale":       sk,
                    "Scale_Label": meta["label"],
                    "Method":      method_name,
                    "rho_1":       round(r1, 4) if not np.isnan(r1) else np.nan,
                    "Sig_AC":      sig_ac,
                    "N":           res.get("n", np.nan),
                    "S":           res.get("S", np.nan),
                    "Var_S":       res.get("Var_S", np.nan),
                    "Var_S_adj":   res.get("Var_S_adj", np.nan),
                    "n_eff":       res.get("n_eff", np.nan),
                    "Z":           res.get("Z", np.nan),
                    "tau":         res.get("tau", np.nan),
                    "p_value":     res.get("p_value", np.nan),
                    "Trend":       res.get("trend", "—"),
                    "sig_05":      res.get("sig_05", False),
                    "sig_01":      res.get("sig_01", False),
                    "Slope_Q":     res.get("slope_Q", np.nan),
                    "Slope_lo":    res.get("slope_lo", np.nan),
                    "Slope_hi":    res.get("slope_hi", np.nan),
                })

    return pd.DataFrame(rows)


def build_comparison(trend_df: pd.DataFrame) -> pd.DataFrame:
    """
    Build MK vs MMK comparison table per (Station × Scale).
    Highlights differences in trend decision and ΔZ.
    """
    rows = []
    for (stn, sk), grp in trend_df.groupby(["Station","Scale"]):
        mk  = grp[grp["Method"]=="Standard MK"].squeeze()
        mmk = grp[grp["Method"]=="Modified MK"].squeeze()
        if isinstance(mk,  pd.DataFrame) or isinstance(mmk, pd.DataFrame):
            continue
        dZ     = float(mmk["Z"]) - float(mk["Z"])
        dp     = float(mmk["p_value"]) - float(mk["p_value"])
        agree  = mk["Trend"] == mmk["Trend"]
        rows.append({
            "Station":       stn,
            "Code":          mk.get("Code", stn),
            "Scale":         sk,
            "Scale_Label":   mk.get("Scale_Label",""),
            "rho_1":         mk.get("rho_1", np.nan),
            "Sig_AC":        bool(mk.get("Sig_AC", False)),
            "MK_Z":          mk["Z"],    "MK_p":    mk["p_value"],
            "MK_Trend":      mk["Trend"],"MK_sig05":mk["sig_05"],
            "MMK_Z":         mmk["Z"],   "MMK_p":   mmk["p_value"],
            "MMK_Trend":     mmk["Trend"],"MMK_sig05":mmk["sig_05"],
            "delta_Z":       round(dZ, 4),
            "delta_p":       round(dp, 6),
            "Agree":         agree,
            "MK_Slope":      mk["Slope_Q"],
            "MMK_Slope":     mmk["Slope_Q"],
            "Slope_lo":      mmk["Slope_lo"],
            "Slope_hi":      mmk["Slope_hi"],
        })
    return pd.DataFrame(rows)


def build_4method_comparison(trend_df: pd.DataFrame) -> pd.DataFrame:
    """
    Build 4-method wide comparison table for each (Station, Scale).

    Requires trend_df to contain rows for "Standard MK", "Modified MK",
    "PW-MK", and "TFPW-MK".  Missing methods yield NaN columns.

    Returns one row per (Station × Scale) with columns:
      Station, Code, Scale, Scale_Label, rho_1, Sig_AC,
      MK_Z/p/slope/sig/trend, MMK_*, PW_*, TFPW_*,
      dZ_MMK/PW/TFPW, dSlope_MMK/PW/TFPW,
      all_agree, n_sig_methods
    """
    if trend_df is None or len(trend_df) == 0:
        return pd.DataFrame()

    rows = []
    method_keys = {
        "Standard MK": "MK",
        "Modified MK": "MMK",
        "PW-MK":       "PW",
        "TFPW-MK":     "TFPW",
    }

    for (stn, sk), grp in trend_df.groupby(["Station", "Scale"]):
        row = {"Station": stn, "Scale": sk}
        mk_Z = mk_slope = np.nan

        for method_name, prefix in method_keys.items():
            sub = grp[grp["Method"] == method_name]
            if len(sub) == 0:
                row[f"{prefix}_Z"]     = np.nan
                row[f"{prefix}_p"]     = np.nan
                row[f"{prefix}_slope"] = np.nan
                row[f"{prefix}_sig"]   = "—"
                row[f"{prefix}_trend"] = "—"
                continue
            r = sub.iloc[0]
            Z = float(r["Z"]) if not pd.isna(r["Z"]) else np.nan
            p = float(r["p_value"]) if not pd.isna(r["p_value"]) else np.nan
            sl = float(r["Slope_Q"]) if not pd.isna(r["Slope_Q"]) else np.nan
            s01 = bool(r.get("sig_01", False))
            s05 = bool(r.get("sig_05", False))
            sig_str = "**" if s01 else ("*" if s05 else "ns")

            row[f"{prefix}_Z"]     = round(Z, 4) if not np.isnan(Z) else np.nan
            row[f"{prefix}_p"]     = round(p, 6) if not np.isnan(p) else np.nan
            row[f"{prefix}_slope"] = round(sl, 3) if not np.isnan(sl) else np.nan
            row[f"{prefix}_sig"]   = sig_str
            row[f"{prefix}_trend"] = str(r.get("Trend", "—"))

            # store rho_1 and Sig_AC from Standard MK row
            if method_name == "Standard MK":
                row["Code"]       = str(r.get("Code", stn))
                row["Scale_Label"]= str(r.get("Scale_Label", sk))
                row["rho_1"]      = float(r.get("rho_1", np.nan)) \
                                    if not pd.isna(r.get("rho_1", np.nan)) else np.nan
                row["Sig_AC"]     = bool(r.get("Sig_AC", False))
                mk_Z     = row["MK_Z"]
                mk_slope = row["MK_slope"]

        # Delta columns (relative to Standard MK)
        for prefix in ["MMK", "PW", "TFPW"]:
            mZ  = row.get(f"{prefix}_Z",     np.nan)
            mSl = row.get(f"{prefix}_slope", np.nan)
            row[f"dZ_{prefix}"]     = (round(mZ  - mk_Z, 4)
                                        if not (np.isnan(mZ) or np.isnan(mk_Z))
                                        else np.nan)
            row[f"dSlope_{prefix}"] = (round(mSl - mk_slope, 4)
                                        if not (np.isnan(mSl) or np.isnan(mk_slope))
                                        else np.nan)

        # Agreement and significance count
        trends = [row.get(f"{p}_trend", "—") for p in method_keys.values()
                  if row.get(f"{p}_trend", "—") not in ("—", "No trend")]
        all_agree = len(set(trends)) <= 1 if trends else True
        n_sig = sum(
            1 for p in method_keys.values()
            if row.get(f"{p}_sig", "ns") in ("*", "**")
        )
        row["all_agree"]      = all_agree
        row["n_sig_methods"]  = n_sig

        rows.append(row)

    return pd.DataFrame(rows)


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §8  FIGURE 1 — Annual Time Series + Trend                             ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def _sens_line(arr, slope, yrs):
    """Anchored Sen trend line at median."""
    if np.isnan(slope): return None
    y_bar  = float(np.nanmedian(arr))
    x_bar  = float(np.median(yrs))
    return slope * (yrs - x_bar) + y_bar


def _sig_label(sig05, sig01, Z):
    if sig01:  return "**"
    if sig05:  return "*"
    return "ns"


def _col_trend(sig05, Z):
    if sig05 and Z > 0: return C["inc"]
    if sig05 and Z < 0: return C["dec"]
    return C["ns_col"]


def fig1_annual_ts(scales, trend_df, stns, smap, period, out_dir, prefix):
    """Fig 1: Annual rainfall time series per station with MMK trend line."""
    df_ann = scales["annual"]
    stns   = [str(s) for s in stns]
    n_s    = len(stns)
    ncols  = min(4, n_s)
    nrows  = math.ceil(n_s / ncols)

    fig, axes = plt.subplots(nrows, ncols,
                             figsize=(5.5*ncols, 4.0*nrows), squeeze=False)
    fig.subplots_adjust(hspace=0.55, wspace=0.30,
                        top=0.93, bottom=0.07, left=0.06, right=0.97)

    for si, stn in enumerate(stns):
        ax = axes[si//ncols][si%ncols]
        if stn not in df_ann.columns: ax.set_visible(False); continue
        s = df_ann[stn].dropna()
        if len(s) < 4: ax.set_visible(False); continue
        yrs  = s.index.year.values.astype(float)
        vals = s.values.astype(float)

        # retrieve MMK result
        sub = trend_df[(trend_df["Station"]==stn) &
                       (trend_df["Scale"]=="annual") &
                       (trend_df["Method"]=="Modified MK")]
        sig05 = bool(sub["sig_05"].values[0]) if len(sub) else False
        sig01 = bool(sub["sig_01"].values[0]) if len(sub) else False
        Z     = float(sub["Z"].values[0]) if len(sub) else np.nan
        slope = float(sub["Slope_Q"].values[0]) if len(sub) else np.nan
        lo    = float(sub["Slope_lo"].values[0]) if len(sub) else np.nan
        hi    = float(sub["Slope_hi"].values[0]) if len(sub) else np.nan
        p_val = float(sub["p_value"].values[0]) if len(sub) else np.nan
        col   = _col_trend(sig05, Z)
        slab  = _sig_label(sig05, sig01, Z)

        ax.bar(yrs, vals, width=0.75, color=col, alpha=0.45,
               edgecolor="none", zorder=2)
        ax.plot(yrs, vals, color=col, lw=1.5, alpha=0.85, zorder=3)

        tl = _sens_line(vals, slope, yrs)
        if tl is not None:
            ax.plot(yrs, tl, "k-", lw=2.2, zorder=5,
                    label=f"β={slope:+.1f} mm/yr")
            if not (np.isnan(lo) or np.isnan(hi)):
                ll = _sens_line(vals, lo, yrs)
                hl = _sens_line(vals, hi, yrs)
                ax.fill_between(yrs, ll, hl, color="grey",
                                alpha=0.18, zorder=4, label="95% CI")

        code = smap.get(stn, stn)
        ax.set_title(f"({chr(97+si)})  {code}  [{stn}]\n"
                     f"Z={Z:.2f}  p={p_val:.3f}  {slab}",
                     loc="left", fontsize=10.5, fontweight="bold", pad=3)
        ax.set_ylabel("mm yr⁻¹", fontsize=10)
        ax.set_xlabel("Year",    fontsize=10)
        ax.set_ylim(bottom=0)
        ax.yaxis.set_minor_locator(ticker.AutoMinorLocator())
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        ax.legend(fontsize=8, frameon=True, edgecolor="#B0BEC5",
                  loc="upper right", handlelength=1.8)

    for si in range(len(stns), nrows*ncols):
        axes[si//ncols][si%ncols].set_visible(False)

    hand = [mpatches.Patch(color=C["inc"],    label="Increasing (sig.)"),
            mpatches.Patch(color=C["dec"],    label="Decreasing (sig.)"),
            mpatches.Patch(color=C["ns_col"], label="Not significant"),
            Line2D([0],[0],color="black",lw=2.2,label="Sen's slope"),
            mpatches.Patch(color="grey",alpha=0.35,label="95% CI")]
    fig.legend(handles=hand, loc="lower center", ncol=5, fontsize=15,
               markerscale=1.5, frameon=True, edgecolor="#B0BEC5", bbox_to_anchor=(0.5,-0.02))
    #fig.suptitle(
    #    f"Figure 1.  Annual Rainfall Time Series and Trend — {period}\n"
    #    "Modified Mann–Kendall (Hamed & Rao 1998) + Sen's Slope  |"
    #    "  * p<0.05  ** p<0.01  ns: not significant",
    #    fontsize=12, fontweight="bold")
    savefig(fig, str(out_dir/f"{prefix}_Fig1_AnnualTimeSeries"))


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §9  FIGURE 2 — Wet & Dry Season Time Series                           ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def fig2_wetdry_ts(scales, trend_df, stns, smap, period, out_dir, prefix):
    """
    Fig 2: Regional-mean time series for Wet and Dry seasons.
    2×2 grid: (a)Wet regional, (b)Dry regional, (c)Wet per station, (d)Dry per station.
    """
    stns = [str(s) for s in stns]
    fig  = plt.figure(figsize=(18, 13))
    gs   = gridspec.GridSpec(2, 2, figure=fig, hspace=0.48, wspace=0.30,
                             top=0.91, bottom=0.08, left=0.07, right=0.97)
    ax1  = fig.add_subplot(gs[0, 0])   # Wet regional
    ax2  = fig.add_subplot(gs[0, 1])   # Dry regional
    ax3  = fig.add_subplot(gs[1, 0])   # Wet per station
    ax4  = fig.add_subplot(gs[1, 1])   # Dry per station

    for ax, sk, col_l, col_lt, panel, season_ttl in [
        (ax1,"wet", C["wet"], C["wet_lt"], "(a)", "Wet Season (May–Oct)"),
        (ax2,"dry", C["dry"], C["dry_lt"], "(b)", "Dry Season (Nov–Apr)"),
    ]:
        df_s  = scales[sk]
        cols  = [s for s in stns if s in df_s.columns]
        reg   = df_s[cols].mean(axis=1).dropna()
        if len(reg) < 4: continue
        yrs   = reg.index.year.values.astype(float)
        vals  = reg.values.astype(float)

        # Regional MMK
        res  = modified_mk(vals)
        Z    = res["Z"]; p   = res["p_value"]
        slope= res["slope_Q"]; lo = res["slope_lo"]; hi = res["slope_hi"]
        sig05= res["sig_05"]; sig01= res["sig_01"]
        slab = _sig_label(sig05, sig01, Z)
        col  = _col_trend(sig05, Z) if sig05 else col_l

        ax.fill_between(yrs, vals, alpha=0.25, color=col_l, zorder=2)
        ax.plot(yrs, vals, color=col_l, lw=2.0, zorder=3, label=season_ttl)
        tl = _sens_line(vals, slope, yrs)
        if tl is not None:
            ax.plot(yrs, tl, "k-", lw=2.2, zorder=5,
                    label=f"β={slope:+.1f} mm/yr")
            if not (np.isnan(lo) or np.isnan(hi)):
                ax.fill_between(yrs,
                                _sens_line(vals, lo, yrs),
                                _sens_line(vals, hi, yrs),
                                color="grey", alpha=0.18, zorder=4, label="95% CI")
        ax.set_title(
            f"{panel}  Regional Mean — {season_ttl}\n"
            f"MMK: Z={Z:.3f}  p={p:.4f}  {slab}",
            loc="left", fontsize=12, fontweight="bold", pad=5)
        ax.set_ylabel(SCALE_META[sk]["unit"], fontsize=11)
        ax.set_xlabel("Year", fontsize=11)
        ax.set_ylim(bottom=0)
        ax.yaxis.set_minor_locator(ticker.AutoMinorLocator())
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        ax.legend(fontsize=10, frameon=True, edgecolor="#B0BEC5",
                  loc="upper right", handlelength=1.8)

    # Per-station slope bars (panels c, d)
    x = np.arange(len(stns))
    codes = [smap.get(s, s) for s in stns]

    for ax, sk, col_l, col_lt, panel, season_ttl in [
        (ax3,"wet",C["wet"],C["wet_lt"],"(c)","Wet Season — Sen's Slope per Station"),
        (ax4,"dry",C["dry"],C["dry_lt"],"(d)","Dry Season — Sen's Slope per Station"),
    ]:
        slopes_v = []; lo_e = []; hi_e = []; bar_cols = []
        for stn in stns:
            sub = trend_df[(trend_df["Station"]==stn) &
                           (trend_df["Scale"]==sk) &
                           (trend_df["Method"]=="Modified MK")]
            if len(sub) == 0:
                slopes_v.append(np.nan); lo_e.append(0); hi_e.append(0)
                bar_cols.append(C["ns_col"]); continue
            sl  = float(sub["Slope_Q"].values[0])
            lo  = float(sub["Slope_lo"].values[0])
            hi  = float(sub["Slope_hi"].values[0])
            s05 = bool(sub["sig_05"].values[0])
            Z   = float(sub["Z"].values[0])
            slopes_v.append(sl)
            lo_e.append(abs(sl - lo) if not np.isnan(lo) else 0)
            hi_e.append(abs(hi - sl) if not np.isnan(hi) else 0)
            bar_cols.append(_col_trend(s05, Z))

        for xi, (sl, le, he, bc) in enumerate(
                zip(slopes_v, lo_e, hi_e, bar_cols)):
            if np.isnan(sl): continue
            ax.bar(xi, sl, width=0.65, color=bc, alpha=0.82,
                   edgecolor="white", linewidth=0.5, zorder=3)
            ax.errorbar(xi, sl, yerr=[[le],[he]],
                        fmt="none", color="black",
                        capsize=5, capthick=1.5, lw=1.5, zorder=5)
        ax.axhline(0, color="black", lw=0.9, ls="--", alpha=0.45)
        ax.set_xticks(x); ax.set_xticklabels(codes,rotation=0,fontsize=11)
        ax.set_ylabel("β (mm yr⁻¹)", fontsize=11)
        ax.set_xlabel("Station", fontsize=11)
        ax.set_title(f"{panel}  {season_ttl}\n"
                     "Error bars: 95% CI  |  * p<0.05  ** p<0.01",
                     loc="left", fontsize=12, fontweight="bold", pad=5)
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        ax.yaxis.set_minor_locator(ticker.AutoMinorLocator())

        for xi, stn in enumerate(stns):
            sub = trend_df[(trend_df["Station"]==stn) &
                           (trend_df["Scale"]==sk) &
                           (trend_df["Method"]=="Modified MK")]
            if len(sub)==0: continue
            s01=bool(sub["sig_01"].values[0]); s05=bool(sub["sig_05"].values[0])
            sl=float(sub["Slope_Q"].values[0])
            if np.isnan(sl): continue
            sig_s="**" if s01 else ("*" if s05 else "")
            if sig_s:
                ax.text(xi, sl+(1.0 if sl>=0 else -2.0), sig_s,
                        ha="center", fontsize=11, fontweight="bold")

    hand = [mpatches.Patch(color=C["inc"],    label="Increasing (sig.)"),
            mpatches.Patch(color=C["dec"],    label="Decreasing (sig.)"),
            mpatches.Patch(color=C["ns_col"], label="Not significant"),
            Line2D([0],[0],color="black",lw=2.2,label="Sen's slope / 95% CI")]
    fig.legend(handles=hand, loc="lower center", ncol=5, fontsize=15,
               markerscale=1.5, frameon=True, edgecolor="#B0BEC5", bbox_to_anchor=(0.5,-0.02))
    #fig.suptitle(
    #    f"Figure 2.  Hydrological Wet & Dry Season Rainfall Trend — {period}\n"
    #    "Wet: May–Oct  |  Dry: Nov–Apr  |  Modified MK + Sen's Slope",
    #    fontsize=12, fontweight="bold")
    savefig(fig, str(out_dir/f"{prefix}_Fig2_WetDryTimeSeries"))


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §10 FIGURE 3 — Sen's Slope All Scales                                  ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def fig3_sens_all(trend_df, stns, smap, period, out_dir, prefix):
    """Fig 3: Sen's slope bar chart for Annual/Wet/Dry — 3 rows × MMK results."""
    stns  = [str(s) for s in stns]
    codes = [smap.get(s, s) for s in stns]
    n_s   = len(stns)
    x     = np.arange(n_s)
    fig, axes = plt.subplots(3, 1, figsize=(max(14,n_s*1.1+4), 14),
                              sharex=True)
    fig.subplots_adjust(hspace=0.42, top=0.93, bottom=0.09,
                        left=0.07, right=0.97)

    for pi, (sk, col_l) in enumerate(
            [("annual",C["annual"]),("wet",C["wet"]),("dry",C["dry"])]):
        ax = axes[pi]
        meta = SCALE_META[sk]
        slopes_v=[]; lo_e=[]; hi_e=[]; bar_cols=[]
        for stn in stns:
            sub = trend_df[(trend_df["Station"]==stn) &
                           (trend_df["Scale"]==sk) &
                           (trend_df["Method"]=="Modified MK")]
            if len(sub)==0:
                slopes_v.append(np.nan); lo_e.append(0); hi_e.append(0)
                bar_cols.append(C["ns_col"]); continue
            sl=float(sub["Slope_Q"].values[0])
            lo=float(sub["Slope_lo"].values[0])
            hi=float(sub["Slope_hi"].values[0])
            s05=bool(sub["sig_05"].values[0])
            Z=float(sub["Z"].values[0])
            slopes_v.append(sl)
            lo_e.append(abs(sl-lo) if not np.isnan(lo) else 0)
            hi_e.append(abs(hi-sl) if not np.isnan(hi) else 0)
            bar_cols.append(_col_trend(s05, Z))

        for xi,(sl,le,he,bc) in enumerate(zip(slopes_v,lo_e,hi_e,bar_cols)):
            if np.isnan(sl): continue
            ax.bar(xi,sl,width=0.65,color=bc,alpha=0.82,
                   edgecolor="white",linewidth=0.5,zorder=3)
            ax.errorbar(xi,sl,yerr=[[le],[he]],fmt="none",color="black",
                        capsize=5,capthick=1.5,lw=1.5,zorder=5)
            yoff=0.8 if sl>=0 else -1.5
            ax.text(xi,sl+yoff,f"{sl:+.1f}",ha="center",
                    va="bottom" if sl>=0 else "top",fontsize=8.5,fontweight="bold")

        # significance stars
        for xi,stn in enumerate(stns):
            sub=trend_df[(trend_df["Station"]==stn) &
                         (trend_df["Scale"]==sk) &
                         (trend_df["Method"]=="Modified MK")]
            if len(sub)==0: continue
            s01=bool(sub["sig_01"].values[0]); s05=bool(sub["sig_05"].values[0])
            sl=float(sub["Slope_Q"].values[0])
            if np.isnan(sl): continue
            sig_s="**" if s01 else ("*" if s05 else "")
            if sig_s:
                ax.text(xi,sl+(1.5 if sl>=0 else -2.5),sig_s,
                        ha="center",fontsize=11,fontweight="bold",color="black")

        ax.axhline(0, color="black", lw=0.9, ls="--", alpha=0.45)
        ax.set_ylabel(f"β (mm yr⁻¹)\n{meta['label']}",fontsize=11)
        ax.set_title(f"({chr(97+pi)})  {meta['label']} — Sen's Slope (Modified MK)",
                     loc="left",fontsize=12,fontweight="bold",pad=4)
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        ax.yaxis.set_minor_locator(ticker.AutoMinorLocator())

    axes[-1].set_xticks(x)
    axes[-1].set_xticklabels(codes,rotation=0,ha="center",fontsize=11)
    axes[-1].set_xlabel("Station",fontsize=12,labelpad=5)
    hand=[mpatches.Patch(color=C["inc"],label="Increasing (sig.)"),
          mpatches.Patch(color=C["dec"],label="Decreasing (sig.)"),
          mpatches.Patch(color=C["ns_col"],label="Not significant"),
          Line2D([0],[0],color="black",lw=1.8,marker="|",ms=8,label="95% CI")]
    fig.legend(handles=hand, loc="lower center", ncol=5, fontsize=15,
               markerscale=1.5, frameon=True, edgecolor="#B0BEC5", bbox_to_anchor=(0.5,-0.02))
    #fig.suptitle(
    #    f"Figure 3.  Sen's Slope Estimator — Annual, Wet, and Dry Season  |  {period}\n"
    #    "Error bars: 95% CI (Gilbert 1987)  |  * p<0.05  ** p<0.01",
    #    fontsize=12,fontweight="bold")
    savefig(fig, str(out_dir/f"{prefix}_Fig3_SenSlope_AllScales"))


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §11 FIGURE 4 — Standard MK vs Modified MK Comparison                  ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def fig4_mk_vs_mmk(comp_df, stns, smap, period, out_dir, prefix):
    """
    Fig 4: Side-by-side comparison of Standard MK and Modified MK.
    (a) Z-statistic scatter: MK Z vs MMK Z (diagonal = no change).
    (b) p-value scatter: MK p vs MMK p with α=0.05 threshold lines.
    (c) ΔZ (MMK−MK) per station × scale bar chart.
    (d) Agreement table heatmap.
    """
    stns  = [str(s) for s in stns]
    codes = [smap.get(s, s) for s in stns]
    scales_plot = ["annual","wet","dry"]
    scale_labels= [SCALE_META[sk]["label"] for sk in scales_plot]

    fig = plt.figure(figsize=(20, 14))
    gs  = gridspec.GridSpec(2, 2, figure=fig, hspace=0.48, wspace=0.32,
                            top=0.91, bottom=0.08, left=0.07, right=0.97)
    ax1 = fig.add_subplot(gs[0, 0])   # Z scatter
    ax2 = fig.add_subplot(gs[0, 1])   # p scatter
    ax3 = fig.add_subplot(gs[1, 0])   # ΔZ bar
    ax4 = fig.add_subplot(gs[1, 1])   # Agreement heatmap

    # Colour per scale
    sc_col = {"annual":C["annual"], "wet":C["wet"], "dry":C["dry"]}

    for sk in scales_plot:
        sub = comp_df[comp_df["Scale"]==sk]
        if len(sub)==0: continue
        ax1.scatter(sub["MK_Z"], sub["MMK_Z"],
                    color=sc_col[sk], s=80, alpha=0.80, zorder=4,
                    edgecolors="white", linewidth=0.6,
                    label=SCALE_META[sk]["label"])
        ax2.scatter(sub["MK_p"], sub["MMK_p"],
                    color=sc_col[sk], s=80, alpha=0.80, zorder=4,
                    edgecolors="white", linewidth=0.6)

    # Panel A: Z scatter
    z_all = pd.concat([comp_df["MK_Z"], comp_df["MMK_Z"]]).dropna()
    zmax  = max(abs(z_all.min()), abs(z_all.max()), 0.5)
    ax1.plot([-zmax,zmax],[-zmax,zmax],"k--",lw=1.3,alpha=0.55,
             label="1:1 (no change)")
    ax1.axhline(0,color="grey",lw=0.7,ls=":",alpha=0.5)
    ax1.axvline(0,color="grey",lw=0.7,ls=":",alpha=0.5)
    ax1.axhline( Z_005,color="orange",lw=1.2,ls="--",alpha=0.7,label=f"±Z₀.₀₅={Z_005}")
    ax1.axhline(-Z_005,color="orange",lw=1.2,ls="--",alpha=0.7)
    ax1.set_xlim(-zmax,zmax); ax1.set_ylim(-zmax,zmax)
    ax1.set_xlabel("Standard MK Z-statistic", fontsize=12)
    ax1.set_ylabel("Modified MK Z-statistic", fontsize=12)
    ax1.set_title("(a)  Z-Statistic: Standard MK vs Modified MK\n"
                  "     Points above 1:1 → autocorr. inflated MK Z",
                  loc="left",fontsize=12,fontweight="bold",pad=5)
    ax1.legend(fontsize=9.5,frameon=True,edgecolor="#B0BEC5",loc="upper left")
    ax1.set_aspect("equal",adjustable="box")
    ax1.spines["top"].set_visible(False); ax1.spines["right"].set_visible(False)

    # Panel B: p scatter
    ax2.plot([0,1],[0,1],"k--",lw=1.3,alpha=0.55,label="1:1 (no change)")
    ax2.axhline(ALPHA_005,color="orange",lw=1.2,ls="--",alpha=0.7,
                label=f"α=0.05")
    ax2.axvline(ALPHA_005,color="orange",lw=1.2,ls="--",alpha=0.7)
    ax2.set_xlim(0,1); ax2.set_ylim(0,1)
    ax2.set_xlabel("Standard MK p-value", fontsize=12)
    ax2.set_ylabel("Modified MK p-value", fontsize=12)
    ax2.set_title("(b)  p-Value: Standard MK vs Modified MK\n"
                  "     Points above 1:1 → MMK more conservative",
                  loc="left",fontsize=12,fontweight="bold",pad=5)
    # annotate significant points
    for _, row in comp_df.iterrows():
        mk_p=float(row["MK_p"]); mmk_p=float(row["MMK_p"])
        if mk_p<0.15 or mmk_p<0.15:
            code=row.get("Code",row["Station"])
            ax2.annotate(f"{code}",xy=(mk_p,mmk_p),
                         fontsize=7.5,color="black",alpha=0.7)
    for sk in scales_plot:
        ax2.scatter([],[], color=sc_col[sk], s=60, label=SCALE_META[sk]["label"])
    ax2.legend(fontsize=9.5,frameon=True,edgecolor="#B0BEC5",loc="upper left")
    ax2.set_aspect("equal",adjustable="box")
    ax2.spines["top"].set_visible(False); ax2.spines["right"].set_visible(False)

    # Panel C: ΔZ per station×scale
    n_s=len(stns); x=np.arange(n_s)
    bw = 0.25
    for di,(sk,off) in enumerate(zip(scales_plot,
                                     [bw*(-1), 0, bw*(1)])):
        sub   = comp_df[comp_df["Scale"]==sk]
        dz_s  = {str(r["Station"]): r["delta_Z"] for _,r in sub.iterrows()}
        dz_arr= [dz_s.get(s,np.nan) for s in stns]
        bars  = ax3.bar(x+off, dz_arr, width=bw*0.88,
                        color=sc_col[sk], alpha=0.82,
                        edgecolor="white", linewidth=0.4,
                        label=SCALE_META[sk]["label"], zorder=3)
    ax3.axhline(0,color="black",lw=0.9,ls="--",alpha=0.5)
    ax3.set_xticks(x); ax3.set_xticklabels(codes,rotation=0,fontsize=11)
    ax3.set_ylabel("ΔZ  =  Z_MMK − Z_MK", fontsize=11)
    ax3.set_xlabel("Station", fontsize=11)
    ax3.set_title("(c)  ΔZ (MMK − Standard MK) per Station\n"
                  "     ΔZ < 0 → autocorrelation reduced |Z|  "
                  "(positive autocorr. inflates standard MK)",
                  loc="left",fontsize=12,fontweight="bold",pad=5)
    ax3.legend(fontsize=10,frameon=True,edgecolor="#B0BEC5",loc="upper right")
    ax3.spines["top"].set_visible(False); ax3.spines["right"].set_visible(False)
    ax3.yaxis.set_minor_locator(ticker.AutoMinorLocator())

    # Panel D: Agreement heatmap (station × scale)
    agree_mat = np.zeros((len(scales_plot), n_s), dtype=float)
    for si, stn in enumerate(stns):
        for sci, sk in enumerate(scales_plot):
            sub = comp_df[(comp_df["Station"]==stn) &
                          (comp_df["Scale"]==sk)]
            if len(sub)==0: agree_mat[sci,si]=np.nan; continue
            agree_mat[sci,si]=1.0 if bool(sub["Agree"].values[0]) else 0.0

    im = ax4.imshow(agree_mat, cmap="RdYlGn", vmin=0, vmax=1,
                    aspect="auto", interpolation="nearest")
    plt.colorbar(im, ax=ax4, orientation="horizontal",
                 pad=0.22, fraction=0.06, shrink=0.8,
                 label="Agreement (1=Same Trend, 0=Different)")
    for si in range(n_s):
        for sci,sk in enumerate(scales_plot):
            v=agree_mat[sci,si]
            if np.isnan(v): continue
            sub=comp_df[(comp_df["Station"]==stns[si]) &
                        (comp_df["Scale"]==sk)]
            if len(sub)==0: continue
            dz=float(sub["delta_Z"].values[0])
            tc="white" if abs(v-0.5)<0.4 else "black"
            ax4.text(si,sci,f"ΔZ={dz:+.2f}",ha="center",va="center",
                     fontsize=8.5,fontweight="bold",color=tc)
    ax4.set_xticks(range(n_s)); ax4.set_xticklabels(codes,fontsize=11,rotation=0)
    ax4.set_yticks(range(len(scales_plot)))
    ax4.set_yticklabels(scale_labels, fontsize=11)
    ax4.set_xlabel("Station", fontsize=11); ax4.set_ylabel("Scale", fontsize=11)
    ax4.set_title("(d)  Agreement Heatmap: Standard MK vs Modified MK\n"
                  "     Cell: ΔZ value  |  Green=agree, Red=disagree",
                  loc="left",fontsize=12,fontweight="bold",pad=5)

    #fig.suptitle(
    #    f"Figure 4.  Standard MK vs Modified MK Comparison — {period}\n"
    #    "Identifies where serial autocorrelation changes trend conclusions",
    #    fontsize=12, fontweight="bold")
    savefig(fig, str(out_dir/f"{prefix}_Fig4_MK_vs_MMK_Comparison"))


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §12 FIGURE 5 — Significance Heatmap (both methods)                     ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def fig5_significance_heatmap(trend_df, stns, smap, period, out_dir, prefix):
    """Fig 5: Z-stat heatmap (station × scale) for Standard MK and Modified MK."""
    stns   = [str(s) for s in stns]
    codes  = [smap.get(s, s) for s in stns]
    scales = ["annual","wet","dry"]
    n_s    = len(stns)

    fig, axes = plt.subplots(2, 1, figsize=(max(14,n_s+6), 13))
    fig.subplots_adjust(hspace=0.55, top=0.91, bottom=0.10,
                        left=0.12, right=0.97)

    for ai, (method, ax) in enumerate([("Standard MK",axes[0]),
                                        ("Modified MK",axes[1])]):
        Z_mat = np.full((len(scales), n_s), np.nan)
        p_mat = np.full((len(scales), n_s), np.nan)
        sl_mat= np.full((len(scales), n_s), np.nan)
        sg_mat= np.zeros((len(scales), n_s), dtype=int)

        for si, stn in enumerate(stns):
            for sci, sk in enumerate(scales):
                sub = trend_df[(trend_df["Station"]==stn) &
                               (trend_df["Scale"]==sk) &
                               (trend_df["Method"]==method)]
                if len(sub)==0: continue
                Z_mat[sci,si]  = float(sub["Z"].values[0])
                p_mat[sci,si]  = float(sub["p_value"].values[0])
                sl_mat[sci,si] = float(sub["Slope_Q"].values[0])
                if bool(sub["sig_01"].values[0]): sg_mat[sci,si]=2
                elif bool(sub["sig_05"].values[0]): sg_mat[sci,si]=1

        abs_max=max(np.nanmax(np.abs(Z_mat)) if not np.all(np.isnan(Z_mat)) else 3, Z_001+0.5)
        im=ax.imshow(Z_mat,cmap="RdBu_r",vmin=-abs_max,vmax=abs_max,
                     aspect="auto",interpolation="nearest")
        cbar=plt.colorbar(im,ax=ax,orientation="vertical",
                          pad=0.02,fraction=0.03,shrink=0.95)
        cbar.set_label("Z-statistic",fontsize=10)
        cbar.ax.axhline( Z_005,color="orange",lw=1.4,ls="--")
        cbar.ax.axhline(-Z_005,color="orange",lw=1.4,ls="--")
        cbar.ax.axhline( Z_001,color="red",   lw=1.4,ls="-")
        cbar.ax.axhline(-Z_001,color="red",   lw=1.4,ls="-")

        for sci in range(len(scales)):
            for si in range(n_s):
                Z_v=Z_mat[sci,si]; p_v=p_mat[sci,si]; sl_v=sl_mat[sci,si]
                if np.isnan(Z_v): continue
                lp=abs(Z_v)/abs_max
                tc="white" if lp>0.70 else "black"
                sig_s=("**" if sg_mat[sci,si]==2 else
                        "*"  if sg_mat[sci,si]==1 else "ns")
                ax.text(si,sci,
                        f"Z={Z_v:.2f}\nβ={sl_v:+.1f}\np={p_v:.3f} {sig_s}",
                        ha="center",va="center",fontsize=7.5,
                        fontweight="bold" if sg_mat[sci,si]>0 else "normal",
                        color=tc,linespacing=1.35)

        ax.set_xticks(range(n_s)); ax.set_xticklabels(codes,fontsize=11,rotation=0)
        ax.set_yticks(range(len(scales)))
        ax.set_yticklabels([SCALE_META[sk]["label"] for sk in scales],fontsize=11)
        ax.set_xlabel("Station",fontsize=11,labelpad=4)
        ax.set_title(f"({chr(97+ai)})  {method} — Z-Statistic Heatmap\n"
                     "     Cell: Z | β (mm/yr) | p | significance",
                     loc="left",fontsize=12,fontweight="bold",pad=5)

    #fig.suptitle(
    #    f"Figure 5.  Trend Significance Heatmap — {period}\n"
    #    "Standard MK (top) vs Modified MK (bottom)  |"
    #    "  * p<0.05  ** p<0.01  |  Blue=increasing  Red=decreasing",
    #    fontsize=12,fontweight="bold")
    savefig(fig, str(out_dir/f"{prefix}_Fig5_Significance_Heatmap"))


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §13 FIGURE 6 — Autocorrelation                                         ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def fig6_autocorrelation(scales, stns, smap, period, out_dir, prefix):
    """Fig 6: Lag-1 r₁ per station (3 scales) + ACF for regional mean."""
    stns  = [str(s) for s in stns]
    codes = [smap.get(s, s) for s in stns]
    n_s   = len(stns)
    sc_list = ["annual","wet","dry"]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 7))
    fig.subplots_adjust(left=0.07,right=0.97,top=0.88,bottom=0.12,wspace=0.30)

    # ── Panel A: Lag-1 r₁ grouped bars ───────────────────────────────────
    x = np.arange(n_s); bw = 0.25
    offsets = [bw*(-1), 0, bw*(1)]
    for di,(sk,off) in enumerate(zip(sc_list, offsets)):
        df_s = scales[sk]
        r1_v = [lag_k_autocorr(df_s[s].dropna().values.astype(float))
                if s in df_s.columns else np.nan for s in stns]
        col  = SCALE_META[sk]["color"]
        bars = ax1.bar(x+off, [r if not np.isnan(r) else 0 for r in r1_v],
                       width=bw*0.88, color=col, alpha=0.82,
                       edgecolor="white", linewidth=0.4,
                       label=SCALE_META[sk]["label"], zorder=3)
        for xi,(r,s) in enumerate(zip(r1_v,stns)):
            if np.isnan(r): continue
            sig=is_sig_autocorr(r, int(df_s[s].dropna().__len__()))
            if sig:
                ax1.text(x[xi]+off, r+(0.02 if r>=0 else -0.04),
                         "*",ha="center",fontsize=11,fontweight="bold",
                         color=col)

    # 95% band (Bartlett, approximate n=34)
    n_approx = int(np.mean([len(scales["annual"][s].dropna())
                             for s in stns if s in scales["annual"].columns]))
    ci95 = scipy_norm.ppf(0.975) / math.sqrt(n_approx)
    ax1.axhline( ci95,color="red",lw=1.5,ls="--",label=f"95% band (±{ci95:.3f})")
    ax1.axhline(-ci95,color="red",lw=1.5,ls="--")
    ax1.axhline(0,color="black",lw=0.8,ls="-",alpha=0.4)
    ax1.set_xticks(x); ax1.set_xticklabels(codes,rotation=0,fontsize=11)
    ax1.set_ylabel("Lag-1 Autocorrelation (r₁)",fontsize=12)
    ax1.set_xlabel("Station",fontsize=12)
    ax1.set_ylim(-1,1)
    ax1.set_title("(a)  Lag-1 Autocorrelation — Annual, Wet, Dry\n"
                  "     * = significant (α=0.05)  |  Dashed: 95% threshold",
                  loc="left",fontsize=12,fontweight="bold",pad=5)
    ax1.legend(fontsize=10,frameon=True,edgecolor="#B0BEC5",loc="upper right")
    ax1.spines["top"].set_visible(False); ax1.spines["right"].set_visible(False)

    # ── Panel B: ACF of regional mean annual ─────────────────────────────
    cols    = [s for s in stns if s in scales["annual"].columns]
    reg_ann = scales["annual"][cols].mean(axis=1).dropna().values.astype(float)
    n_reg   = len(reg_ann)
    max_lag = min(10, n_reg // 3)
    rho_all = all_lag_autocorr(reg_ann, max_lag=max_lag)
    lags    = np.arange(1, len(rho_all)+1)
    ci_acf  = scipy_norm.ppf(0.975) / math.sqrt(n_reg)

    ax2.bar(lags, rho_all, width=0.65, color=C["annual"],
            alpha=0.75, edgecolor="white", linewidth=0.5, zorder=3)
    ax2.axhline( ci_acf,color="red",lw=1.5,ls="--",label=f"95% CI (±{ci_acf:.3f})")
    ax2.axhline(-ci_acf,color="red",lw=1.5,ls="--")
    ax2.axhline(0,color="black",lw=0.8,ls="-",alpha=0.4)
    for lg,rho_v in zip(lags,rho_all):
        if abs(rho_v)>ci_acf:
            ax2.text(lg,rho_v+(0.03 if rho_v>=0 else -0.05),
                     f"{rho_v:.2f}*",ha="center",va="bottom" if rho_v>=0 else "top",
                     fontsize=9,fontweight="bold",color=C["dec"])
    ax2.set_xticks(lags)
    ax2.set_xlabel("Lag (years)",fontsize=12)
    ax2.set_ylabel("Autocorrelation Coefficient (rₖ)",fontsize=12)
    ax2.set_ylim(-1,1)
    ax2.set_title(f"(b)  ACF — Regional Mean Annual Rainfall (Lag 1–{max_lag})\n"
                  "     Significant rₖ → Modified MK essential",
                  loc="left",fontsize=12,fontweight="bold",pad=5)
    ax2.legend(fontsize=10,frameon=True,edgecolor="#B0BEC5")
    ax2.spines["top"].set_visible(False); ax2.spines["right"].set_visible(False)

    #fig.suptitle(
    #    f"Figure 6.  Autocorrelation Diagnostics — {period}\n"
    #    "Significant autocorrelation → Modified MK required  "
    #    "(Hamed & Rao 1998)",
    #    fontsize=12,fontweight="bold")
    savefig(fig, str(out_dir/f"{prefix}_Fig6_Autocorrelation"))


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §14 FIGURE 7 — Monthly Climatology                                     ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def fig7_monthly_climatology(scales, stns, smap, period, out_dir, prefix):
    """
    Fig 7: Mean monthly rainfall per station (line) + regional mean (bars).
    Clearly shows wet/dry season partitioning.
    """
    stns = [str(s) for s in stns]
    df_m = scales["monthly_all"]
    n_s  = len(stns)

    fig, axes = plt.subplots(math.ceil(n_s/4)+1, min(4,n_s),
                              figsize=(18, 4*(math.ceil(n_s/4)+1)),
                              squeeze=False)
    fig.subplots_adjust(hspace=0.55, wspace=0.30,
                        top=0.93, bottom=0.06, left=0.06, right=0.97)

    # Regional mean monthly
    cols    = [s for s in stns if s in df_m.columns]
    reg_m   = df_m[cols].copy()
    reg_m["month"] = reg_m.index.month
    clim_reg= reg_m.groupby("month")[cols].mean().mean(axis=1)

    # Ax0: Regional mean
    ax0 = axes[0][0]
    months = np.arange(1,13)
    bar_col = [C["wet_lt"] if m in WET_MONTHS else C["dry_lt"] for m in months]
    ax0.bar(months, clim_reg.values, color=bar_col, edgecolor="grey",
            linewidth=0.6, alpha=0.85, zorder=3)
    ax0.set_xticks(months)
    ax0.set_xticklabels(MONTH_ABBR, rotation=0, fontsize=9)
    ax0.set_ylabel("Mean Monthly\nRainfall (mm)", fontsize=10)
    ax0.set_title("(a)  Regional Mean Monthly Climatology\n"
                  "     Blue=Wet (May–Oct)  |  Orange=Dry (Nov–Apr)",
                  loc="left", fontsize=10.5, fontweight="bold", pad=4)
    ax0.set_ylim(bottom=0)
    ax0.spines["top"].set_visible(False); ax0.spines["right"].set_visible(False)

    # Axhline to mark wet/dry boundary
    for ax_ref in [ax0]:
        for m in [4.5, 10.5]:
            ax_ref.axvline(m, color="black", lw=1.0, ls="--", alpha=0.5)

    # Per-station panels
    for si, stn in enumerate(stns):
        row = (si + 1) // 4;  col_i = (si + 1) % 4
        ax  = axes[row][col_i] if row < axes.shape[0] else None
        if ax is None: continue
        if stn not in df_m.columns: ax.set_visible(False); continue

        stn_m = df_m[[stn]].copy()
        stn_m["month"] = stn_m.index.month
        clim_s = stn_m.groupby("month")[stn].mean()
        bar_c2 = [C["wet"] if m in WET_MONTHS else C["dry"] for m in months]
        ax.bar(months, clim_s.values, color=bar_c2, edgecolor="white",
               linewidth=0.4, alpha=0.78, zorder=3)
        ax.set_xticks(months)
        ax.set_xticklabels(MONTH_ABBR, rotation=45, fontsize=7.5)
        ax.set_ylabel("mm", fontsize=9, labelpad=2)
        ax.set_title(f"({chr(97+si+1)})  {smap.get(stn,stn)} [{stn}]",
                     loc="left", fontsize=9.5, fontweight="bold", pad=3)
        ax.set_ylim(bottom=0)
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)

    for idx in range(len(stns)+1, axes.size):
        axes.flat[idx].set_visible(False)

    # wet/dry legend
    hand = [mpatches.Patch(color=C["wet_lt"],label="Wet Season (May–Oct)"),
            mpatches.Patch(color=C["dry_lt"],label="Dry Season (Nov–Apr)")]
    fig.legend(handles=hand, loc="lower center", ncol=5, fontsize=15,
               markerscale=1.5, frameon=True, edgecolor="#B0BEC5", bbox_to_anchor=(0.5,-0.02))
    #fig.suptitle(
    #    f"Figure 7.  Monthly Rainfall Climatology — {period}\n"
    #    "Mean 1981–2014  |  Wet Season: May–Oct  |  Dry Season: Nov–Apr",
    #    fontsize=12,fontweight="bold")
    savefig(fig, str(out_dir/f"{prefix}_Fig7_MonthlyClimatology"))


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §15 FIGURE 8 — Spatial Trend Summary (bubble/matrix)                  ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def fig8_spatial_summary(trend_df, comp_df, stns, smap, period,
                          out_dir, prefix):
    """
    Fig 8: Multi-panel spatial trend summary.
    (a) Bubble chart: x=Sen's slope, y=Z, size=n, colour=scale.
    (b) Stacked bar: % increasing / decreasing / no-trend per scale+method.
    (c) ΔSlope (Wet−Dry) per station.
    (d) Slope ratio heatmap (station × scale × method).
    """
    stns  = [str(s) for s in stns]
    codes = [smap.get(s, s) for s in stns]
    n_s   = len(stns)
    scales_plot = ["annual","wet","dry"]

    fig = plt.figure(figsize=(20,14))
    gs  = gridspec.GridSpec(2,2,figure=fig,hspace=0.48,wspace=0.32,
                            top=0.91,bottom=0.08,left=0.07,right=0.97)
    ax1=fig.add_subplot(gs[0,0]); ax2=fig.add_subplot(gs[0,1])
    ax3=fig.add_subplot(gs[1,0]); ax4=fig.add_subplot(gs[1,1])

    # ── Panel A: Bubble — slope vs Z ────────────────────────────────────
    for sk in scales_plot:
        sub=trend_df[(trend_df["Scale"]==sk) &
                     (trend_df["Method"]=="Modified MK")]
        ax1.scatter(sub["Slope_Q"], sub["Z"],
                    s=80, color=SCALE_META[sk]["color"],
                    alpha=0.80, edgecolors="white", linewidth=0.6,
                    label=SCALE_META[sk]["label"], zorder=4)
    ax1.axhline( Z_005,color="orange",lw=1.2,ls="--",alpha=0.7,label=f"±Z₀.₀₅")
    ax1.axhline(-Z_005,color="orange",lw=1.2,ls="--",alpha=0.7)
    ax1.axhline(0,color="grey",lw=0.8,ls=":",alpha=0.5)
    ax1.axvline(0,color="grey",lw=0.8,ls=":",alpha=0.5)
    ax1.set_xlabel("Sen's Slope β (mm yr⁻¹)",fontsize=12)
    ax1.set_ylabel("Modified MK Z-statistic",fontsize=12)
    ax1.set_title("(a)  Sen's Slope vs Z-Statistic\n"
                  "     Modified MK  |  All scales",
                  loc="left",fontsize=12,fontweight="bold",pad=5)
    ax1.legend(fontsize=10,frameon=True,edgecolor="#B0BEC5",loc="upper left")
    ax1.spines["top"].set_visible(False); ax1.spines["right"].set_visible(False)

    # ── Panel B: Stacked bar — trend counts ──────────────────────────────
    method_list = ["Standard MK","Modified MK"]
    cats  = ["Increasing","Decreasing","No trend"]
    cat_col={"Increasing":C["inc"],"Decreasing":C["dec"],"No trend":C["ns_col"]}
    bar_labs = []
    for sk in scales_plot:
        for meth in method_list:
            bar_labs.append(f"{SCALE_META[sk]['label'][:3]}\n{meth[:3]}")
    x2 = np.arange(len(bar_labs))

    for cat_i, cat in enumerate(cats):
        counts = []
        for sk in scales_plot:
            for meth in method_list:
                sub=trend_df[(trend_df["Scale"]==sk) &
                             (trend_df["Method"]==meth)]
                n_cat = sub["sig_05"].sum() if cat=="Increasing" else 0
                if cat=="Decreasing":
                    n_cat=int(((sub["sig_05"]==True) & (sub["Z"]<0)).sum())
                elif cat=="Increasing":
                    n_cat=int(((sub["sig_05"]==True) & (sub["Z"]>0)).sum())
                elif cat=="No trend":
                    n_cat=int((sub["sig_05"]==False).sum())
                counts.append(n_cat)
        if cat_i==0:
            bottom=np.zeros(len(bar_labs))
        ax2.bar(x2, counts, bottom=bottom, color=cat_col[cat],
                alpha=0.85, edgecolor="white", linewidth=0.5,
                label=cat, zorder=3)
        for xi,(c,b) in enumerate(zip(counts,bottom)):
            if c>0:
                ax2.text(xi,b+c/2,str(c),ha="center",va="center",
                         fontsize=9,fontweight="bold",color="white")
        bottom += np.array(counts)

    ax2.set_xticks(x2); ax2.set_xticklabels(bar_labs,fontsize=8,rotation=0)
    ax2.set_ylabel("Number of stations",fontsize=11)
    ax2.set_title("(b)  Trend Count by Scale and Method\n"
                  "     (Increasing / Decreasing / No trend at p<0.05)",
                  loc="left",fontsize=12,fontweight="bold",pad=5)
    ax2.legend(fontsize=10,frameon=True,edgecolor="#B0BEC5",loc="upper right")
    ax2.spines["top"].set_visible(False); ax2.spines["right"].set_visible(False)
    ax2.set_ylim(bottom=0)

    # ── Panel C: ΔSlope (Wet − Dry) ─────────────────────────────────────
    wet_sl = {str(r["Station"]):r["Slope_Q"]
               for _,r in trend_df[(trend_df["Scale"]=="wet") &
                                    (trend_df["Method"]=="Modified MK")].iterrows()}
    dry_sl = {str(r["Station"]):r["Slope_Q"]
               for _,r in trend_df[(trend_df["Scale"]=="dry") &
                                    (trend_df["Method"]=="Modified MK")].iterrows()}
    delta  = [wet_sl.get(s,np.nan)-dry_sl.get(s,np.nan) for s in stns]
    col_d  = [C["wet"] if d>0 else C["dry"] for d in delta]
    ax3.bar(range(n_s), delta, width=0.65, color=col_d, alpha=0.82,
            edgecolor="white", linewidth=0.5, zorder=3)
    ax3.axhline(0,color="black",lw=0.9,ls="--",alpha=0.45)
    for xi,d in enumerate(delta):
        if np.isnan(d): continue
        ax3.text(xi,d+(0.5 if d>=0 else -1.0),f"{d:+.1f}",
                 ha="center",va="bottom" if d>=0 else "top",
                 fontsize=8.5,fontweight="bold")
    ax3.set_xticks(range(n_s)); ax3.set_xticklabels(codes,rotation=0,fontsize=11)
    ax3.set_ylabel("ΔSlope = β_Wet − β_Dry  (mm yr⁻¹)",fontsize=11)
    ax3.set_xlabel("Station",fontsize=11)
    ax3.set_title("(c)  Wet–Dry Slope Difference (ΔSlope)\n"
                  "     Blue > 0: stronger wet-season trend  |  "
                  "Orange < 0: stronger dry-season trend",
                  loc="left",fontsize=12,fontweight="bold",pad=5)
    ax3.spines["top"].set_visible(False); ax3.spines["right"].set_visible(False)
    ax3.yaxis.set_minor_locator(ticker.AutoMinorLocator())

    # ── Panel D: Slope ratio heatmap (method × scale, station mean) ──────
    meth_list2 = ["Standard MK","Modified MK"]
    mat = np.full((len(meth_list2)*3, n_s), np.nan)
    ylabels = []
    for ri,(meth,sk) in enumerate(
            [(m,sk2) for sk2 in scales_plot for m in meth_list2]):
        sub=trend_df[(trend_df["Scale"]==sk) & (trend_df["Method"]==meth)]
        sl_dict={str(r["Station"]):r["Slope_Q"] for _,r in sub.iterrows()}
        for si,stn in enumerate(stns):
            mat[ri,si]=sl_dict.get(stn,np.nan)
        ylabels.append(f"{SCALE_META[sk]['label'][:10]}\n({meth[:3]})")

    abs_m=np.nanmax(np.abs(mat)) if not np.all(np.isnan(mat)) else 10
    im=ax4.imshow(mat,cmap="RdYlGn",vmin=-abs_m,vmax=abs_m,
                  aspect="auto",interpolation="nearest")
    plt.colorbar(im,ax=ax4,orientation="horizontal",
                 pad=0.20,fraction=0.06,shrink=0.85,
                 label="Sen's Slope β (mm yr⁻¹)")
    for ri in range(mat.shape[0]):
        for si in range(n_s):
            v=mat[ri,si]
            if np.isnan(v): continue
            lp=v/abs_m
            tc="white" if abs(lp)>0.70 else "black"
            ax4.text(si,ri,f"{v:+.1f}",ha="center",va="center",
                     fontsize=8.5,fontweight="bold",color=tc)
    ax4.set_xticks(range(n_s)); ax4.set_xticklabels(codes,fontsize=11,rotation=0)
    ax4.set_yticks(range(len(ylabels))); ax4.set_yticklabels(ylabels,fontsize=9)
    ax4.set_xlabel("Station",fontsize=11)
    ax4.set_title("(d)  Sen's Slope Heatmap — All Methods × Scales\n"
                  "     Green=increasing, Red=decreasing",
                  loc="left",fontsize=12,fontweight="bold",pad=5)

    #fig.suptitle(
    #    f"Figure 8.  Spatial Trend Summary — {period}\n"
    #    "Modified Mann–Kendall + Sen's Slope  |  Annual / Wet / Dry Season",
    #    fontsize=12,fontweight="bold")
    savefig(fig, str(out_dir/f"{prefix}_Fig8_SpatialTrend_Summary"))


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §16 EXCEL OUTPUT — 6 SHEETS                                            ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def write_excel(out_xlsx, stns, smap, trend_df, comp_df,
                desc_df, qc_dict, period):
    """
    6 Excel sheets:
      S1 — Standard MK Results
      S2 — Modified MK Results (Hamed & Rao 1998)
      S3 — MK vs MMK Comparison
      S4 — Sen's Slope Summary (with 95% CI)
      S5 — Descriptive Statistics
      S6 — Methods & References
    """
    stns  = [str(s) for s in stns]
    wb    = Workbook(); wb.remove(wb.active)

    def _title(ws, nc, t1, t2=""):
        mxsc(ws,1,1,nc,t1,bold=True,fc="FFFFFF",bg=XC["title"],sz=12,align="left")
        rh(ws,1,24)
        if t2:
            mxsc(ws,2,1,nc,t2,italic=True,fc="FFFFFF",bg=XC["sub"],sz=9)
            rh(ws,2,14)

    def _hdr(ws, r, hdrs, bg=XC["hdr"]):
        for ci,h in enumerate(hdrs,1):
            xsc(ws,r,ci,h,bold=True,fc="FFFFFF",bg=bg,border=tb(),sz=9,wrap=True)
        rh(ws,r,40)

    sc_bg = {"annual":XC["ann_h"],"wet":XC["wet_h"],"dry":XC["dry_h"]}

    def _write_trend_sheet(ws, method_filter, t1, t2):
        sub_df = trend_df[trend_df["Method"]==method_filter].reset_index(drop=True)
        is_mmk = "Modified" in method_filter
        nc = 15 if is_mmk else 13
        _title(ws, nc, t1, t2)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "E4"
        if is_mmk:
            hdr = ["Station","Code","Scale","N","S","Var(S)","Var*(S)","n_eff",
                   "ρ₁","Z","τ (Kendall)","p-value","Trend","* p<0.05","** p<0.01"]
        else:
            hdr = ["Station","Code","Scale","N","S","Var(S)",
                   "Z","τ (Kendall)","p-value","Trend","* p<0.05","** p<0.01","rho_1"]
        _hdr(ws,3,hdr)
        ri=4
        for _,row in sub_df.iterrows():
            sk=row["Scale"]; s05=bool(row["sig_05"]); s01=bool(row["sig_01"])
            bg=sc_bg.get(sk,XC["ann_h"])
            if s01: bg=XC["sig01"]
            elif s05: bg=XC["sig05"]
            if is_mmk:
                vals=[str(row["Station"]),str(row["Code"]),row["Scale_Label"],
                      row["N"],row["S"],row["Var_S"],row["Var_S_adj"],row["n_eff"],
                      row["rho_1"],row["Z"],row["tau"],row["p_value"],
                      str(row["Trend"]),"**" if s01 else ("*" if s05 else "ns"),
                      "**" if s01 else "ns"]
            else:
                vals=[str(row["Station"]),str(row["Code"]),row["Scale_Label"],
                      row["N"],row["S"],row["Var_S"],
                      row["Z"],row["tau"],row["p_value"],
                      str(row["Trend"]),"**" if s01 else ("*" if s05 else "ns"),
                      "**" if s01 else "ns",row["rho_1"]]
            for ci,v in enumerate(vals,1):
                if isinstance(v,float) and np.isnan(v): v="—"
                elif isinstance(v,float): v=round(v,4)
                cell=xsc(ws,ri,ci,v,bg=bg,border=tb(),sz=9,
                         align="left" if ci<=3 else "right")
                if ci==len(vals)-2 and "Increasing" in str(v):
                    cell.font=Font(bold=True,color="1B5E20",name="Calibri",size=9)
                if ci==len(vals)-2 and "Decreasing" in str(v):
                    cell.font=Font(bold=True,color="B71C1C",name="Calibri",size=9)
            rh(ws,ri,15); ri+=1
        for ci,w in enumerate([10,8,18]+([7]*2)+([11]*(nc-5))+[14,9,9],1):
            cw(ws,ci,w)

    # S1: Standard MK
    ws1 = wb.create_sheet("S1 Standard MK")
    _write_trend_sheet(
        ws1, "Standard MK",
        f"Standard Mann–Kendall Test Results  |  {period}",
        "Mann (1945) / Kendall (1975)  |  NO autocorrelation correction  |"
        "  * p<0.05  ** p<0.01  |  Two-tailed test")

    # S2: Modified MK
    ws2 = wb.create_sheet("S2 Modified MK (H&R98)")
    _write_trend_sheet(
        ws2, "Modified MK",
        f"Modified Mann–Kendall Test Results (Hamed & Rao 1998)  |  {period}",
        "Autocorrelation-corrected  |  n* = effective sample size  |"
        "  Var*(S) = Var(S)×(n/n*)  |  * p<0.05  ** p<0.01")

    # S3: MK vs MMK Comparison
    ws3 = wb.create_sheet("S3 MK vs MMK Comparison")
    ws3.sheet_view.showGridLines = False
    nc3 = 17
    _title(ws3, nc3,
           f"Comparison: Standard MK vs Modified MK (Hamed & Rao 1998)  |  {period}",
           "ΔZ = Z_MMK − Z_MK  (negative → autocorr. reduces |Z|)  |"
           "  Agree = both methods reach same trend conclusion  |"
           "  Red Agree=False → autocorrelation changes trend decision")
    hdr3=["Station","Code","Scale","ρ₁","Sig.AC",
          "MK Z","MK p","MK Trend","MK *",
          "MMK Z","MMK p","MMK Trend","MMK *",
          "ΔZ","Δp","Agree","Note"]
    _hdr(ws3,3,hdr3)
    ri3=4
    for _,row in comp_df.iterrows():
        sk=row["Scale"]
        agree=bool(row["Agree"])
        sig_ac=bool(row["Sig_AC"])
        bg=sc_bg.get(sk,XC["ann_h"])
        if not agree: bg="FFE0E0"
        elif sig_ac:  bg="FFFDE7"
        vals=[str(row["Station"]),str(row.get("Code","")),
              SCALE_META.get(sk,{}).get("label",sk),
              row["rho_1"],
              "Yes*" if sig_ac else "No",
              row["MK_Z"],row["MK_p"],str(row["MK_Trend"]),
              "**" if row.get("MK_sig05",False) else "ns",
              row["MMK_Z"],row["MMK_p"],str(row["MMK_Trend"]),
              "**" if row.get("MMK_sig05",False) else "ns",
              row["delta_Z"],row["delta_p"],
              "Yes" if agree else "No",
              ("AC changed conclusion" if not agree else
               ("AC corrected" if sig_ac else ""))]
        for ci,v in enumerate(vals,1):
            if isinstance(v,float) and np.isnan(v): v="—"
            elif isinstance(v,float): v=round(v,4)
            bg_cell=bg
            cell=xsc(ws3,ri3,ci,v,bg=bg_cell,border=tb(),sz=9,
                     align="left" if ci in(1,2,3,8,12,17) else "right")
            if ci==16 and v=="No":
                cell.fill=xfill("FFCCBC")
                cell.font=Font(bold=True,color="B71C1C",name="Calibri",size=9)
            if ci==5 and "Yes" in str(v):
                cell.font=Font(bold=True,color="E65100",name="Calibri",size=9)
        rh(ws3,ri3,15); ri3+=1
    for ci,w in enumerate([10,8,18,8,7]+[10,10,18,5]*2+[9,9,5,30],1): cw(ws3,ci,w)

    # S4: Sen's Slope
    ws4 = wb.create_sheet("S4 Sens Slope")
    ws4.sheet_view.showGridLines = False
    _title(ws4, 11,
           f"Sen's Slope Estimator + 95% CI  |  {period}",
           "Sen (1968) JASA 63:1379  |  95% CI: Gilbert (1987) rank-based  |"
           "  β = Sen's slope (mm yr⁻¹)  |  Positive β = increasing rainfall")
    hdr4=["Station","Code","Scale","Method","N","β (mm/yr)",
          "CI_Lower (mm/yr)","CI_Upper (mm/yr)","Z","p-value","Trend"]
    _hdr(ws4,3,hdr4)
    ri4=4
    for _,row in trend_df.iterrows():
        sk=row["Scale"]; s05=bool(row["sig_05"]); s01=bool(row["sig_01"])
        bg=sc_bg.get(sk,XC["ann_h"])
        if s01: bg=XC["sig01"]
        elif s05: bg=XC["sig05"]
        vals=[str(row["Station"]),str(row["Code"]),row["Scale_Label"],
              row["Method"],row["N"],row["Slope_Q"],
              row["Slope_lo"],row["Slope_hi"],
              row["Z"],row["p_value"],str(row["Trend"])]
        for ci,v in enumerate(vals,1):
            if isinstance(v,float) and np.isnan(v): v="—"
            elif isinstance(v,float): v=round(v,3)
            cell=xsc(ws4,ri4,ci,v,bg=bg,border=tb(),sz=9,
                     align="left" if ci<=4 else "right")
            if ci==6 and isinstance(v,float):
                fc_v="1B5E20" if v>0 else "B71C1C"
                cell.font=Font(bold=s05,color=fc_v,name="Calibri",size=9)
        rh(ws4,ri4,15); ri4+=1
    for ci,w in enumerate([10,8,18,16,7,12,14,14,10,10,18],1): cw(ws4,ci,w)

    # S5: Descriptive Statistics
    ws5 = wb.create_sheet("S5 Descriptive Statistics")
    ws5.sheet_view.showGridLines = False
    _title(ws5, 11,
           f"Descriptive Statistics of Annual Rainfall  |  {period}",
           f"Wet-day threshold: ≥{WET_THR} mm/day (WMO)  |  "
           "CV = Coefficient of Variation  |  "
           "Skewness/Kurtosis: Fisher-Pearson")
    hdr5=["Station","Code","N (yr)","Mean (mm)","Median (mm)",
          "Max (mm)","Min (mm)","Std (mm)","CV (%)","Wet-days/yr","Skewness"]
    _hdr(ws5,3,hdr5)
    ri5=4
    alt=[xfill("E8F4FD"), xfill("FFFFFF")]
    for ni,stn in enumerate(stns,1):
        bg_fill = alt[ni%2]
        if stn not in desc_df.index:
            cell=xsc(ws5,ri5,1,stn,border=tb(),sz=9)
            cell.fill=bg_fill; rh(ws5,ri5,15); ri5+=1; continue
        d=desc_df.loc[stn]
        vals=[stn,smap.get(stn,stn),d["N (yr)"],d["Mean (mm)"],d["Median (mm)"],
              d["Max (mm)"],d["Min (mm)"],d["Std (mm)"],d["CV (%)"],
              d["Wet-days/yr"],d["Skewness"]]
        for ci,v in enumerate(vals,1):
            if isinstance(v,float) and np.isnan(v): v="—"
            elif isinstance(v,float): v=round(v,1)
            cell=xsc(ws5,ri5,ci,v,border=tb(),sz=9,
                     align="left" if ci<=2 else "right")
            cell.fill=bg_fill
        rh(ws5,ri5,16); ri5+=1
    for ci,w in enumerate([10,8,8,12,12,12,12,10,8,10,10],1): cw(ws5,ci,w)

    # S6: Methods & References
    ws6 = wb.create_sheet("S6 Methods & References")
    ws6.sheet_view.showGridLines = False
    mxsc(ws6,1,1,3,"Statistical Methods & References — Rainfall Trend Analysis v2.0",
         bold=True,fc="FFFFFF",bg=XC["title"],sz=13)
    rh(ws6,1,26)
    refs=[
        ("Standard Mann–Kendall",
         "Mann (1945); Kendall (1975)",
         "Non-parametric trend test. S = Σ Σ sgn(xⱼ−xᵢ). "
         "Var(S) with tie correction. Z=(S±1)/√Var(S). p two-tailed. "
         "Does NOT account for serial autocorrelation — may overestimate significance."),
        ("Modified Mann–Kendall",
         "Hamed & Rao (1998) J. Hydrol. 204:182–196",
         "Corrects Var(S) using ranked-series autocorrelations. "
         "n* = n / [1 + 2Σ(1−k/n)ρ_k(ranks)]. Var*(S)=Var(S)×(n/n*). "
         "Only significant ρ_k retained. Recommended when serial autocorrelation present."),
        ("Sen's Slope + 95% CI",
         "Sen (1968) JASA 63:1379; Gilbert (1987)",
         "Q = median[(xⱼ−xᵢ)/(j−i)] ∀j>i. Magnitude of trend. "
         "95% CI: Cα=z₀.₀₂₅×√Var(S); lo=(N−Cα)/2, hi=(N+Cα)/2+1. "
         "Non-parametric; robust to non-normality and outliers."),
        ("Lag-1 Autocorrelation",
         "Pearson; Box & Jenkins (1976)",
         "r₁ = Σ(xᵢ−x̄)(xᵢ₊₁−x̄)/Σ(xᵢ−x̄)². "
         "Significance: |r₁|>z₀.₀₂₅/√n. "
         "Significant r₁ → use Modified MK (Önöz & Bayazit 2003)."),
        ("Hydrological Seasons",
         "Thai hydro-climatological standard",
         "Wet: May–Oct (monsoon onset to withdrawal). "
         "Dry: Nov–Apr (Nov-Dec of year Y + Jan-Apr of year Y+1). "
         "Wet-day threshold: ≥1.0 mm/day (WMO 2008)."),
        ("Data Quality Control",
         "Tukey (1977); WMO (2008)",
         "Missing data: linear interpolation ≤5 consecutive days. "
         "Outlier detection: upper fence = Q3 + 3×IQR (extreme outlier threshold). "
         "Flagged values retained for transparency."),
        ("Season Definition",
         "Thai Meteorological Department; RID",
         "Wet: May 1 – October 31  (6 months, 184 days). "
         "Dry: November 1 – April 30  (6 months, 181/182 days). "
         "Annual: Calendar year January–December."),
        ("All References",
         "",
         "Mann HB (1945) Econometrica 13:245–259.\n"
         "Kendall MG (1975) Rank Correlation Methods. Griffin, London.\n"
         "Sen PK (1968) JASA 63:1379–1389.\n"
         "Hamed KH, Rao AR (1998) J. Hydrol. 204:182–196.\n"
         "Gilbert RO (1987) Statistical Methods for Environmental Pollution. Van Nostrand.\n"
         "Önöz B, Bayazit M (2003) Hydrol. Sci. J. 48:25–34.\n"
         "Yue S, Wang C (2004) Water Resour. Res. 40:W08307.\n"
         "Box GEP, Jenkins GM (1976) Time Series Analysis. Holden-Day.\n"
         "WMO (2008) Guide to Hydrological Practices. WMO-No. 168."),
    ]
    alt2=[PatternFill("solid",fgColor="DEEAF1"),PatternFill("solid",fgColor="FFFFFF")]
    for ri,(met,ref,desc) in enumerate(refs,3):
        fl=alt2[ri%2]
        for ci,v in enumerate([met,ref,desc],1):
            cell=xsc(ws6,ri,ci,v,bold=(ci<=2),sz=9.5,align="left",border=tb())
            cell.fill=fl
            if ci==3:
                cell.alignment=Alignment(horizontal="left",vertical="top",
                                          wrap_text=True)
        rh(ws6,ri,68)
    for ci,w in enumerate([26,42,68],1): cw(ws6,ci,w)

    wb.save(str(out_xlsx))
    print(f"  ✓  Excel: {Path(out_xlsx).name}  (6 sheets)")


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §17 RESEARCH SUMMARY DOCUMENT (Markdown)                               ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def write_summary_md(out_md: Path, stns, smap, trend_df, comp_df,
                     desc_df, period, n_sig_mk, n_sig_mmk, n_total,
                     any_sig_ac):
    """
    Write a comprehensive Markdown research summary ready for paper writing.
    Includes: Study area, methods, key results, tables, statistical summary.
    """
    stns  = [str(s) for s in stns]
    codes = [smap.get(s,s) for s in stns]
    n_s   = len(stns)
    now   = datetime.now().strftime("%Y-%m-%d")

    def fmt_row(stn, sk, method):
        sub=trend_df[(trend_df["Station"]==stn) &
                     (trend_df["Scale"]==sk) &
                     (trend_df["Method"]==method)]
        if len(sub)==0: return "—","—","—","—"
        r=sub.iloc[0]
        sig="**" if r["sig_01"] else ("*" if r["sig_05"] else "ns")
        sl=f"{r['Slope_Q']:+.2f}" if not np.isnan(r["Slope_Q"]) else "—"
        Z=f"{r['Z']:.3f}"
        p=f"{r['p_value']:.4f}"
        return Z,p,sig,sl

    lines = []
    lines += [
        f"# Rainfall Trend Analysis — Research Summary",
        f"",
        f"> **Generated**: {now}  |  "
        f"**Study Period**: {period}  |  "
        f"**Script**: Rainfall Trend Analysis v{VERSION}",
        f"",
        "---",
        "",
        "## 1. Study Area and Data",
        "",
        f"- **Study area**: Phetchaburi–Prachuap Khiri Khan River Basin, Western Thailand",
        f"- **Data**: Daily observed rainfall from {n_s} meteorological stations",
        f"- **Period**: {period}",
        f"- **Stations**: {', '.join([f'{c} ({s})' for c,s in zip(codes,stns)])}",
        f"- **Wet-day threshold**: ≥{WET_THR} mm day⁻¹ (WMO standard)",
        "",
        "## 2. Methods",
        "",
        "### 2.1 Temporal Scales (Hydrological Year)",
        "",
        "| Scale | Period | Description |",
        "|-------|--------|-------------|",
        "| Annual | Jan–Dec | Calendar year total |",
        "| Wet Season | May–Oct | Monsoon / wet season (6 months) |",
        "| Dry Season | Nov–Apr | Dry season — hydrological year approach (6 months) |",
        "",
        "### 2.2 Statistical Methods",
        "",
        "**Standard Mann–Kendall Test** (Mann 1945; Kendall 1975):",
        "- Non-parametric trend test for monotonic trends in time series.",
        "- S statistic with tie correction; Z-statistic from standard normal.",
        "- *Limitation*: Does not account for serial autocorrelation.",
        "",
        "**Modified Mann–Kendall Test** (Hamed & Rao 1998):",
        "- Corrects Var(S) using autocorrelation of the ranked series.",
        "- Effective sample size: $n^* = n / [1 + (2/n) \\sum_{k=1}^{n-1}(n-k)\\rho_k]$",
        "- Adjusted variance: $\\text{Var}^*(S) = \\text{Var}(S) \\times (n/n^*)$",
        f"- **Autocorrelation detected**: {'Yes → Modified MK essential' if any_sig_ac else 'No → both methods appropriate'}",
        "",
        "**Sen's Slope Estimator** (Sen 1968):",
        "- $Q = \\text{median}\\left[\\frac{x_j - x_i}{j - i}\\right]$ for all $j > i$",
        "- 95% CI: rank-based method (Gilbert 1987)",
        "- Interpretation: magnitude of change in mm per year",
        "",
        "**Significance levels**: α = 0.05 (|Z| > 1.96) and α = 0.01 (|Z| > 2.58)",
        "",
        "## 3. Results",
        "",
        "### 3.1 Descriptive Statistics",
        "",
        "| Station | Code | Mean (mm) | Std (mm) | CV (%) | Wet-days/yr |",
        "|---------|------|-----------|----------|--------|-------------|",
    ]
    for stn in stns:
        if stn not in desc_df.index: continue
        d=desc_df.loc[stn]
        lines.append(f"| {stn} | {smap.get(stn,stn)} | "
                     f"{d['Mean (mm)']:.1f} | {d['Std (mm)']:.1f} | "
                     f"{d['CV (%)']:.1f} | {d['Wet-days/yr']:.1f} |")

    lines += [
        "",
        f"*Regional mean annual rainfall: {desc_df['Mean (mm)'].mean():.1f} mm/yr "
        f"(range: {desc_df['Mean (mm)'].min():.1f}–{desc_df['Mean (mm)'].max():.1f} mm/yr)*",
        "",
        "### 3.2 Autocorrelation Results",
        "",
        "| Station | Code | r₁ (Annual) | Significant? | → Modified MK? |",
        "|---------|------|-------------|--------------|----------------|",
    ]
    for stn in stns:
        arr=scales_global.get("annual",{})
        if hasattr(arr,"__contains__") and stn in arr.columns if hasattr(arr,"columns") else False:
            a=arr[stn].dropna().values.astype(float)
        else:
            a=np.array([])
        r1=lag_k_autocorr(a) if len(a)>4 else np.nan
        sig=is_sig_autocorr(r1,len(a)) if not np.isnan(r1) else False
        r1_str = f"{r1:.4f}" if not np.isnan(r1) else "—"
        lines.append(f"| {stn} | {smap.get(stn,stn)} | "
                     f"{r1_str} | "
                     f"{'Yes ***' if sig else 'No'} | "
                     f"{'Recommended' if sig else 'Optional'} |")

    lines += [
        "",
        "### 3.3 Trend Analysis — Annual Scale",
        "",
        "| Station | Code | MK Z | MK p | MMK Z | MMK p | β (mm/yr) | Trend | Sig. |",
        "|---------|------|------|------|-------|-------|-----------|-------|------|",
    ]
    for stn in stns:
        mk_Z,mk_p,mk_sig,_ = fmt_row(stn,"annual","Standard MK")
        mmk_Z,mmk_p,mmk_sig,sl = fmt_row(stn,"annual","Modified MK")
        code=smap.get(stn,stn)
        sub=trend_df[(trend_df["Station"]==stn) &
                     (trend_df["Scale"]=="annual") &
                     (trend_df["Method"]=="Modified MK")]
        tr=str(sub["Trend"].values[0]) if len(sub) else "—"
        lines.append(f"| {stn} | {code} | {mk_Z} | {mk_p} | "
                     f"{mmk_Z} | {mmk_p} | {sl} | {tr} | {mmk_sig} |")

    lines += [
        "",
        "### 3.4 Trend Analysis — Wet Season (May–Oct)",
        "",
        "| Station | Code | MK Z | MMK Z | MMK p | β (mm/yr) | Trend | Sig. |",
        "|---------|------|------|-------|-------|-----------|-------|------|",
    ]
    for stn in stns:
        mk_Z,_,_,_ = fmt_row(stn,"wet","Standard MK")
        mmk_Z,mmk_p,mmk_sig,sl = fmt_row(stn,"wet","Modified MK")
        code=smap.get(stn,stn)
        sub=trend_df[(trend_df["Station"]==stn) &
                     (trend_df["Scale"]=="wet") &
                     (trend_df["Method"]=="Modified MK")]
        tr=str(sub["Trend"].values[0]) if len(sub) else "—"
        lines.append(f"| {stn} | {code} | {mk_Z} | {mmk_Z} | {mmk_p} | "
                     f"{sl} | {tr} | {mmk_sig} |")

    lines += [
        "",
        "### 3.5 Trend Analysis — Dry Season (Nov–Apr)",
        "",
        "| Station | Code | MK Z | MMK Z | MMK p | β (mm/yr) | Trend | Sig. |",
        "|---------|------|------|-------|-------|-----------|-------|------|",
    ]
    for stn in stns:
        mk_Z,_,_,_ = fmt_row(stn,"dry","Standard MK")
        mmk_Z,mmk_p,mmk_sig,sl = fmt_row(stn,"dry","Modified MK")
        code=smap.get(stn,stn)
        sub=trend_df[(trend_df["Station"]==stn) &
                     (trend_df["Scale"]=="dry") &
                     (trend_df["Method"]=="Modified MK")]
        tr=str(sub["Trend"].values[0]) if len(sub) else "—"
        lines.append(f"| {stn} | {code} | {mk_Z} | {mmk_Z} | {mmk_p} | "
                     f"{sl} | {tr} | {mmk_sig} |")

    # Agreement summary
    n_agree = int(comp_df["Agree"].sum())
    n_total_comp = len(comp_df)
    n_changed = n_total_comp - n_agree
    lines += [
        "",
        "### 3.6 MK vs Modified MK Comparison",
        "",
        f"| Metric | Value |",
        f"|--------|-------|",
        f"| Total comparisons (station × scale) | {n_total_comp} |",
        f"| Agreement (same trend conclusion) | {n_agree} ({100*n_agree/n_total_comp:.1f}%) |",
        f"| Changed by autocorrelation correction | {n_changed} ({100*n_changed/n_total_comp:.1f}%) |",
        f"| Stations with significant autocorr. (annual) | "
        f"{sum(1 for _,r in comp_df[comp_df['Scale']=='annual'].iterrows() if r['Sig_AC'])} / {n_s} |",
        f"| Sig. trends (Standard MK, p<0.05) | {n_sig_mk} / {n_total} |",
        f"| Sig. trends (Modified MK, p<0.05) | {n_sig_mmk} / {n_total} |",
        "",
        "### 3.7 Key Findings",
        "",
    ]

    # Auto-generate key findings from data
    inc_ann = trend_df[(trend_df["Scale"]=="annual") &
                       (trend_df["Method"]=="Modified MK") &
                       (trend_df["sig_05"]==True) &
                       (trend_df["Z"]>0)]
    dec_ann = trend_df[(trend_df["Scale"]=="annual") &
                       (trend_df["Method"]=="Modified MK") &
                       (trend_df["sig_05"]==True) &
                       (trend_df["Z"]<0)]
    inc_wet = trend_df[(trend_df["Scale"]=="wet") &
                       (trend_df["Method"]=="Modified MK") &
                       (trend_df["sig_05"]==True) &
                       (trend_df["Z"]>0)]
    dec_dry = trend_df[(trend_df["Scale"]=="dry") &
                       (trend_df["Method"]=="Modified MK") &
                       (trend_df["sig_05"]==True) &
                       (trend_df["Z"]<0)]

    if len(inc_ann)>0:
        sl_mean=float(inc_ann["Slope_Q"].mean())
        stns_inc=[smap.get(str(s),str(s)) for s in inc_ann["Station"]]
        lines.append(f"- **Annual increasing trend**: {', '.join(stns_inc)} show "
                     f"significant increasing trends (mean β = {sl_mean:+.2f} mm/yr, p<0.05).")
    if len(dec_ann)>0:
        sl_mean=float(dec_ann["Slope_Q"].mean())
        stns_dec=[smap.get(str(s),str(s)) for s in dec_ann["Station"]]
        lines.append(f"- **Annual decreasing trend**: {', '.join(stns_dec)} show "
                     f"significant decreasing trends (mean β = {sl_mean:+.2f} mm/yr, p<0.05).")
    if len(inc_ann)==0 and len(dec_ann)==0:
        lines.append("- **Annual**: No statistically significant trends detected "
                     "at p<0.05 level in annual rainfall.")
    if len(inc_wet)>0:
        stns_iw=[smap.get(str(s),str(s)) for s in inc_wet["Station"]]
        lines.append(f"- **Wet season**: {', '.join(stns_iw)} show increasing trends.")
    else:
        lines.append("- **Wet season**: No significant trends detected.")
    if len(dec_dry)>0:
        stns_dd=[smap.get(str(s),str(s)) for s in dec_dry["Station"]]
        lines.append(f"- **Dry season**: {', '.join(stns_dd)} show decreasing trends.")
    else:
        lines.append("- **Dry season**: No significant trends detected.")
    if any_sig_ac:
        lines.append(f"- **Autocorrelation effect**: Serial autocorrelation was significant in "
                     f"several stations. Modified MK corrects for this bias; "
                     f"{n_changed} trend conclusions changed after applying the correction.")
    else:
        lines.append("- **Autocorrelation**: No significant autocorrelation detected; "
                     "Standard MK and Modified MK results are highly consistent.")

    lines += [
        "",
        "## 4. Discussion Points",
        "",
        "- The Modified Mann–Kendall test (Hamed & Rao 1998) is the recommended approach "
        "when serial autocorrelation is present in hydro-climatic time series data.",
        "- Positive serial autocorrelation inflates the Standard MK Z-statistic, "
        "leading to false positive trend detection (Type I error inflation).",
        "- Sen's slope provides a physically meaningful estimate of the rate of change, "
        "which is essential for water resource planning.",
        "- Wet/dry season separation is hydrologically important: "
        "changes in wet season rainfall affect flood risk, "
        "while dry season trends affect irrigation demand and reservoir management.",
        "- The 95% CI of Sen's slope should be reported alongside trend significance "
        "to convey the uncertainty in the magnitude of change.",
        "",
        "## 5. Suggested Paper Language",
        "",
        "### Methods Section (Draft)",
        "",
        "Long-term trends in daily, annual, and seasonal rainfall were analysed using "
        "the Modified Mann–Kendall (MMK) trend test proposed by Hamed and Rao (1998), "
        "which accounts for the effect of positive serial autocorrelation commonly found "
        "in hydro-climatic time series. The standard Mann–Kendall test (Mann 1945; "
        "Kendall 1975) was also applied for comparison. The magnitude of detected trends "
        "was quantified using the non-parametric Sen's slope estimator (Sen 1968), "
        "together with its 95% confidence interval derived from the rank-based method "
        "of Gilbert (1987). All analyses were conducted separately for the annual "
        f"({period}) and two hydrological seasons: the wet season "
        "(May–October) and the dry season (November–April). Significance was assessed "
        "at the 5% (α = 0.05) and 1% (α = 0.01) levels.",
        "",
        "### Results Section (Template)",
        "",
        "Of the {N} station–scale combinations tested, {n_sig_mmk} showed statistically "
        "significant trends (p < 0.05) according to the Modified MK test. "
        "The serial autocorrelation analysis indicated that {n_ac} stations exhibited "
        "significant Lag-1 autocorrelation at the annual scale, justifying the use of "
        "the Modified MK correction. Agreement between Standard MK and Modified MK was "
        "high ({n_agree}/{n_total_comp} combinations, {pct:.1f}%), indicating that "
        "autocorrelation had a limited but non-negligible effect on trend conclusions.",
        "",
        "## 6. References",
        "",
        "- Mann, H. B. (1945). Nonparametric tests against trend. *Econometrica*, 13, 245–259.",
        "- Kendall, M. G. (1975). *Rank Correlation Methods* (4th ed.). Griffin, London.",
        "- Sen, P. K. (1968). Estimates of regression coefficient based on Kendall's tau. "
        "*Journal of the American Statistical Association*, 63, 1379–1389.",
        "- Hamed, K. H., & Rao, A. R. (1998). A modified Mann–Kendall trend test for "
        "autocorrelated data. *Journal of Hydrology*, 204, 182–196.",
        "- Gilbert, R. O. (1987). *Statistical Methods for Environmental Pollution "
        "Monitoring*. Van Nostrand Reinhold, New York.",
        "- Önöz, B., & Bayazit, M. (2003). The power of statistical tests for trend "
        "detection. *Hydrological Sciences Journal*, 48, 93–98.",
        "- Yue, S., & Wang, C. (2004). The Mann–Kendall test modified by effective "
        "sample size to detect trend in serially correlated hydrological series. "
        "*Water Resources Research*, 40, W08307.",
        "- WMO (2008). *Guide to Hydrological Practices* (WMO-No. 168). "
        "World Meteorological Organization, Geneva.",
        "",
        "---",
        f"*End of Research Summary  |  Generated: {now}  |  Script v{VERSION}*",
    ]

    # Fill in template placeholders
    n_ac_sig = sum(1 for _,r in comp_df[comp_df["Scale"]=="annual"].iterrows()
                   if r["Sig_AC"])
    pct_agree = 100*n_agree/n_total_comp if n_total_comp>0 else 0
    text = "\n".join(lines)
    text = (text
            .replace("{N}", str(n_total))
            .replace("{n_sig_mmk}", str(n_sig_mmk))
            .replace("{n_ac}", str(n_ac_sig))
            .replace("{n_agree}", str(n_agree))
            .replace("{n_total_comp}", str(n_total_comp))
            .replace("{pct:.1f}", f"{pct_agree:.1f}"))

    out_md.write_text(text, encoding="utf-8")
    print(f"  ✓  Summary: {out_md.name}")


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  §18 MAIN                                                               ║
# ╚══════════════════════════════════════════════════════════════════════════╝

# Global for summary doc (needed inside write_summary_md)
scales_global = {}


def short_labels(stns):
    return {str(s): f"S{i+1}" for i, s in enumerate(stns)}


def main():
    SEP = "═" * 72
    _ext = " + PW-MK + TFPW-MK" if _RTA_AVAILABLE else ""
    print(SEP)
    print(f"  Rainfall Trend Analysis  v{VERSION}  — Publication Edition")
    print(f"  Standard MK + Modified MK (H&R98){_ext} + Sen's Slope")
    print("  Hydrological Year: Wet (May–Oct) | Dry (Nov–Apr)")
    print("  Output: 8 Figures + Excel + Research Summary Markdown")
    if _RTA_AVAILABLE:
        print("  Extensions: Field Significance | Checkpoint/Resume | Spatial Support")
    print(SEP)

    # ── CLI arguments: [folder] [--no-resume] [--no-pdf] ────────────────
    args      = sys.argv[1:]
    work_dir  = None
    no_resume = False
    for a in args:
        if a == "--no-resume":
            no_resume = True
        elif a == "--no-pdf":
            import rta.config as _cfg; _cfg.SAVE_PDF = False
        elif not a.startswith("--"):
            work_dir = a.strip('"').strip("'")
    if work_dir is None:
        try:
            work_dir = str(Path(os.path.abspath(__file__)).parent)
        except Exception:
            work_dir = os.getcwd()

    out_dir  = Path(work_dir)
    cp_dir   = out_dir / "checkpoints"
    if _RTA_AVAILABLE:
        cp_dir.mkdir(exist_ok=True)

    csv_path = find_csv(work_dir)
    base     = Path(csv_path).stem
    prefix   = f"Output_TrendV2_{base}"

    print(f"  Input : {csv_path}")
    print(f"  Output: {work_dir}\n")

    # ── Checkpoint: prompt resume ────────────────────────────────────────
    resume_from = 0
    if _RTA_AVAILABLE:
        resume_from = _ckpt_prompt(cp_dir, no_resume=no_resume)

    # ════════════════════════════════════════════════════════════════════
    # Step 1: Load + QC
    # ════════════════════════════════════════════════════════════════════
    if _RTA_AVAILABLE and resume_from >= 1:
        print("  Step 1: Loading QC checkpoint ...")
        _ck = _ckpt_load("01_qc", cp_dir)
        df, qc_dict = _ck["df"], _ck["qc_dict"]
        stns_str, smap, period = _ck["stns_str"], _ck["smap"], _ck["period"]
    else:
        print("  Step 1: Loading data and Quality Control ...")
        df_raw     = load_daily(csv_path)
        df, qc_dict= quality_control(df_raw.copy())
        stns_str   = [str(s) for s in df.columns.tolist()]
        smap       = short_labels(stns_str)
        period     = f"{df.index[0].year}–{df.index[-1].year}"
        if _RTA_AVAILABLE:
            _ckpt_save("01_qc",
                       {"df": df, "qc_dict": qc_dict,
                        "stns_str": stns_str, "smap": smap, "period": period},
                       cp_dir)

    global scales_global
    print(f"  Stations: {len(stns_str)} | Period: {period} | "
          f"Records: {len(df):,}")
    for s, q in qc_dict.items():
        print(f"    {smap[s]:5s} [{s}]  "
              f"missing={q['n_missing']}d ({q['pct_miss']}%)  "
              f"outliers={q['n_outlier']}  filled={q['n_filled']}d")

    # ════════════════════════════════════════════════════════════════════
    # Step 2: Temporal Aggregation
    # ════════════════════════════════════════════════════════════════════
    if _RTA_AVAILABLE and resume_from >= 2:
        print("\n  Step 2: Loading aggregation checkpoint ...")
        _ck2 = _ckpt_load("02_aggregation", cp_dir)
        scales, desc_df = _ck2["scales"], _ck2["desc_df"]
    else:
        print("\n  Step 2: Temporal aggregation ...")
        scales = aggregate_all(df)
        desc_df = descriptive_stats(scales, df)
        if _RTA_AVAILABLE:
            _ckpt_save("02_aggregation",
                       {"scales": scales, "desc_df": desc_df}, cp_dir)

    scales_global = scales
    for sk in ["annual", "wet", "dry"]:
        df_s = scales[sk]
        print(f"    {SCALE_META[sk]['label']:22s}: "
              f"{len(df_s)} years × {df_s.shape[1]} stations")

    # ════════════════════════════════════════════════════════════════════
    # Step 3: Descriptive Statistics
    # ════════════════════════════════════════════════════════════════════
    print("\n  Step 3: Descriptive statistics ...")
    print(desc_df[["Mean (mm)", "CV (%)", "Wet-days/yr", "Skewness"]].to_string())

    # ════════════════════════════════════════════════════════════════════
    # Step 4: Autocorrelation
    # ════════════════════════════════════════════════════════════════════
    print("\n  Step 4: Lag-1 Autocorrelation (annual scale) ...")
    any_sig_ac = False
    print(f"  {'Code':6s} [{' Station ':8s}]  "
          f"{'r₁':>8s}  {'Sig.':>6s}  Use Modified MK?")
    for stn in stns_str:
        arr = (scales["annual"][stn].dropna().values.astype(float)
               if stn in scales["annual"].columns else np.array([]))
        r1  = lag_k_autocorr(arr)
        sig = is_sig_autocorr(r1, len(arr))
        if sig: any_sig_ac = True
        print(f"  {smap[stn]:6s} [{stn:8s}]  "
              f"{r1:8.4f}  {'Yes ***' if sig else 'No':>6s}  "
              f"{'→ Essential' if sig else '→ Optional'}")
    print(f"\n  → {'Modified MK applied' if any_sig_ac else 'No significant AC'}")

    # ════════════════════════════════════════════════════════════════════
    # Step 5-6: Trend tests (MK + MMK + PW-MK + TFPW-MK if available)
    # ════════════════════════════════════════════════════════════════════
    if _RTA_AVAILABLE and resume_from >= 4:
        print("\n  Steps 5–6: Loading trend checkpoint ...")
        _ck4 = _ckpt_load("04_trends", cp_dir)
        trend_df = _ck4["trend_df"]
    else:
        _method_label = ("Standard MK + Modified MK + PW-MK + TFPW-MK"
                         if _RTA_AVAILABLE else "Standard MK + Modified MK")
        print(f"\n  Steps 5–6: {_method_label} ...")
        trend_df = run_all(scales, stns_str, smap)
        if _RTA_AVAILABLE:
            _ckpt_save("04_trends", {"trend_df": trend_df}, cp_dir)

    # Print summary table (show MK and MMK; PW/TFPW printed if available)
    _print_methods = ["Standard MK", "Modified MK"]
    if _RTA_AVAILABLE:
        _print_methods += ["PW-MK", "TFPW-MK"]
    print(f"\n  {'Code':6s} {'Scale':12s} {'Method':15s} "
          f"{'Z':>7s} {'p':>7s} {'β mm/yr':>8s}  Trend")
    print("  " + "-" * 68)
    for sk in ["annual", "wet", "dry"]:
        for meth in _print_methods:
            sub = trend_df[(trend_df["Scale"] == sk) &
                           (trend_df["Method"] == meth)]
            for _, row in sub.iterrows():
                slab = _sig_label(row["sig_05"], row["sig_01"], row["Z"])
                beta = (f"{row['Slope_Q']:+.2f}"
                        if not np.isnan(row["Slope_Q"]) else "—")
                print(f"  {row['Code']:6s} {sk:12s} {meth[:14]:15s} "
                      f"{row['Z']:7.3f} {row['p_value']:7.4f} "
                      f"{beta:>8s}  {row['Trend']} {slab}")
        print()

    n_mk_rows = int((trend_df["Method"] == "Standard MK").sum())
    n_total   = n_mk_rows  # used for reporting fractions
    n_sig_mk  = int(trend_df[trend_df["Method"] == "Standard MK"]["sig_05"].sum())
    n_sig_mmk = int(trend_df[trend_df["Method"] == "Modified MK"]["sig_05"].sum())
    print(f"  Significant (p<0.05): "
          f"Standard MK={n_sig_mk}/{n_mk_rows}  "
          f"Modified MK={n_sig_mmk}/{n_mk_rows}")
    if _RTA_AVAILABLE:
        for meth in ["PW-MK", "TFPW-MK"]:
            n_s = int(trend_df[trend_df["Method"] == meth]["sig_05"].sum())
            print(f"  Significant (p<0.05): {meth}={n_s}/{n_mk_rows}")

    # ════════════════════════════════════════════════════════════════════
    # Step 7: Comparison tables
    # ════════════════════════════════════════════════════════════════════
    if _RTA_AVAILABLE and resume_from >= 5:
        print("\n  Step 7: Loading comparison checkpoint ...")
        _ck5 = _ckpt_load("05_comparison", cp_dir)
        comp_df  = _ck5["comp_df"]
        comp4_df = _ck5.get("comp4_df")
    else:
        print("\n  Step 7: Building comparison tables ...")
        comp_df  = build_comparison(trend_df)
        comp4_df = build_4method_comparison(trend_df) if _RTA_AVAILABLE else None
        if _RTA_AVAILABLE:
            _ckpt_save("05_comparison",
                       {"comp_df": comp_df, "comp4_df": comp4_df}, cp_dir)

    n_agree   = int(comp_df["Agree"].sum())
    n_changed = len(comp_df) - n_agree
    print(f"  MK vs MMK agreement: {n_agree}/{len(comp_df)} "
          f"({100*n_agree/len(comp_df):.1f}%)")
    if n_changed > 0:
        print(f"  ⚠  {n_changed} cases where AC correction changed conclusion:")
        for _, r in comp_df[~comp_df["Agree"]].iterrows():
            print(f"     {smap.get(r['Station'], r['Station'])} "
                  f"({r['Scale']}): MK={r['MK_Trend']}  MMK={r['MMK_Trend']}")
    if comp4_df is not None and len(comp4_df) > 0:
        n_all = int(comp4_df["all_agree"].sum()) if "all_agree" in comp4_df.columns else 0
        print(f"  4-method full agreement: {n_all}/{len(comp4_df)}")

    # ════════════════════════════════════════════════════════════════════
    # Step 7b: Field significance (rta extension)
    # ════════════════════════════════════════════════════════════════════
    field_sig_df = None
    if _RTA_AVAILABLE:
        if resume_from >= 6:
            print("\n  Step 7b: Loading field significance checkpoint ...")
            _ck6 = _ckpt_load("06_field_sig", cp_dir)
            field_sig_df = _ck6["field_sig_df"]
        else:
            print("\n  Step 7b: Field significance (Walker + Livezey-Chen MC) ...")
            field_sig_df = field_sig_summary(scales, stns_str,
                                             alpha=ALPHA_005, n_perm=1000)
            _ckpt_save("06_field_sig", {"field_sig_df": field_sig_df}, cp_dir)
        if field_sig_df is not None and len(field_sig_df) > 0:
            print(field_sig_df[["Scale", "N_sig_MK", "N_sig_MMK",
                                 "Walker_sig_MK", "LC_sig_MK"]].to_string(index=False))

    # ── Station coordinates (optional) ──────────────────────────────────
    coords = load_coords(work_dir) if _RTA_AVAILABLE else None
    if coords:
        print(f"\n  ✓ Station coordinates loaded ({len(coords)} stations)")
    else:
        print("\n  ℹ  No coordinates file found — spatial maps use index axis")

    # ════════════════════════════════════════════════════════════════════
    # Step 8: Original 8 figures (Output_TrendV2_ prefix — backward compat)
    # ════════════════════════════════════════════════════════════════════
    print(f"\n{'─'*72}")
    print("  Step 8: Generating publication figures (600 DPI) ...")

    print("\n  Figure 1: Annual Time Series ...")
    fig1_annual_ts(scales, trend_df, stns_str, smap, period, out_dir, prefix)

    print("\n  Figure 2: Wet & Dry Season Time Series ...")
    fig2_wetdry_ts(scales, trend_df, stns_str, smap, period, out_dir, prefix)

    print("\n  Figure 3: Sen's Slope All Scales ...")
    fig3_sens_all(trend_df, stns_str, smap, period, out_dir, prefix)

    print("\n  Figure 4: MK vs MMK Comparison ...")
    fig4_mk_vs_mmk(comp_df, stns_str, smap, period, out_dir, prefix)

    print("\n  Figure 5: Significance Heatmap ...")
    fig5_significance_heatmap(trend_df, stns_str, smap, period, out_dir, prefix)

    print("\n  Figure 6: Autocorrelation ...")
    fig6_autocorrelation(scales, stns_str, smap, period, out_dir, prefix)

    print("\n  Figure 7: Monthly Climatology ...")
    fig7_monthly_climatology(scales, stns_str, smap, period, out_dir, prefix)

    print("\n  Figure 8: Spatial Trend Summary ...")
    fig8_spatial_summary(trend_df, comp_df, stns_str, smap, period,
                         out_dir, prefix)

    # ════════════════════════════════════════════════════════════════════
    # Step 8b: New v4 figures (Output_TrendV2_ same prefix, additional figs)
    # ════════════════════════════════════════════════════════════════════
    if _RTA_AVAILABLE:
        try:
            from rta.figures.method_comparison import (
                fig10_z_comparison_matrix,
                fig11_method_comparison_scatter)
            from rta.figures.taylor      import fig9_taylor_diagram
            from rta.figures.acf_plots   import fig12_acf_diagnostics
            from rta.figures.field_sig_plot import fig13_field_significance
            from rta.figures.spatial_maps   import (fig14_spatial_maps,
                    fig_station_distribution, fig_spatial_methods,
                    fig_spatial_field_sig, fig_spatial_full)

            print("\n  Figure 9: Taylor Diagram ...")
            fig9_taylor_diagram(scales, stns_str, smap, period, out_dir, prefix)

            print("\n  Figure 10: Z-Comparison Matrix (4 methods) ...")
            fig10_z_comparison_matrix(trend_df, stns_str, smap, period,
                                      out_dir, prefix)

            print("\n  Figure 11: Method Comparison Scatter ...")
            fig11_method_comparison_scatter(trend_df, stns_str, smap, period,
                                            out_dir, prefix)

            print("\n  Figure 12: ACF Diagnostics ...")
            fig12_acf_diagnostics(scales, stns_str, smap, period, out_dir, prefix)

            if field_sig_df is not None and len(field_sig_df) > 0:
                print("\n  Figure 13: Field Significance ...")
                fig13_field_significance(field_sig_df, period, out_dir, prefix)

            print("\n  Figure 14: Spatial Trend Maps (geographic, MMK) ...")
            fig14_spatial_maps(trend_df, stns_str, smap, coords, period,
                               out_dir, prefix)

            if coords:
                print("\n  Figure 14b: Station Distribution Map ...")
                fig_station_distribution(coords, stns_str, smap, period,
                                         out_dir, prefix)
                print("\n  Figure 14c: All-Methods Spatial Maps ...")
                fig_spatial_methods(trend_df, stns_str, smap, coords, period,
                                    out_dir, prefix)
                print("\n  Figure 14d: Field Significance Spatial Map ...")
                fig_spatial_field_sig(trend_df, field_sig_df, stns_str,
                                       smap, coords, period, out_dir, prefix)
                print("\n  Figure 14e: Comprehensive Spatial Overview ...")
                fig_spatial_full(trend_df, stns_str, smap, coords,
                                  field_sig_df, period, out_dir, prefix)
        except Exception as _fig_err:
            print(f"  ⚠  New figures skipped: {_fig_err}")

    # ════════════════════════════════════════════════════════════════════
    # Step 9: Excel
    # ════════════════════════════════════════════════════════════════════
    print(f"\n{'─'*72}")
    out_xlsx = out_dir / f"Output_TrendV2_{base}_Results.xlsx"
    print(f"  Step 9: Building Excel → {out_xlsx.name} ...")
    write_excel(out_xlsx, stns_str, smap, trend_df, comp_df,
                desc_df, qc_dict, period)

    # ════════════════════════════════════════════════════════════════════
    # Step 10: Research Summary Markdown
    # ════════════════════════════════════════════════════════════════════
    print(f"\n  Step 10: Writing Research Summary ...")
    out_md = out_dir / f"Output_TrendV2_{base}_Research_Summary.md"
    write_summary_md(out_md, stns_str, smap, trend_df, comp_df,
                     desc_df, period, n_sig_mk, n_sig_mmk, n_total,
                     any_sig_ac)

    # ════════════════════════════════════════════════════════════════════
    # Final Summary
    # ════════════════════════════════════════════════════════════════════
    n_fig = len(list(out_dir.glob(f"{prefix}_Fig*.png")))
    n_sheets = 6
    print()
    print(SEP)
    print(f"  ✓  DONE — Rainfall Trend Analysis v{VERSION}")
    print(f"  {'─'*62}")
    print(f"  Period           : {period}")
    print(f"  Stations         : {len(stns_str)}  "
          f"({', '.join(smap.values())})")
    print(f"  Temporal scales  : Annual / Wet (May–Oct) / Dry (Nov–Apr)")
    _m_str = "MK + MMK + PW-MK + TFPW-MK" if _RTA_AVAILABLE else "MK + MMK"
    print(f"  Methods          : {_m_str} + Sen's slope")
    print(f"  Autocorr. (Lag-1): "
          f"{'Significant → Modified MK essential' if any_sig_ac else 'Not significant'}")
    print(f"  Sig. (p<0.05)    : "
          f"MK={n_sig_mk}/{n_mk_rows}  "
          f"MMK={n_sig_mmk}/{n_mk_rows}")
    print(f"  MK vs MMK agree  : "
          f"{n_agree}/{len(comp_df)} ({100*n_agree/len(comp_df):.1f}%)")
    if field_sig_df is not None and len(field_sig_df) > 0:
        row = field_sig_df[field_sig_df["Scale"] == "annual"]
        if len(row):
            r = row.iloc[0]
            print(f"  Field sig (ann)  : "
                  f"Walker={'Yes' if r['Walker_sig_MK'] else 'No'}  "
                  f"LC={'Yes' if r['LC_sig_MK'] else 'No'}")
    print(f"  Figures          : {n_fig} PNG" + (" + PDF" if SAVE_PDF else ""))
    print(f"  Excel ({n_sheets} sheets): {out_xlsx.name}")
    print(f"  Summary (MD)     : {out_md.name}")
    print(f"  Saved in         : {work_dir}")
    print(SEP)


if __name__ == "__main__":
    main()
